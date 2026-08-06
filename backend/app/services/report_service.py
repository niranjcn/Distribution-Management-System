from datetime import datetime, timedelta, timezone
from typing import Optional, List, Dict, Any
import asyncio
import csv
import io

from openpyxl import Workbook

from app.database_sqlalchemy import async_session_factory
from sqlalchemy import text, bindparam
from app.utils.roles import normalize_role


REPORT_MANAGEMENT_ROLES = {"super_admin", "md_director", "manager", "pdic_staff"}


async def _resolve_report_scope(current_user: Optional[dict]) -> Optional[dict]:
    """Resolve the hierarchy-report scope for a user.

    Returns None for management roles (full system report), or a dict like
    {"scope": "sub", "sub_id": int} / {"scope": "cluster", "cluster_id": int}
    / {"scope": "operator", "operator_id": int} for field roles.
    """
    if not current_user:
        return None
    role = normalize_role(current_user.get("role"))
    if role in REPORT_MANAGEMENT_ROLES:
        return None
    user_id = int(current_user["id"])
    if role == "sub_distributor":
        return {"scope": "sub", "sub_id": user_id}
    if role in ("sub_distribution_manager", "sub_distribution_employee"):
        async with async_session_factory() as session:
            row = (await session.execute(
                text("SELECT parent_id FROM users WHERE id = :uid"),
                {"uid": user_id},
            )).mappings().first()
        sub_id = int(row["parent_id"]) if row and row.get("parent_id") is not None else None
        return {"scope": "sub", "sub_id": sub_id}
    if role == "cluster":
        return {"scope": "cluster", "cluster_id": user_id}
    if role == "operator":
        return {"scope": "operator", "operator_id": user_id}
    return None


ALLOWED_REPORT_TABLES = {
    "devices",
    "distributions",
    "defects",
    "returns",
    "users",
}


async def _count(session, table: str, condition: str = "1=1", params: Optional[dict] = None) -> int:
    if table not in ALLOWED_REPORT_TABLES:
        raise ValueError(f"Invalid table name: {table}")
    result = await session.execute(text(f"SELECT COUNT(*) FROM {table} WHERE {condition}"), params or {})
    return result.scalar()


def _build_date_filter(base_condition: str, base_params: dict, start_date: Optional[str], end_date: Optional[str]) -> tuple:
    conds = [base_condition]
    params = dict(base_params)
    if start_date:
        conds.append("created_at >= :start_date")
        params["start_date"] = start_date
    if end_date:
        conds.append("created_at <= :end_date")
        params["end_date"] = end_date
    return " AND ".join(conds), params


async def get_inventory_report(start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
    """Generate device inventory report"""
    async with async_session_factory() as session:
        cond, prm = _build_date_filter("1=1", {}, start_date, end_date)

        # Single-pass aggregation: one query groups by status, device_type and
        # current_holder_type together. The result is split below into the three
        # summary buckets, so we scan the device table exactly once instead of
        # running four separate COUNT/GROUP BY queries. Rows with a NULL
        # current_holder_type still count toward by_status/by_type and the
        # total, but are excluded from by_location (matching the prior query
        # which filtered current_holder_type IS NOT NULL).
        result = await session.execute(
            text(f"SELECT status, device_type, current_holder_type, COUNT(*) as cnt FROM devices WHERE {cond} GROUP BY status, device_type, current_holder_type"),
            prm,
        )

        total = 0
        by_status = {}
        by_type = {}
        by_location = {}
        for row in result.fetchall():
            status = row[0]
            device_type = row[1]
            holder_type = row[2]
            cnt = int(row[3])
            total += cnt
            by_status[status] = by_status.get(status, 0) + cnt
            by_type[device_type] = by_type.get(device_type, 0) + cnt
            if holder_type is not None:
                by_location[holder_type] = by_location.get(holder_type, 0) + cnt

        return {
            "total_devices": total,
            "by_status": by_status,
            "by_type": by_type,
            "by_location": by_location,
            "generated_at": datetime.now().replace(tzinfo=None).isoformat()
        }


async def get_distribution_summary(start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
    """Generate distribution summary report"""
    async with async_session_factory() as session:
        cond, prm = _build_date_filter("1=1", {}, start_date, end_date)
        total = await _count(session, "distributions", cond, prm)

        by_status = {}
        result = await session.execute(text(f"SELECT status, COUNT(*) as cnt FROM distributions WHERE {cond} GROUP BY status"), prm)
        for row in result.fetchall():
            by_status[row[0]] = row[1]

        by_month = []
        now = datetime.now().replace(tzinfo=None)
        six_months_ago = (datetime(now.year, now.month, 1) - timedelta(days=180)).isoformat()
        mc, mp = _build_date_filter("created_at >= :six_months_ago", {"six_months_ago": six_months_ago}, start_date, end_date)
        result = await session.execute(
            text(f"SELECT SUBSTR(created_at, 1, 7) as ym, COUNT(*) as cnt FROM distributions WHERE {mc} GROUP BY SUBSTR(created_at, 1, 7)"),
            mp,
        )
        month_counts = {row[0]: row[1] for row in result.fetchall()}
        for i in range(5, -1, -1):
            month_start = datetime(now.year, now.month, 1) - timedelta(days=i * 30)
            ym = month_start.strftime("%Y-%m")
            by_month.append({"month": month_start.strftime("%B %Y"), "count": month_counts.get(ym, 0)})

        # Top distributors
        c, p = _build_date_filter("status = 'delivered'", {}, start_date, end_date)
        result = await session.execute(
            text(f"""SELECT to_user_name, SUM(device_count) as total
            FROM distributions WHERE {c}
            GROUP BY to_user_name ORDER BY total DESC LIMIT 5"""), p
        )
        top = result.fetchall()

        return {
            "total": total,
            "by_status": by_status,
            "by_month": by_month,
            "top_distributors": [{"name": r[0], "devices": r[1]} for r in top],
            "generated_at": datetime.now().replace(tzinfo=None).isoformat()
        }


async def get_defect_summary(start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
    """Generate defect summary report"""
    async with async_session_factory() as session:
        cond, prm = _build_date_filter("1=1", {}, start_date, end_date)
        total = await _count(session, "defects", cond, prm)

        by_status = {}
        result = await session.execute(text(f"SELECT status, COUNT(*) as cnt FROM defects WHERE {cond} GROUP BY status"), prm)
        for row in result.fetchall():
            by_status[row[0]] = row[1]

        by_severity = {}
        result = await session.execute(text(f"SELECT severity, COUNT(*) as cnt FROM defects WHERE {cond} GROUP BY severity"), prm)
        for row in result.fetchall():
            by_severity[row[0]] = row[1]

        by_type = {}
        result = await session.execute(text(f"SELECT defect_type, COUNT(*) as cnt FROM defects WHERE {cond} GROUP BY defect_type"), prm)
        for row in result.fetchall():
            by_type[row[0]] = row[1]

        by_month = []
        now = datetime.now().replace(tzinfo=None)
        six_months_ago = (datetime(now.year, now.month, 1) - timedelta(days=180)).isoformat()
        mc, mp = _build_date_filter("created_at >= :six_months_ago", {"six_months_ago": six_months_ago}, start_date, end_date)
        result = await session.execute(
            text(f"SELECT SUBSTR(created_at, 1, 7) as ym, COUNT(*) as cnt FROM defects WHERE {mc} GROUP BY SUBSTR(created_at, 1, 7)"),
            mp,
        )
        month_counts = {row[0]: row[1] for row in result.fetchall()}
        for i in range(5, -1, -1):
            month_start = datetime(now.year, now.month, 1) - timedelta(days=i * 30)
            ym = month_start.strftime("%Y-%m")
            by_month.append({"month": month_start.strftime("%B %Y"), "count": month_counts.get(ym, 0)})

        return {
            "total": total,
            "by_status": by_status,
            "by_severity": by_severity,
            "by_type": by_type,
            "by_month": by_month,
            "generated_at": datetime.now().replace(tzinfo=None).isoformat()
        }


async def get_return_summary(start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
    """Generate return summary report"""
    async with async_session_factory() as session:
        cond, prm = _build_date_filter("1=1", {}, start_date, end_date)
        total = await _count(session, "returns", cond, prm)

        by_status = {}
        result = await session.execute(text(f"SELECT status, COUNT(*) as cnt FROM returns WHERE {cond} GROUP BY status"), prm)
        for row in result.fetchall():
            by_status[row[0]] = row[1]

        by_reason = {}
        result = await session.execute(text(f"SELECT reason, COUNT(*) as cnt FROM returns WHERE {cond} GROUP BY reason"), prm)
        for row in result.fetchall():
            by_reason[row[0]] = row[1]

        by_month = []
        now = datetime.now().replace(tzinfo=None)
        six_months_ago = (datetime(now.year, now.month, 1) - timedelta(days=180)).isoformat()
        mc, mp = _build_date_filter("created_at >= :six_months_ago", {"six_months_ago": six_months_ago}, start_date, end_date)
        result = await session.execute(
            text(f"SELECT SUBSTR(created_at, 1, 7) as ym, COUNT(*) as cnt FROM returns WHERE {mc} GROUP BY SUBSTR(created_at, 1, 7)",
            ), mp,
        )
        month_counts = {row[0]: row[1] for row in result.fetchall()}
        for i in range(5, -1, -1):
            month_start = datetime(now.year, now.month, 1) - timedelta(days=i * 30)
            ym = month_start.strftime("%Y-%m")
            by_month.append({"month": month_start.strftime("%B %Y"), "count": month_counts.get(ym, 0)})

        return {
            "total": total,
            "by_status": by_status,
            "by_reason": by_reason,
            "by_month": by_month,
            "generated_at": datetime.now().replace(tzinfo=None).isoformat()
        }


async def get_user_activity_report(start_date: Optional[str] = None,
                                    end_date: Optional[str] = None) -> Dict[str, Any]:
    """Generate user activity report"""
    async with async_session_factory() as session:
        by_role = {}
        result = await session.execute(text("SELECT role, COUNT(*) as cnt FROM users GROUP BY role"))
        for row in result.fetchall():
            by_role[row[0]] = row[1]

        if start_date:
            active_users = await _count(session, "users", "last_login >= :last_login", {"last_login": start_date})
        else:
            thirty_days_ago = (datetime.now().replace(tzinfo=None) - timedelta(days=30)).isoformat()
            active_users = await _count(session, "users", "last_login >= :thirty_days_ago", {"thirty_days_ago": thirty_days_ago})
        total_users = await _count(session, "users")

        result = await session.execute(text("SELECT * FROM device_history ORDER BY timestamp DESC LIMIT 50"))
        rows = result.mappings().all()

        return {
            "total_users": total_users,
            "active_users": active_users,
            "by_role": by_role,
            "recent_activities": [dict(r) for r in rows],
            "generated_at": datetime.now().replace(tzinfo=None).isoformat()
        }


async def get_device_utilization_report(start_date: Optional[str] = None,
                                         end_date: Optional[str] = None) -> Dict[str, Any]:
    """Generate device utilization report"""
    async with async_session_factory() as session:
        total_devices = await _count(session, "devices")
        if start_date or end_date:
            cond, prm = _build_date_filter("status IN ('distributed', 'in_use')", {}, start_date, end_date)
            in_use = await _count(session, "devices", cond, prm)
            ac, ap = _build_date_filter("status = 'available'", {}, start_date, end_date)
            available = await _count(session, "devices", ac, ap)
            dfc, dfp = _build_date_filter("status = 'defective'", {}, start_date, end_date)
            defective = await _count(session, "devices", dfc, dfp)
        else:
            in_use = await _count(session, "devices", "status IN ('distributed', 'in_use')")
            available = await _count(session, "devices", "status = 'available'")
            defective = await _count(session, "devices", "status = 'defective'")

        utilization_rate = (in_use / total_devices * 100) if total_devices > 0 else 0

        return {
            "total_devices": total_devices,
            "in_use": in_use,
            "available": available,
            "defective": defective,
            "utilization_rate": round(utilization_rate, 2),
            "generated_at": datetime.now().replace(tzinfo=None).isoformat()
        }


def _build_device_journey(history_rows: List[Dict[str, Any]], device_row: Dict[str, Any]) -> Dict[str, str]:
    """Build start, intermediate path, and current location strings for a device."""
    sorted_history = sorted(history_rows, key=lambda row: str(row.get("timestamp") or ""))

    start_location = "PDIC"
    for row in sorted_history:
        action = str(row.get("action") or "").lower()
        location = str(row.get("location") or "").strip()
        if action == "registered" and location:
            start_location = location
            break

    path_nodes = [start_location]
    for row in sorted_history:
        action = str(row.get("action") or "").lower()
        if action != "distributed":
            continue

        next_point = str(row.get("to_user_name") or "").strip() or str(row.get("location") or "").strip()
        if next_point and next_point != path_nodes[-1]:
            path_nodes.append(next_point)

    current_location = (
        str(device_row.get("current_holder_name") or "").strip()
        or str(device_row.get("current_location") or "").strip()
        or (path_nodes[-1] if path_nodes else "")
    )

    full_path_nodes = path_nodes[:]
    if current_location and (not full_path_nodes or current_location != full_path_nodes[-1]):
        full_path_nodes.append(current_location)

    passed_through_nodes = []
    if len(full_path_nodes) >= 3:
        passed_through_nodes = full_path_nodes[1:-1]

    return {
        "started_from": start_location,
        "passed_through": " -> ".join(passed_through_nodes),
        "current_at": current_location,
        "journey_path": " -> ".join(full_path_nodes),
    }


def _build_device_backup_file(rows: List[Dict[str, Any]], file_format: str) -> Dict[str, Any]:
    """Build downloadable backup payload for devices in CSV or XLSX format."""
    headers = [
        "device_db_id",
        "device_id",
        "serial_number",
        "mac_address",
        "nuid",
        "device_type",
        "model",
        "manufacturer",
        "status",
        "current_holder_name",
        "current_holder_type",
        "started_from",
        "passed_through",
        "current_at",
        "journey_path",
        "created_at",
        "updated_at",
    ]

    normalized = str(file_format or "xlsx").strip().lower()
    generated_ts = datetime.now().replace(tzinfo=None).strftime("%Y%m%d_%H%M%S")

    if normalized == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Device Backup"
        sheet.append(headers)

        for row in rows:
            sheet.append([row.get(col, "") for col in headers])

        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)
        return {
            "content": payload.getvalue(),
            "filename": f"device-backup-{generated_ts}.xlsx",
            "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

    if normalized == "csv":
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in headers})

        return {
            "content": buffer.getvalue().encode("utf-8"),
            "filename": f"device-backup-{generated_ts}.csv",
            "media_type": "text/csv",
        }

    raise ValueError("Unsupported export format. Use 'csv' or 'xlsx'")


async def get_device_backup_export(file_format: str = "xlsx") -> Dict[str, Any]:
    """Generate a full device backup export including journey path details."""
    MAX_EXPORT = 100000
    async with async_session_factory() as session:
        result = await session.execute(text("SELECT * FROM devices ORDER BY id ASC LIMIT :max_export"), {"max_export": MAX_EXPORT})
        device_rows = [dict(r) for r in result.mappings().all()]

        result = await session.execute(
            text("SELECT * FROM device_history ORDER BY timestamp ASC LIMIT :max_history"),
            {"max_history": MAX_EXPORT * 5}
        )
        history_rows = [dict(r) for r in result.mappings().all()]

    history_by_device: Dict[str, List[Dict[str, Any]]] = {}
    for row in history_rows:
        key = str(row.get("device_id") or "")
        if not key:
            continue
        history_by_device.setdefault(key, []).append(row)

    export_rows: List[Dict[str, Any]] = []
    for device in device_rows:
        device_key = str(device.get("id") or "")
        journey = _build_device_journey(history_by_device.get(device_key, []), device)

        export_rows.append(
            {
                "device_db_id": str(device.get("id") or ""),
                "device_id": str(device.get("device_id") or ""),
                "serial_number": str(device.get("serial_number") or ""),
                "mac_address": str(device.get("mac_address") or ""),
                "nuid": str(device.get("nuid") or ""),
                "device_type": str(device.get("device_type") or ""),
                "model": str(device.get("model") or ""),
                "manufacturer": str(device.get("manufacturer") or ""),
                "status": str(device.get("status") or ""),
                "current_holder_name": str(device.get("current_holder_name") or ""),
                "current_holder_type": str(device.get("current_holder_type") or ""),
                "started_from": journey["started_from"],
                "passed_through": journey["passed_through"],
                "current_at": journey["current_at"],
                "journey_path": journey["journey_path"],
                "created_at": str(device.get("created_at") or ""),
                "updated_at": str(device.get("updated_at") or ""),
            }
        )

    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, _build_device_backup_file, export_rows, file_format)


def _build_returns_defects_backup_file(
    returns_rows: List[Dict[str, Any]],
    defects_rows: List[Dict[str, Any]],
    file_format: str,
) -> Dict[str, Any]:
    """Build downloadable backup payload for returns and defects."""
    def _xlsx_cell_value(value: Any) -> Any:
        if value is None:
            return ""
        return value

    normalized = str(file_format or "xlsx").strip().lower()
    generated_ts = datetime.now().replace(tzinfo=None).strftime("%Y%m%d_%H%M%S")

    return_headers = [
        "return_id",
        "defect_report_id",
        "device_identifier",
        "device_model",
        "device_serial",
        "device_nuid",
        "device_type",
        "requested_by_name",
        "reason",
        "description",
        "status",
        "request_date",
        "return_approved_at",
        "return_approved_by_name",
        "received_date",
        "created_at",
        "updated_at",
    ]

    defect_headers = [
        "report_id",
        "device_identifier",
        "device_model",
        "device_serial",
        "device_nuid",
        "device_type",
        "reported_by_name",
        "operator_name",
        "sub_distributor_name",
        "defect_type",
        "severity",
        "description",
        "status",
        "resolution",
        "replacement_by_name",
        "resolved_at",
        "defect_approved_by_name",
        "defect_approved_at",
        "return_approved_by_name",
        "return_approved_at",
        "created_at",
        "updated_at",
    ]

    if normalized == "xlsx":
        workbook = Workbook()

        returns_sheet = workbook.active
        returns_sheet.title = "Returned Devices"
        returns_sheet.append(return_headers)
        for row in returns_rows:
            returns_sheet.append([_xlsx_cell_value(row.get(col)) for col in return_headers])

        defects_sheet = workbook.create_sheet("Defect Reports")
        defects_sheet.append(defect_headers)
        for row in defects_rows:
            defects_sheet.append([_xlsx_cell_value(row.get(col)) for col in defect_headers])

        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)
        return {
            "content": payload.getvalue(),
            "filename": f"returns-defects-backup-{generated_ts}.xlsx",
            "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

    if normalized == "csv":
        combined_headers = [
            "record_type",
            "id",
            "device_identifier",
            "device_model",
            "device_serial",
            "device_nuid",
            "device_type",
            "person_name",
            "category",
            "status",
            "description",
            "created_at",
            "updated_at",
        ]

        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=combined_headers)
        writer.writeheader()

        for row in returns_rows:
            writer.writerow(
                {
                    "record_type": "return",
                    "id": str(row.get("return_id") or ""),
                    "device_identifier": str(row.get("device_identifier") or ""),
                    "device_model": str(row.get("device_model") or ""),
                    "device_serial": str(row.get("device_serial") or ""),
                    "device_nuid": str(row.get("device_nuid") or ""),
                    "device_type": str(row.get("device_type") or ""),
                    "person_name": str(row.get("requested_by_name") or ""),
                    "category": str(row.get("reason") or ""),
                    "status": str(row.get("status") or ""),
                    "description": str(row.get("description") or ""),
                    "created_at": str(row.get("created_at") or ""),
                    "updated_at": str(row.get("updated_at") or ""),
                }
            )

        for row in defects_rows:
            writer.writerow(
                {
                    "record_type": "defect",
                    "id": str(row.get("report_id") or ""),
                    "device_identifier": str(row.get("device_identifier") or ""),
                    "device_model": str(row.get("device_model") or ""),
                    "device_serial": str(row.get("device_serial") or ""),
                    "device_nuid": str(row.get("device_nuid") or ""),
                    "device_type": str(row.get("device_type") or ""),
                    "person_name": str(row.get("reported_by_name") or ""),
                    "category": str(row.get("defect_type") or ""),
                    "status": str(row.get("status") or ""),
                    "description": str(row.get("description") or ""),
                    "created_at": str(row.get("created_at") or ""),
                    "updated_at": str(row.get("updated_at") or ""),
                }
            )

        return {
            "content": buffer.getvalue().encode("utf-8"),
            "filename": f"returns-defects-backup-{generated_ts}.csv",
            "media_type": "text/csv",
        }

    raise ValueError("Unsupported export format. Use 'csv' or 'xlsx'")


async def get_returns_defects_backup_export(file_format: str = "xlsx") -> Dict[str, Any]:
    """Generate backup export for return requests and defect reports."""
    async with async_session_factory() as session:
        result = await session.execute(text("SELECT * FROM returns ORDER BY created_at DESC"))
        returns_rows = [dict(r) for r in result.mappings().all()]

        result = await session.execute(text("SELECT * FROM defects ORDER BY created_at DESC"))
        defects_rows = [dict(r) for r in result.mappings().all()]

        result = await session.execute(text("SELECT id, device_id, model, serial_number, mac_address, nuid, device_type FROM devices"))
        devices_rows = [dict(r) for r in result.mappings().all()]

        result = await session.execute(text("SELECT id, name FROM users"))
        users_rows = [dict(r) for r in result.mappings().all()]

    device_lookup: Dict[str, Dict[str, Any]] = {}
    for device in devices_rows:
        db_id = str(device.get("id") or "").strip()
        business_id = str(device.get("device_id") or "").strip()
        if db_id:
            device_lookup[db_id] = device
        if business_id:
            device_lookup[business_id] = device

    user_name_lookup: Dict[str, str] = {
        str(user.get("id") or "").strip(): str(user.get("name") or "").strip()
        for user in users_rows
        if str(user.get("id") or "").strip()
    }

    defect_lookup: Dict[str, Dict[str, Any]] = {}
    for defect in defects_rows:
        db_id = str(defect.get("id") or "").strip()
        if db_id:
            defect_lookup[db_id] = defect

    for row in returns_rows:
        raw_device_id = str(row.get("device_id") or "").strip()
        resolved_device = device_lookup.get(raw_device_id)
        row["device_identifier"] = str(
            (resolved_device or {}).get("device_id")
            or raw_device_id
        )
        row["device_model"] = str((resolved_device or {}).get("model") or "")
        row["device_nuid"] = str((row.get("device_nuid") or "") or ((resolved_device or {}).get("nuid") or ""))

        resolved_type = str((resolved_device or {}).get("device_type") or row.get("device_type") or "")
        is_sb = resolved_type.strip().lower() in {"set-top box", "set top box", "sb", "stb"}
        if is_sb:
            row["device_serial"] = ""

        linked_defect = defect_lookup.get(str(row.get("defect_id") or "").strip()) or {}
        row["defect_report_id"] = str(linked_defect.get("report_id") or "")
        row["requested_by_name"] = str(
            row.get("requested_by_name") or linked_defect.get("reported_by_name") or ""
        )
        row["description"] = str(row.get("description") or linked_defect.get("description") or "")
        row["return_approved_at"] = row.get("return_approved_at") or linked_defect.get("return_approved_at") or ""
        row["return_approved_by_name"] = str(
            row.get("return_approved_by_name")
            or linked_defect.get("return_approved_by_name")
            or ""
        )

    for row in defects_rows:
        raw_device_id = str(row.get("device_id") or "").strip()
        resolved_device = device_lookup.get(raw_device_id)
        row["device_identifier"] = str(
            (resolved_device or {}).get("device_id")
            or raw_device_id
        )
        row["device_model"] = str((resolved_device or {}).get("model") or "")
        row["device_nuid"] = str((row.get("device_nuid") or "") or ((resolved_device or {}).get("nuid") or ""))

        resolved_type = str((resolved_device or {}).get("device_type") or row.get("device_type") or "")
        is_sb = resolved_type.strip().lower() in {"set-top box", "set top box", "sb", "stb"}
        if is_sb:
            row["device_serial"] = ""

        if not str(row.get("reported_by_name") or "").strip():
            row["reported_by_name"] = user_name_lookup.get(str(row.get("reported_by") or "").strip(), "")
        resolved_name = str(row.get("replacement_by_name") or "").strip()
        if not resolved_name:
            resolved_name = user_name_lookup.get(str(row.get("replacement_by") or "").strip(), resolved_name)
        row["replacement_by_name"] = resolved_name

        if not str(row.get("defect_approved_by_name") or "").strip():
            row["defect_approved_by_name"] = user_name_lookup.get(str(row.get("defect_approved_by") or "").strip(), "")
        if not str(row.get("return_approved_by_name") or "").strip():
            row["return_approved_by_name"] = user_name_lookup.get(str(row.get("return_approved_by") or "").strip(), "")

        row["operator_name"] = user_name_lookup.get(str(row.get("operator_id") or "").strip(), "")
        row["sub_distributor_name"] = user_name_lookup.get(str(row.get("sub_distributor_id") or "").strip(), "")

    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(
        None,
        _build_returns_defects_backup_file,
        returns_rows,
        defects_rows,
        file_format,
    )


def _canonical_device_type(value) -> str:
    """Return a canonical lowercase key for a stored device_type string."""
    return str(value or "").strip().lower().replace("_", " ").replace("-", " ")


def _is_sb_device_type(value) -> bool:
    return _canonical_device_type(value) in {"sb", "stb", "set top box", "settopbox"}


def _is_ont_device_type(value) -> bool:
    return _canonical_device_type(value) == "ont"


def _build_device_date_filter(base_params: Optional[dict], start_date: Optional[str], end_date: Optional[str], prefix: str = "") -> str:
    """Build a ``devices.created_at`` date filter fragment with named params.

    Mirrors the dashboard date-filter semantics (device counts restricted to
    devices created within the requested window). When no date is provided,
    returns ``1=1`` so callers can append the fragment unconditionally.
    """
    params = base_params if base_params is not None else {}
    conds = []
    if start_date:
        conds.append(f"d.created_at >= :{prefix}start_date")
        params[f"{prefix}start_date"] = start_date
    if end_date:
        conds.append(f"d.created_at <= :{prefix}end_date")
        params[f"{prefix}end_date"] = end_date
    return " AND ".join(conds) if conds else "1=1"


async def get_sub_distribution_report(start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
    """Build the sub-distribution hierarchy report.

    Returns one row per sub-distributor with hierarchy rollups (operators,
    clusters), identity columns (digital id / broadband id), and device
    counts broken into SB / ONT / other plus per-vendor SB and ONT counts.

    When ``start_date`` / ``end_date`` are provided, device counts are
    restricted to devices created within that window (matching the dashboard
    date-filter semantics).
    """
    async with async_session_factory() as session:
        sub_result = await session.execute(text("""
            SELECT u.id, u.name, u.email, u.phone
            FROM users u
            WHERE u.role = 'sub_distributor'
            ORDER BY u.name
        """))
        sub_rows = sub_result.mappings().all()
        sub_ids = [int(r["id"]) for r in sub_rows]

        sub_identities = {
            int(r["id"]): {"digital_id": None, "broadband_id": None, "digital_ids": []}
            for r in sub_rows
        }

        if sub_ids:
            identity_result = await session.execute(
                text("""
                    SELECT di.id, di.user_id, di.digital_id, di.broadband_id, di.is_primary, di.created_at
                    FROM digital_identities di
                    WHERE di.user_id IN :sub_ids
                    ORDER BY di.is_primary DESC, di.id ASC
                """).bindparams(bindparam("sub_ids", expanding=True)),
                {"sub_ids": sub_ids},
            )
            for row in identity_result.mappings().all():
                entry = sub_identities[int(row["user_id"])]
                digital = str(row["digital_id"] or "").strip() or None
                broadband = str(row["broadband_id"] or "").strip() or None
                if digital and entry["digital_id"] is None:
                    entry["digital_id"] = digital
                if broadband and entry["broadband_id"] is None:
                    entry["broadband_id"] = broadband
                entry["digital_ids"].append({
                    "id": row["id"],
                    "user_id": row["user_id"],
                    "digital_id": digital,
                    "broadband_id": broadband,
                    "is_primary": bool(row["is_primary"]),
                    "created_at": row["created_at"].isoformat() if row["created_at"] else None,
                })

        member_counts = {
            int(r["id"]): {"operators": set(), "clusters": 0}
            for r in sub_rows
        }
        operator_identities = {
            int(r["id"]): {"with_digital": set(), "with_broadband": set()}
            for r in sub_rows
        }
        device_totals = {
            int(r["id"]): {
                "total": 0,
                "sb": 0,
                "ont": 0,
                "other": 0,
                "sb_by_vendor": {},
                "ont_by_vendor": {},
            }
            for r in sub_rows
        }

        if sub_ids:
            # Load the user hierarchy once instead of running the recursive
            # CTE in three separate queries. The users table is small (it holds
            # accounts, not devices), so one flat read plus in-memory traversal
            # is far cheaper than re-walking the tree per rollup.
            user_result = await session.execute(text("SELECT id, parent_id, role FROM users"))
            role_by_id = {}
            children_by_parent = {}
            for row in user_result.mappings().all():
                uid = int(row["id"])
                role_by_id[uid] = str(row["role"] or "")
                parent_id = row["parent_id"]
                if parent_id is not None:
                    children_by_parent.setdefault(int(parent_id), []).append(uid)

            sub_to_members = {}
            sub_to_operators = {}
            holder_to_sub = {}
            for sub_id in sub_ids:
                members = set()
                stack = [sub_id]
                while stack:
                    uid = stack.pop()
                    if uid in members:
                        continue
                    members.add(uid)
                    holder_to_sub[uid] = sub_id
                    stack.extend(children_by_parent.get(uid, []))
                sub_to_members[sub_id] = members
                sub_to_operators[sub_id] = {
                    uid for uid in members if role_by_id.get(uid) == "operator"
                }

            for sub_id, members in sub_to_members.items():
                for uid in members:
                    role = role_by_id.get(uid)
                    if role == "operator":
                        member_counts[sub_id]["operators"].add(uid)
                    elif role == "cluster":
                        member_counts[sub_id]["clusters"] += 1

            all_operator_ids = [uid for ids in sub_to_operators.values() for uid in ids]
            if all_operator_ids:
                identity_result = await session.execute(
                    text("""
                        SELECT di.user_id,
                               MAX(CASE WHEN TRIM(COALESCE(di.digital_id, '')) != '' THEN 1 ELSE 0 END) AS has_digital,
                               MAX(CASE WHEN TRIM(COALESCE(di.broadband_id, '')) != '' THEN 1 ELSE 0 END) AS has_broadband
                        FROM digital_identities di
                        WHERE di.user_id IN :operator_ids
                        GROUP BY di.user_id
                    """).bindparams(bindparam("operator_ids", expanding=True)),
                    {"operator_ids": all_operator_ids},
                )
                for row in identity_result.mappings().all():
                    uid = int(row["user_id"])
                    for sub_id, op_ids in sub_to_operators.items():
                        if uid not in op_ids:
                            continue
                        if row["has_digital"]:
                            operator_identities[sub_id]["with_digital"].add(uid)
                        if row["has_broadband"]:
                            operator_identities[sub_id]["with_broadband"].add(uid)

            all_holder_ids = list(holder_to_sub.keys())
            device_date_params = {}
            device_date_cond = _build_device_date_filter(device_date_params, start_date, end_date)
            device_result = await session.execute(
                text(f"""
                    SELECT d.current_holder_id, d.device_type, d.manufacturer, COUNT(*) AS cnt
                    FROM devices d
                    WHERE d.current_holder_id IS NOT NULL
                      AND d.current_holder_id IN :holder_ids
                      AND {device_date_cond}
                    GROUP BY d.current_holder_id, d.device_type, d.manufacturer
                """).bindparams(bindparam("holder_ids", expanding=True)),
                {**device_date_params, "holder_ids": all_holder_ids},
            )
            for row in device_result.mappings().all():
                sub_id = holder_to_sub[int(row["current_holder_id"])]
                device_type = str(row["device_type"] or "")
                vendor = str(row["manufacturer"] or "").strip() or "Unknown"
                count = int(row["cnt"])
                bucket = device_totals[sub_id]
                bucket["total"] += count
                if _is_sb_device_type(device_type):
                    bucket["sb"] += count
                    bucket["sb_by_vendor"][vendor] = bucket["sb_by_vendor"].get(vendor, 0) + count
                elif _is_ont_device_type(device_type):
                    bucket["ont"] += count
                    bucket["ont_by_vendor"][vendor] = bucket["ont_by_vendor"].get(vendor, 0) + count
                else:
                    bucket["other"] += count

        rows = []
        for sub_row in sub_rows:
            sub_id = int(sub_row["id"])
            identity = sub_identities[sub_id]
            members = member_counts[sub_id]
            op_ids = operator_identities[sub_id]
            devices = device_totals[sub_id]
            rows.append({
                "sub_id": sub_id,
                "sub_name": str(sub_row["name"] or ""),
                "email": str(sub_row["email"] or ""),
                "phone": str(sub_row["phone"] or ""),
                "digital_id": identity["digital_id"],
                "broadband_id": identity["broadband_id"],
                "digital_ids": identity["digital_ids"],
                "total_operators": len(members["operators"]),
                "operators_with_digital_id": len(op_ids["with_digital"]),
                "operators_with_broadband_id": len(op_ids["with_broadband"]),
                "total_clusters": members["clusters"],
                "device_count": devices["total"],
                "sb_device_count": devices["sb"],
                "ont_device_count": devices["ont"],
                "other_device_count": devices["other"],
                "sb_by_vendor": dict(devices["sb_by_vendor"]),
                "ont_by_vendor": dict(devices["ont_by_vendor"]),
            })

        return {
            "sub_distributions": rows,
            "generated_at": datetime.now().replace(tzinfo=None).isoformat(),
        }


async def get_cluster_report(scope: Optional[dict] = None, start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
    """Build the cluster hierarchy report.

    Returns one row per cluster with its parent sub-distribution, hierarchy
    rollups (operators), identity columns (digital id / broadband id), and
    device counts broken into SB / ONT / other plus per-vendor SB and ONT
    counts. Mirrors the sub-distribution report.

    When a ``scope`` dict (from ``_resolve_report_scope``) is provided, only
    clusters within that user's chain are returned. When ``start_date`` /
    ``end_date`` are provided, device counts are restricted to devices created
    within that window (matching the dashboard date-filter semantics).
    """
    async with async_session_factory() as session:
        cluster_result = await session.execute(text("""
            SELECT c.id, c.name, c.email, c.phone, p.id AS sub_id, p.name AS sub_name
            FROM users c
            LEFT JOIN users p ON p.id = c.parent_id
            WHERE c.role = 'cluster'
            ORDER BY p.name, c.name
        """))
        cluster_rows = cluster_result.mappings().all()
        cluster_ids = [int(r["id"]) for r in cluster_rows]

        allowed_cluster_ids = None
        scope_mode = (scope or {}).get("scope")
        if scope_mode == "sub":
            scope_sub_id = (scope or {}).get("sub_id")
            allowed_cluster_ids = {
                int(r["id"])
                for r in cluster_rows
                if scope_sub_id is not None and r["sub_id"] is not None
                and int(r["sub_id"]) == scope_sub_id
            }
        elif scope_mode == "cluster":
            allowed_cluster_ids = {int((scope or {}).get("cluster_id"))}
        elif scope_mode == "operator":
            allowed_cluster_ids = set()

        cluster_identities = {
            int(r["id"]): {"digital_id": None, "broadband_id": None, "digital_ids": []}
            for r in cluster_rows
        }

        if cluster_ids:
            identity_result = await session.execute(
                text("""
                    SELECT di.id, di.user_id, di.digital_id, di.broadband_id, di.is_primary, di.created_at
                    FROM digital_identities di
                    WHERE di.user_id IN :cluster_ids
                    ORDER BY di.is_primary DESC, di.id ASC
                """).bindparams(bindparam("cluster_ids", expanding=True)),
                {"cluster_ids": cluster_ids},
            )
            for row in identity_result.mappings().all():
                entry = cluster_identities[int(row["user_id"])]
                digital = str(row["digital_id"] or "").strip() or None
                broadband = str(row["broadband_id"] or "").strip() or None
                if digital and entry["digital_id"] is None:
                    entry["digital_id"] = digital
                if broadband and entry["broadband_id"] is None:
                    entry["broadband_id"] = broadband
                entry["digital_ids"].append({
                    "id": row["id"],
                    "user_id": row["user_id"],
                    "digital_id": digital,
                    "broadband_id": broadband,
                    "is_primary": bool(row["is_primary"]),
                    "created_at": row["created_at"].isoformat() if row["created_at"] else None,
                })

        member_counts = {
            int(r["id"]): {"operators": set()}
            for r in cluster_rows
        }
        operator_identities = {
            int(r["id"]): {"with_digital": set(), "with_broadband": set()}
            for r in cluster_rows
        }
        device_totals = {
            int(r["id"]): {
                "total": 0,
                "sb": 0,
                "ont": 0,
                "other": 0,
                "sb_by_vendor": {},
                "ont_by_vendor": {},
            }
            for r in cluster_rows
        }

        if cluster_ids:
            active_cluster_ids = (
                [cid for cid in cluster_ids if cid in allowed_cluster_ids]
                if allowed_cluster_ids is not None
                else cluster_ids
            )

            # Load the user hierarchy once instead of running the recursive
            # CTE in three separate queries (see get_sub_distribution_report).
            user_result = await session.execute(text("SELECT id, parent_id, role FROM users"))
            role_by_id = {}
            children_by_parent = {}
            for row in user_result.mappings().all():
                uid = int(row["id"])
                role_by_id[uid] = str(row["role"] or "")
                parent_id = row["parent_id"]
                if parent_id is not None:
                    children_by_parent.setdefault(int(parent_id), []).append(uid)

            cluster_to_members = {}
            cluster_to_operators = {}
            holder_to_cluster = {}
            for cluster_id in active_cluster_ids:
                members = set()
                stack = [cluster_id]
                while stack:
                    uid = stack.pop()
                    if uid in members:
                        continue
                    members.add(uid)
                    holder_to_cluster[uid] = cluster_id
                    stack.extend(children_by_parent.get(uid, []))
                cluster_to_members[cluster_id] = members
                cluster_to_operators[cluster_id] = {
                    uid for uid in members if role_by_id.get(uid) == "operator"
                }

            for cluster_id, members in cluster_to_members.items():
                for uid in members:
                    if role_by_id.get(uid) == "operator":
                        member_counts[cluster_id]["operators"].add(uid)

            all_operator_ids = [uid for ids in cluster_to_operators.values() for uid in ids]
            if all_operator_ids:
                identity_result = await session.execute(
                    text("""
                        SELECT di.user_id,
                               MAX(CASE WHEN TRIM(COALESCE(di.digital_id, '')) != '' THEN 1 ELSE 0 END) AS has_digital,
                               MAX(CASE WHEN TRIM(COALESCE(di.broadband_id, '')) != '' THEN 1 ELSE 0 END) AS has_broadband
                        FROM digital_identities di
                        WHERE di.user_id IN :operator_ids
                        GROUP BY di.user_id
                    """).bindparams(bindparam("operator_ids", expanding=True)),
                    {"operator_ids": all_operator_ids},
                )
                for row in identity_result.mappings().all():
                    uid = int(row["user_id"])
                    for cluster_id, op_ids in cluster_to_operators.items():
                        if uid not in op_ids:
                            continue
                        if row["has_digital"]:
                            operator_identities[cluster_id]["with_digital"].add(uid)
                        if row["has_broadband"]:
                            operator_identities[cluster_id]["with_broadband"].add(uid)

            all_holder_ids = list(holder_to_cluster.keys())
            device_date_params = {}
            device_date_cond = _build_device_date_filter(device_date_params, start_date, end_date)
            device_result = await session.execute(
                text(f"""
                    SELECT d.current_holder_id, d.device_type, d.manufacturer, COUNT(*) AS cnt
                    FROM devices d
                    WHERE d.current_holder_id IS NOT NULL
                      AND d.current_holder_id IN :holder_ids
                      AND {device_date_cond}
                    GROUP BY d.current_holder_id, d.device_type, d.manufacturer
                """).bindparams(bindparam("holder_ids", expanding=True)),
                {**device_date_params, "holder_ids": all_holder_ids},
            )
            for row in device_result.mappings().all():
                cluster_id = holder_to_cluster[int(row["current_holder_id"])]
                device_type = str(row["device_type"] or "")
                vendor = str(row["manufacturer"] or "").strip() or "Unknown"
                count = int(row["cnt"])
                bucket = device_totals[cluster_id]
                bucket["total"] += count
                if _is_sb_device_type(device_type):
                    bucket["sb"] += count
                    bucket["sb_by_vendor"][vendor] = bucket["sb_by_vendor"].get(vendor, 0) + count
                elif _is_ont_device_type(device_type):
                    bucket["ont"] += count
                    bucket["ont_by_vendor"][vendor] = bucket["ont_by_vendor"].get(vendor, 0) + count
                else:
                    bucket["other"] += count

        rows = []
        for cluster_row in cluster_rows:
            cluster_id = int(cluster_row["id"])
            if allowed_cluster_ids is not None and cluster_id not in allowed_cluster_ids:
                continue
            identity = cluster_identities[cluster_id]
            members = member_counts[cluster_id]
            op_ids = operator_identities[cluster_id]
            devices = device_totals[cluster_id]
            rows.append({
                "cluster_id": cluster_id,
                "cluster_name": str(cluster_row["name"] or ""),
                "email": str(cluster_row["email"] or ""),
                "phone": str(cluster_row["phone"] or ""),
                "sub_id": int(cluster_row["sub_id"]) if cluster_row["sub_id"] else None,
                "sub_name": str(cluster_row["sub_name"] or ""),
                "digital_id": identity["digital_id"],
                "broadband_id": identity["broadband_id"],
                "digital_ids": identity["digital_ids"],
                "total_operators": len(members["operators"]),
                "operators_with_digital_id": len(op_ids["with_digital"]),
                "operators_with_broadband_id": len(op_ids["with_broadband"]),
                "device_count": devices["total"],
                "sb_device_count": devices["sb"],
                "ont_device_count": devices["ont"],
                "other_device_count": devices["other"],
                "sb_by_vendor": dict(devices["sb_by_vendor"]),
                "ont_by_vendor": dict(devices["ont_by_vendor"]),
            })

        return {
            "clusters": rows,
            "generated_at": datetime.now().replace(tzinfo=None).isoformat(),
        }


async def get_operator_report(scope: Optional[dict] = None, start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
    """Build the operator hierarchy report.

    Returns one row per operator with its parent sub-distribution and cluster
    (if any), identity columns (digital id / broadband id), and device counts
    broken into SB / ONT / other plus per-vendor SB and ONT counts.

    When a ``scope`` dict (from ``_resolve_report_scope``) is provided, only
    operators within that user's chain are returned. When ``start_date`` /
    ``end_date`` are provided, device counts are restricted to devices created
    within that window (matching the dashboard date-filter semantics).
    """
    async with async_session_factory() as session:
        operator_result = await session.execute(text("""
            SELECT o.id, o.name, o.email, o.phone,
                   p.id AS parent_id, p.name AS parent_name, p.role AS parent_role,
                   g.id AS sub_id, g.name AS sub_name
            FROM users o
            LEFT JOIN users p ON p.id = o.parent_id
            LEFT JOIN users g ON g.id = CASE WHEN p.role = 'cluster' THEN p.parent_id ELSE p.id END
            WHERE o.role = 'operator'
            ORDER BY g.name, p.name, o.name
        """))
        operator_rows = operator_result.mappings().all()

        scope_mode = (scope or {}).get("scope")
        if scope_mode == "sub":
            scope_sub_id = (scope or {}).get("sub_id")
            operator_rows = [
                r for r in operator_rows
                if scope_sub_id is not None and r["sub_id"] is not None
                and int(r["sub_id"]) == scope_sub_id
            ]
        elif scope_mode == "cluster":
            scope_cluster_id = int((scope or {}).get("cluster_id"))
            operator_rows = [
                r for r in operator_rows
                if r["parent_id"] is not None and int(r["parent_id"]) == scope_cluster_id
            ]
        elif scope_mode == "operator":
            scope_operator_id = int((scope or {}).get("operator_id"))
            operator_rows = [
                r for r in operator_rows if int(r["id"]) == scope_operator_id
            ]

        operator_ids = [int(r["id"]) for r in operator_rows]

        operator_identities = {
            int(r["id"]): {"digital_id": None, "broadband_id": None, "digital_ids": []}
            for r in operator_rows
        }
        device_totals = {
            int(r["id"]): {
                "total": 0,
                "sb": 0,
                "ont": 0,
                "other": 0,
                "sb_by_vendor": {},
                "ont_by_vendor": {},
            }
            for r in operator_rows
        }

        if operator_ids:
            identity_result = await session.execute(
                text("""
                    SELECT di.id, di.user_id, di.digital_id, di.broadband_id, di.is_primary, di.created_at
                    FROM digital_identities di
                    WHERE di.user_id IN :operator_ids
                    ORDER BY di.is_primary DESC, di.id ASC
                """).bindparams(bindparam("operator_ids", expanding=True)),
                {"operator_ids": operator_ids},
            )
            for row in identity_result.mappings().all():
                entry = operator_identities[int(row["user_id"])]
                digital = str(row["digital_id"] or "").strip() or None
                broadband = str(row["broadband_id"] or "").strip() or None
                if digital and entry["digital_id"] is None:
                    entry["digital_id"] = digital
                if broadband and entry["broadband_id"] is None:
                    entry["broadband_id"] = broadband
                entry["digital_ids"].append({
                    "id": row["id"],
                    "user_id": row["user_id"],
                    "digital_id": digital,
                    "broadband_id": broadband,
                    "is_primary": bool(row["is_primary"]),
                    "created_at": row["created_at"].isoformat() if row["created_at"] else None,
                })

            device_date_params = {}
            device_date_cond = _build_device_date_filter(device_date_params, start_date, end_date)
            device_result = await session.execute(
                text("""
                    SELECT d.current_holder_id AS user_id, d.device_type, d.manufacturer, COUNT(*) AS cnt
                    FROM devices d
                    WHERE d.current_holder_id IS NOT NULL
                      AND d.current_holder_id IN :operator_ids
                      AND """ + device_date_cond + """
                    GROUP BY d.current_holder_id, d.device_type, d.manufacturer
                """).bindparams(bindparam("operator_ids", expanding=True)),
                {**device_date_params, "operator_ids": operator_ids},
            )
            for row in device_result.mappings().all():
                user_id = int(row["user_id"])
                device_type = str(row["device_type"] or "")
                vendor = str(row["manufacturer"] or "").strip() or "Unknown"
                count = int(row["cnt"])
                bucket = device_totals[user_id]
                bucket["total"] += count
                if _is_sb_device_type(device_type):
                    bucket["sb"] += count
                    bucket["sb_by_vendor"][vendor] = bucket["sb_by_vendor"].get(vendor, 0) + count
                elif _is_ont_device_type(device_type):
                    bucket["ont"] += count
                    bucket["ont_by_vendor"][vendor] = bucket["ont_by_vendor"].get(vendor, 0) + count
                else:
                    bucket["other"] += count

        rows = []
        for operator_row in operator_rows:
            operator_id = int(operator_row["id"])
            identity = operator_identities[operator_id]
            devices = device_totals[operator_id]
            parent_role = str(operator_row["parent_role"] or "")
            cluster_id = int(operator_row["parent_id"]) if parent_role == "cluster" and operator_row["parent_id"] else None
            cluster_name = str(operator_row["parent_name"] or "") if cluster_id else None
            rows.append({
                "operator_id": operator_id,
                "operator_name": str(operator_row["name"] or ""),
                "email": str(operator_row["email"] or ""),
                "phone": str(operator_row["phone"] or ""),
                "sub_id": int(operator_row["sub_id"]) if operator_row["sub_id"] else None,
                "sub_name": str(operator_row["sub_name"] or ""),
                "cluster_id": cluster_id,
                "cluster_name": cluster_name,
                "digital_id": identity["digital_id"],
                "broadband_id": identity["broadband_id"],
                "digital_ids": identity["digital_ids"],
                "device_count": devices["total"],
                "sb_device_count": devices["sb"],
                "ont_device_count": devices["ont"],
                "other_device_count": devices["other"],
                "sb_by_vendor": dict(devices["sb_by_vendor"]),
                "ont_by_vendor": dict(devices["ont_by_vendor"]),
            })

        return {
            "operators": rows,
            "generated_at": datetime.now().replace(tzinfo=None).isoformat(),
        }

