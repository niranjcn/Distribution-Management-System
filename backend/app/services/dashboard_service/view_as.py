import io
from datetime import datetime
from typing import Dict, Any, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import text

from app.database_sqlalchemy import async_session_factory
from app.services import device_service

from .helpers import _build_date_filter, _resolve_scope_root_for_sub_distribution_manager, _get_descendant_user_ids
from .stats import get_dashboard_stats
from .analytics import get_advanced_dashboard_metrics


async def get_view_as_dashboard(
    target_user: Dict[str, Any],
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
) -> Dict[str, Any]:
    role = target_user.get("role")
    user_id = str(target_user.get("_id", target_user.get("id", "")))

    stats = await get_dashboard_stats(target_user, start_date, end_date)
    advanced = await get_advanced_dashboard_metrics(target_user, start_date, end_date)

    async with async_session_factory() as session:
        scope_root_id = _resolve_scope_root_for_sub_distribution_manager(target_user, user_id)

        if role in ("sub_distributor",):
            rows = (await session.execute(
                text("""SELECT id, name, email, role, status, phone, location FROM users
                WHERE role = 'operator' AND parent_id IN (
                    SELECT id FROM users WHERE role = 'cluster' AND parent_id IN (
                        SELECT id FROM users WHERE role = 'sub_distribution_manager' AND parent_id = :uid
                    )
                    UNION
                    SELECT id FROM users WHERE role = 'sub_distribution_manager' AND parent_id = :uid2
                )"""),
                {"uid": int(user_id), "uid2": int(user_id)}
            )).mappings().all()
            target_users = [dict(r) for r in rows]
        elif role == "cluster":
            rows = (await session.execute(
                text("SELECT id, name, email, role, status FROM users WHERE role = 'operator' AND parent_id = :uid"),
                {"uid": int(user_id)}
            )).mappings().all()
            target_users = [dict(r) for r in rows]
        else:
            target_users = []

        scope_ids = sorted({scope_root_id} | await _get_descendant_user_ids(session, scope_root_id)) if role in ("sub_distribution_manager", "sub_distributor") else [str(user_id)]

        dev_ph = ",".join([f":d_{i}" for i in range(len(scope_ids))]) if scope_ids else "''"
        dev_params: Dict[str, Any] = {f"d_{i}": sid for i, sid in enumerate(scope_ids)}
        dts, dte = (start_date, end_date) if (start_date or end_date) else (None, None)
        date_conds = []
        if dts:
            date_conds.append("created_at >= :ds")
            dev_params["ds"] = dts
        if dte:
            date_conds.append("created_at <= :de")
            dev_params["de"] = dte
        date_clause = " AND ".join(date_conds) if date_conds else "1=1"
        dc = f"current_holder_id IN ({dev_ph})" if scope_ids else "1=0"

        rows = (await session.execute(
            text(f"SELECT * FROM devices WHERE {dc} AND {date_clause}"), dev_params
        )).mappings().all()
        target_devices = [dict(r) for r in rows]

        str_scope_ids = [str(s) for s in scope_ids]
        str_ph = ",".join([f":s_{i}" for i in range(len(str_scope_ids))]) if str_scope_ids else "''"
        def_params: Dict[str, Any] = {f"s_{i}": sid for i, sid in enumerate(str_scope_ids)}
        if dts:
            def_params["ds"] = dts
        if dte:
            def_params["de"] = dte
        def_dc = f"reported_by IN ({str_ph})" if str_scope_ids else "1=0"

        rows = (await session.execute(
            text(f"SELECT * FROM defects WHERE {def_dc} AND {date_clause} LIMIT 1000"), def_params
        )).mappings().all()
        target_defects = [dict(r) for r in rows]

        ret_params: Dict[str, Any] = {f"s_{i}": sid for i, sid in enumerate(str_scope_ids)}
        if dts:
            ret_params["ds"] = dts
        if dte:
            ret_params["de"] = dte
        ret_dc = f"requested_by IN ({str_ph})" if str_scope_ids else "1=0"

        rows = (await session.execute(
            text(f"SELECT * FROM returns WHERE {ret_dc} AND {date_clause} LIMIT 1000"), ret_params
        )).mappings().all()
        target_returns = [dict(r) for r in rows]

        dist_ph1 = ",".join([f":df_{i}" for i in range(len(scope_ids))]) if scope_ids else "''"
        dist_ph2 = ",".join([f":dt_{i}" for i in range(len(scope_ids))]) if scope_ids else "''"
        dist_params: Dict[str, Any] = {}
        if scope_ids:
            for i, sid in enumerate(scope_ids):
                dist_params[f"df_{i}"] = sid
                dist_params[f"dt_{i}"] = sid
        if dts:
            dist_params["ds"] = dts
        if dte:
            dist_params["de"] = dte
        dist_dc = f"(from_user_id IN ({dist_ph1}) OR to_user_id IN ({dist_ph2}))" if scope_ids else "1=0"

        rows = (await session.execute(
            text(f"SELECT * FROM distributions WHERE {dist_dc} AND {date_clause}"), dist_params
        )).mappings().all()
        target_distributions = [dict(r) for r in rows]

    return {
        "user": {"id": target_user.get("id"), "name": target_user.get("name", ""), "role": target_user.get("role", "")},
        "stats": stats,
        "advanced": advanced,
        "devices": target_devices,
        "defects": target_defects,
        "returns": target_returns,
        "distributions": target_distributions,
        "users": target_users,
    }


async def generate_report(
    current_user: Dict[str, Any],
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
) -> Dict[str, Any]:
    role = current_user.get("role")
    user_id = str(current_user.get("_id", current_user.get("id", "")))
    scope_root_id = _resolve_scope_root_for_sub_distribution_manager(current_user, user_id)

    async with async_session_factory() as session:
        date_cond, date_params_tup = _build_date_filter("1=1", {}, start_date, end_date)

        scope_ids: List[str] = []
        if role not in ["super_admin", "md_director", "manager", "pdic_staff"]:
            scoped_ids = sorted({scope_root_id} | await _get_descendant_user_ids(session, scope_root_id))
            scope_ids = [str(s) for s in scoped_ids]

        device_stats = await device_service.get_device_stats(start_date, end_date)

        cursor = await session.execute(
            text(f"""SELECT
                   COALESCE(NULLIF(TRIM(device_type), ''), 'Unknown') AS device_type,
                   COUNT(*) AS total
                FROM devices
                WHERE {date_cond}
                GROUP BY COALESCE(NULLIF(TRIM(device_type), ''), 'Unknown')
                ORDER BY total DESC"""),
            date_params_tup
        )
        all_devices_by_type = {(r["device_type"]): int(r["total"]) for r in cursor.mappings().all()}

        cursor = await session.execute(
            text(f"""SELECT
                   COALESCE(NULLIF(TRIM(device_type), ''), 'Unknown') AS device_type,
                   COUNT(*) AS total
                FROM devices
                WHERE status IN ('distributed', 'in_use') AND {date_cond}
                GROUP BY COALESCE(NULLIF(TRIM(device_type), ''), 'Unknown')
                ORDER BY total DESC"""),
            date_params_tup
        )
        distributed_by_type = {(r["device_type"]): int(r["total"]) for r in cursor.mappings().all()}

        date_cond_d = date_cond.replace("created_at", "d.created_at")

        cursor = await session.execute(
            text(f"""SELECT
                   CAST(d.current_holder_id AS CHAR) AS holder_id,
                   COALESCE(NULLIF(TRIM(u.name), ''), 'Unknown') AS holder_name,
                   COUNT(*) AS total_sent
                FROM devices d
                LEFT JOIN users u ON CAST(d.current_holder_id AS UNSIGNED) = u.id
                WHERE d.status IN ('distributed', 'in_use') AND {date_cond_d}
                GROUP BY CAST(d.current_holder_id AS CHAR), COALESCE(NULLIF(TRIM(u.name), ''), 'Unknown')
                ORDER BY total_sent DESC"""),
            date_params_tup
        )
        subdistributor_rows = cursor.mappings().all()

        cursor = await session.execute(
            text(f"""SELECT *
                FROM devices
                WHERE {date_cond}
                ORDER BY id ASC"""),
            date_params_tup
        )
        all_device_rows = [dict(r) for r in cursor.mappings().all()]

        cursor = await session.execute(
            text(f"""SELECT d.*,
                       COALESCE(NULLIF(TRIM(u.name), ''), 'Unknown') AS holder_name
                FROM devices d
                LEFT JOIN users u ON CAST(d.current_holder_id AS UNSIGNED) = u.id
                WHERE d.status IN ('distributed', 'in_use') AND {date_cond_d}
                ORDER BY d.id ASC"""),
            date_params_tup
        )
        distributed_device_rows = [dict(r) for r in cursor.mappings().all()]

    total_distributed = sum(distributed_by_type.values())

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    date_label = "All Time"
    if start_date and end_date:
        date_label = f"{start_date[:10]} to {end_date[:10]}"
    elif start_date:
        date_label = f"From {start_date[:10]}"

    wb = Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.merge_cells("A1:D1")
    cell_title = ws1["A1"]
    cell_title.value = f"System Report - Generated {now_str}  |  Range: {date_label}"
    cell_title.font = Font(name="Calibri", size=14, bold=True)
    ws2 = wb.create_sheet("All Devices")
    ws3 = wb.create_sheet("Distributed Devices")
    ws4 = wb.create_sheet("Devices by Holder")
    ws5 = wb.create_sheet("Device Type Summary")

    for ws in [ws2, ws3, ws4, ws5]:
        ws["A1"] = f"System Report - Generated {now_str}  |  Range: {date_label}"

    summary_headers = ["Metric", "Value"]
    ws1.append(summary_headers)
    ws1.append(["Filter Period", date_label])
    ws1.append(["Total Devices (filtered)", device_stats.get("total", 0)])
    ws1.append(["Total Distributed/In Use", total_distributed])
    ws1.append(["Total Available (current)", device_stats.get("available", 0)])
    ws1.append(["Total Defective (filtered)", device_stats.get("defective", 0)])
    ws1.append([])
    ws1.append(["Distribution by Holder"])
    ws1.append(["Holder", "Total Devices Sent"])
    for row in subdistributor_rows:
        ws1.append([row["holder_name"], int(row["total_sent"])])

    # Sheet 2: All Devices
    all_headers = [
        "ID", "Device ID", "Serial Number", "MAC Address", "NUID",
        "Device Type", "Model", "Manufacturer", "Status",
        "Current Holder", "Current Location", "Created At"
    ]
    ws2.append(all_headers)
    for dev in all_device_rows:
        ws2.append([
            dev.get("id", ""),
            dev.get("device_id", ""),
            dev.get("serial_number", ""),
            dev.get("mac_address", ""),
            dev.get("nuid", ""),
            dev.get("device_type", ""),
            dev.get("model", ""),
            dev.get("manufacturer", ""),
            dev.get("status", ""),
            dev.get("current_holder_name", ""),
            dev.get("current_location", ""),
            dev.get("created_at", ""),
        ])

    # Sheet 3: Distributed Devices
    dist_headers = [
        "ID", "Device ID", "Serial Number", "MAC Address", "NUID",
        "Device Type", "Model", "Manufacturer",
        "Holder", "Current Location", "Status", "Created At"
    ]
    ws3.append(dist_headers)
    for dev in distributed_device_rows:
        ws3.append([
            dev.get("id", ""),
            dev.get("device_id", ""),
            dev.get("serial_number", ""),
            dev.get("mac_address", ""),
            dev.get("nuid", ""),
            dev.get("device_type", ""),
            dev.get("model", ""),
            dev.get("manufacturer", ""),
            dev.get("holder_name", ""),
            dev.get("current_location", ""),
            dev.get("status", ""),
            dev.get("created_at", ""),
        ])

    # Sheet 4: Devices by Holder
    sd_headers = ["Holder", "Total Devices Sent"]
    ws4.append(sd_headers)
    for row in subdistributor_rows:
        ws4.append([row["holder_name"], int(row["total_sent"])])

    # Sheet 5: Device Type Summary
    type_headers = ["Device Type", "Total (filtered)", "Distributed/In Use", "Remaining"]
    ws5.append(type_headers)
    all_types = sorted(set(list(all_devices_by_type.keys()) + list(distributed_by_type.keys())))
    for dt in all_types:
        total = all_devices_by_type.get(dt, 0)
        dist = distributed_by_type.get(dt, 0)
        remaining = total - dist
        ws5.append([dt, total, dist, remaining])

    payload = io.BytesIO()
    wb.save(payload)
    payload.seek(0)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    date_suffix = f"_{start_date[:10]}_{end_date[:10]}" if start_date and end_date else ""
    return {
        "content": payload.getvalue(),
        "filename": f"report{date_suffix}_{ts}.xlsx",
        "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
