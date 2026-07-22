import io
from datetime import datetime, timedelta, timezone
from typing import Dict, Any, List, Optional, Set

from openpyxl import Workbook
from openpyxl.styles import Font

from app.database import get_db, rows_to_list
from app.core.activity_logger import log_api_activity
from app.services import device_service, distribution_service, defect_service, return_service, user_service, approval_service, operator_service


ACTIVE_DEVICE_STATUSES = {"active", "available", "distributed", "in_use"}


def _build_date_filter(base_condition: str, base_params: tuple, start_date: Optional[str], end_date: Optional[str]) -> tuple:
    conds = [base_condition]
    params = list(base_params)
    if start_date:
        conds.append("created_at >= ?")
        params.append(start_date)
    if end_date:
        conds.append("created_at <= ?")
        params.append(end_date)
    return " AND ".join(conds), tuple(params)


async def _get_descendant_user_ids(db, root_user_id: str) -> Set[str]:
    descendants: Set[str] = set()
    if not root_user_id or not str(root_user_id).isdigit():
        return descendants

    pending: List[int] = [int(root_user_id)]
    visited: Set[int] = set()

    while pending:
        current_parent_id = pending.pop()
        if current_parent_id in visited:
            continue
        visited.add(current_parent_id)

        cursor = await db.execute(
            "SELECT id FROM users WHERE parent_id = ?",
            (current_parent_id,)
        )
        rows = await cursor.fetchall()
        for row in rows:
            child_id = int(row["id"])
            child_id_str = str(child_id)
            if child_id_str not in descendants:
                descendants.add(child_id_str)
                pending.append(child_id)

    return descendants


def _resolve_scope_root_for_sub_distribution_manager(user: Dict[str, Any], user_id: str) -> str:
    role = str(user.get("role") or "").lower()
    parent_id = str(user.get("parent_id") or "")
    if role == "sub_distribution_manager" and parent_id.isdigit():
        return parent_id
    return user_id


def _month_start(dt: datetime) -> datetime:
    return datetime(dt.year, dt.month, 1)


def _shift_months(dt: datetime, months: int) -> datetime:
    year = dt.year + (dt.month - 1 + months) // 12
    month = (dt.month - 1 + months) % 12 + 1
    return datetime(year, month, 1)


def _active_inactive_from_status_counts(status_counts: Dict[str, int]) -> Dict[str, int]:
    active = sum(int(total) for status, total in status_counts.items() if status in ACTIVE_DEVICE_STATUSES)
    total = sum(int(total) for total in status_counts.values())
    return {
        "active": int(active),
        "inactive": int(max(0, total - active)),
    }


async def _get_user_status_split_by_role(db, role: str, parent_id: Optional[str] = None) -> Dict[str, int]:
    query = """
        SELECT
            SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) AS active_total,
            SUM(CASE WHEN status != 'active' THEN 1 ELSE 0 END) AS inactive_total
        FROM users
        WHERE role = ?
    """
    params = [role]
    if parent_id is not None:
        query += " AND parent_id = ?"
        params.append(int(parent_id))

    cursor = await db.execute(query, tuple(params))
    row = await cursor.fetchone()
    return {
        "active": int((row[0] if row and row[0] is not None else 0)),
        "inactive": int((row[1] if row and row[1] is not None else 0)),
    }


async def _get_device_status_counts_for_holder(db, holder_id: str) -> Dict[str, int]:
    cursor = await db.execute(
        "SELECT status, COUNT(*) AS total FROM devices WHERE current_holder_id = ? GROUP BY status",
        (str(holder_id),)
    )
    rows = await cursor.fetchall()
    return {str(row[0]): int(row[1]) for row in rows}


async def get_dashboard_stats(user: Dict[str, Any],
                              start_date: Optional[str] = None,
                              end_date: Optional[str] = None) -> Dict[str, Any]:
    """Get dashboard statistics based on user role"""
    role = user.get("role")
    user_id = str(user.get("_id", user.get("id", "")))
    scope_root_id = _resolve_scope_root_for_sub_distribution_manager(user, user_id)

    stats = {}

    if role in ["super_admin", "md_director", "manager", "pdic_staff"]:
        total_stats = await device_service.get_device_stats()
        device_stats = await device_service.get_device_stats(start_date, end_date)
        dist_stats = await distribution_service.get_distribution_stats(start_date, end_date)
        defect_stats = await defect_service.get_defect_stats(start_date, end_date)
        return_stats = await return_service.get_return_stats(start_date, end_date)
        user_stats = await user_service.get_user_stats()
        approval_stats = await approval_service.get_approval_stats()

        async with get_db() as db:
            cond, prm = _build_date_filter("1=1", (), start_date, end_date)
            cursor = await db.execute(
                f"SELECT COUNT(*) FROM distributions WHERE {cond}", prm
            )
            distributions_filtered = (await cursor.fetchone())[0]

            cursor = await db.execute(
                f"SELECT COUNT(*) FROM defects WHERE replacement_device_id IS NOT NULL AND {cond}", prm
            )
            replacements_in_range = (await cursor.fetchone())[0]

            cursor = await db.execute(
                "SELECT COUNT(*) FROM defects WHERE replacement_device_id IS NOT NULL"
            )
            total_replacements = (await cursor.fetchone())[0]

        total_active = (
            total_stats.get("available", 0) +
            total_stats.get("distributed", 0) +
            total_stats.get("in_use", 0)
        )
        total_distributed = total_stats.get("distributed", 0) + total_stats.get("in_use", 0)
        total_inactive = max(0, total_stats.get("total", 0) - total_active)

        filtered_total = device_stats.get("total", 0)
        filtered_active = (
            device_stats.get("available", 0) +
            device_stats.get("distributed", 0) +
            device_stats.get("in_use", 0)
        )

        stats = {
            "total_devices": total_stats.get("total", 0),
            "total_active_devices": total_active,
            "total_distributed_devices": total_distributed,
            "total_inactive_devices": total_inactive,
            "total_defective_devices": total_stats.get("defective", 0),
            "total_replaced_devices": total_replacements,
            "registered_in_range": filtered_total,
            "distributed_in_range": device_stats.get("distributed", 0) + device_stats.get("in_use", 0),
            "inactive_in_range": max(0, filtered_total - filtered_active),
            "defective_in_range": device_stats.get("defective", 0),
            "replacements_in_range": replacements_in_range,
            "available_devices": device_stats.get("available", 0),
            "in_use_devices": device_stats.get("in_use", 0),
            "defective_devices": device_stats.get("defective", 0),
            "returned_devices": device_stats.get("returned", 0),
            "active_devices": filtered_active,
            "distributed_devices": device_stats.get("distributed", 0) + device_stats.get("in_use", 0),
            "total_distributions": dist_stats.get("total", 0),
            "pending_distributions": dist_stats.get("pending", 0),
            "approved_distributions": dist_stats.get("approved", 0),
            "delivered_distributions": dist_stats.get("delivered", 0),
            "rejected_distributions": dist_stats.get("rejected", 0),
            "distribution_this_month": distributions_filtered,
            "total_defects": defect_stats.get("total", 0),
            "defect_reports": defect_stats.get("total", 0),
            "reported_defects": defect_stats.get("by_status", {}).get("reported", 0),
            "under_review_defects": defect_stats.get("by_status", {}).get("under_review", 0),
            "resolved_defects": defect_stats.get("by_status", {}).get("resolved", 0),
            "total_returns": return_stats.get("total", 0),
            "return_requests": return_stats.get("total", 0),
            "pending_returns": return_stats.get("by_status", {}).get("pending", 0),
            "approved_returns": return_stats.get("by_status", {}).get("approved", 0),
            "received_returns": return_stats.get("by_status", {}).get("received", 0),
            "rejected_returns": return_stats.get("by_status", {}).get("rejected", 0),
            "total_users": user_stats.get("total", 0),
            "active_users": user_stats.get("active", 0),
            "pending_approvals": approval_stats.get("total_pending", 0),
            "pending_receipts": dist_stats.get("pending_receipt", 0),
            "total_approved": approval_stats.get("approved", 0),
            "total_rejected": approval_stats.get("rejected", 0),
            "devices": device_stats,
            "distributions": dist_stats,
            "defects": defect_stats,
            "returns": return_stats,
            "users": user_stats,
            "approvals": approval_stats
        }

    elif role == "sub_distributor":
        async with get_db() as db:
            dc, dp = _build_date_filter("current_holder_id = ?", (user_id,), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM devices WHERE {dc}", dp)
            my_devices = (await cursor.fetchone())[0]

            ac, ap = _build_date_filter("current_holder_id = ? AND status = 'available'", (user_id,), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM devices WHERE {ac}", ap)
            available_devices = (await cursor.fetchone())[0]

            sc, sp = _build_date_filter("from_user_id = ?", (user_id,), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM distributions WHERE {sc}", sp)
            sent = (await cursor.fetchone())[0]

            rc, rp = _build_date_filter("to_user_id = ?", (user_id,), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM distributions WHERE {rc}", rp)
            received = (await cursor.fetchone())[0]

            pc, pp = _build_date_filter("from_user_id = ? AND status = 'pending'", (user_id,), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM distributions WHERE {pc}", pp)
            pending = (await cursor.fetchone())[0]

            cursor = await db.execute("SELECT id FROM users WHERE role = 'sub_distribution_manager' AND parent_id = ?", (int(user_id),))
            sub_dist_manager_ids = [row[0] for row in await cursor.fetchall()]
            candidate_cluster_parent_ids = [int(user_id)] + sub_dist_manager_ids
            if candidate_cluster_parent_ids:
                placeholders = ",".join("?" * len(candidate_cluster_parent_ids))
                cursor = await db.execute(
                    f"SELECT id FROM users WHERE role = 'cluster' AND parent_id IN ({placeholders})",
                    tuple(candidate_cluster_parent_ids)
                )
                cluster_ids = [row[0] for row in await cursor.fetchall()]
            else:
                cluster_ids = []

            if cluster_ids:
                placeholders = ",".join("?" * len(cluster_ids))
                cursor = await db.execute(
                    f"SELECT COUNT(*) FROM users WHERE role = 'operator' AND parent_id IN ({placeholders})",
                    tuple(cluster_ids)
                )
                operator_count = (await cursor.fetchone())[0]
            else:
                operator_count = 0

        stats = {
            "my_devices": my_devices,
            "received_devices": my_devices,
            "available_devices": available_devices,
            "distributions_sent": sent,
            "distributions_received": received,
            "pending_distributions": pending,
            "operator_count": operator_count,
        }

    elif role == "sub_distribution_manager":
        async with get_db() as db:
            scope_ids = sorted({scope_root_id} | await _get_descendant_user_ids(db, scope_root_id))
            placeholders = ",".join(["?"] * len(scope_ids)) if scope_ids else "?"

            dc, dp = _build_date_filter(f"current_holder_id IN ({placeholders})", tuple(scope_ids), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM devices WHERE {dc}", dp)
            branch_devices = (await cursor.fetchone())[0]

            ac, ap = _build_date_filter(f"current_holder_id IN ({placeholders}) AND status = 'available'", tuple(scope_ids), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM devices WHERE {ac}", ap)
            available_devices = (await cursor.fetchone())[0]

            sc, sp = _build_date_filter(f"from_user_id IN ({placeholders})", tuple(scope_ids), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM distributions WHERE {sc}", sp)
            sent = (await cursor.fetchone())[0]

            rc, rp = _build_date_filter(f"to_user_id IN ({placeholders})", tuple(scope_ids), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM distributions WHERE {rc}", rp)
            received = (await cursor.fetchone())[0]

            cursor = await db.execute(
                f"SELECT COUNT(*) FROM distributions WHERE to_user_id = ? AND status = 'pending_receipt'",
                (user_id,)
            )
            pending = (await cursor.fetchone())[0]

            cursor = await db.execute(
                f"SELECT COUNT(*) FROM users WHERE role = 'operator' AND id IN ({placeholders})",
                tuple(scope_ids)
            )
            operator_count = (await cursor.fetchone())[0]

            dec, dep = _build_date_filter(f"CAST(reported_by AS TEXT) IN ({placeholders})", tuple(scope_ids), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM defects WHERE {dec}", dep)
            defect_reports = (await cursor.fetchone())[0]

            rec, rep = _build_date_filter(f"CAST(requested_by AS TEXT) IN ({placeholders})", tuple(scope_ids), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM returns WHERE {rec}", rep)
            return_requests = (await cursor.fetchone())[0]

        stats = {
            "my_devices": branch_devices,
            "received_devices": branch_devices,
            "available_devices": available_devices,
            "distributions_sent": sent,
            "distributions_received": received,
            "pending_distributions": pending,
            "operator_count": operator_count,
            "defect_reports": defect_reports,
            "return_requests": return_requests,
            "assigned_to_operators": sent,
        }

    elif role == "cluster":
        async with get_db() as db:
            dc, dp = _build_date_filter("current_holder_id = ?", (user_id,), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM devices WHERE {dc}", dp)
            my_devices = (await cursor.fetchone())[0]

            sc, sp = _build_date_filter("from_user_id = ?", (user_id,), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM distributions WHERE {sc}", sp)
            sent = (await cursor.fetchone())[0]

            rc, rp = _build_date_filter("to_user_id = ?", (user_id,), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM distributions WHERE {rc}", rp)
            received = (await cursor.fetchone())[0]
        operator_stats_data = await operator_service.get_operator_stats(user_id)
        stats = {
            "my_devices": my_devices,
            "operators": operator_stats_data,
            "distributions_sent": sent,
            "distributions_received": received
        }

    elif role == "operator":
        async with get_db() as db:
            dc, dp = _build_date_filter("current_holder_id = ?", (user_id,), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM devices WHERE {dc}", dp)
            my_devices = (await cursor.fetchone())[0]

            dfc, dfp = _build_date_filter("reported_by = ?", (user_id,), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM defects WHERE {dfc}", dfp)
            my_defects = (await cursor.fetchone())[0]

            rc, rp = _build_date_filter("requested_by = ?", (user_id,), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM returns WHERE {rc}", rp)
            my_returns = (await cursor.fetchone())[0]
        stats = {
            "my_devices": my_devices,
            "my_defects": my_defects,
            "my_returns": my_returns
        }

    return stats


async def get_recent_activities(user: Dict[str, Any], limit: int = 10) -> list:
    """Get recent activities based on user role"""
    role = user.get("role")
    user_id = str(user.get("_id", user.get("id", "")))

    activities = []

    async with get_db() as db:
        if role in ["super_admin", "md_director", "manager", "pdic_staff"]:
            cursor = await db.execute(
                "SELECT * FROM device_history ORDER BY timestamp DESC LIMIT ?", (limit,)
            )
        else:
            cursor = await db.execute(
                """SELECT * FROM device_history
                WHERE performed_by = ? OR from_user_id = ? OR to_user_id = ?
                ORDER BY timestamp DESC LIMIT ?""",
                (user_id, user_id, user_id, limit)
            )
        rows = await cursor.fetchall()

        for h in rows:
            hd = dict(h)
            activities.append({
                "id": str(hd["id"]),
                "action": hd["action"],
                "description": f"{hd.get('performed_by_name', 'Unknown')} {hd['action']} device",
                "user_name": hd.get("performed_by_name", "Unknown"),
                "timestamp": hd["timestamp"],
                "category": "device",
                "link": None
            })

    return activities


async def get_admin_activities(
    page: int = 1,
    page_size: int = 50,
    actor: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> Dict[str, Any]:
    """Get unified activity stream for admin users with filtering."""
    normalized_category = (category or "all").strip().lower()
    activities: List[Dict[str, Any]] = []

    async with get_db() as db:
        if normalized_category in {"all", "device"}:
            conditions = ["1=1"]
            params: List[Any] = []

            # Keep bulk device uploads as one summary entry in API activity logs.
            conditions.append("action != ?")
            params.append("bulk_registered")

            # Replacement workflow already writes a curated API business activity.
            # Hide matching device-history rows so replacement appears as a single activity.
            conditions.append("(notes IS NULL OR notes NOT LIKE ?)")
            params.append("Device replaced by % for defect %")
            conditions.append("(notes IS NULL OR notes NOT LIKE ?)")
            params.append("Device serviced and reassigned for defect %")

            if actor:
                conditions.append("performed_by_name LIKE ?")
                params.append(f"%{actor}%")

            if search:
                like = f"%{search}%"
                conditions.append("(action LIKE ? OR notes LIKE ? OR device_id LIKE ? OR performed_by_name LIKE ?)")
                params.extend([like, like, like, like])

            if start_date:
                conditions.append("timestamp >= ?")
                params.append(start_date)

            if end_date:
                conditions.append("timestamp <= ?")
                params.append(end_date)

            where_clause = " AND ".join(conditions)
            cursor = await db.execute(
                f"""SELECT id, device_id, action, notes, performed_by_name, timestamp
                    FROM device_history
                    WHERE {where_clause}
                    ORDER BY timestamp DESC""",
                params,
            )
            rows = await cursor.fetchall()
            for row in rows:
                item = dict(row)
                actor_name = item.get("performed_by_name") or "Unknown"
                description = (
                    item.get("notes")
                    or f"{item.get('action', 'updated')} on device {item.get('device_id', '-')}."
                )
                activities.append(
                    {
                        "id": f"device-{item.get('id')}",
                        "category": "device",
                        "action": item.get("action") or "device_update",
                        "actor": actor_name,
                        "description": description,
                        "date": item.get("timestamp"),
                        "link": None,
                    }
                )

        if normalized_category in {"all", "inventory"}:
            conditions = ["1=1"]
            params = []

            if actor:
                conditions.append("performed_by_name LIKE ?")
                params.append(f"%{actor}%")

            if search:
                like = f"%{search}%"
                conditions.append("(movement_type LIKE ? OR notes LIKE ? OR item_sku LIKE ? OR item_name LIKE ? OR performed_by_name LIKE ?)")
                params.extend([like, like, like, like, like])

            if start_date:
                conditions.append("created_at >= ?")
                params.append(start_date)

            if end_date:
                conditions.append("created_at <= ?")
                params.append(end_date)

            where_clause = " AND ".join(conditions)
            cursor = await db.execute(
                f"""SELECT id, item_sku, item_name, movement_type, notes, performed_by_name, created_at
                    FROM inventory_stock_movements
                    WHERE {where_clause}
                    ORDER BY created_at DESC""",
                params,
            )
            rows = await cursor.fetchall()
            for row in rows:
                item = dict(row)
                actor_name = item.get("performed_by_name") or "Unknown"
                description = (
                    item.get("notes")
                    or f"{item.get('movement_type', 'movement')} for {item.get('item_sku') or item.get('item_name') or '-'}."
                )
                activities.append(
                    {
                        "id": f"inventory-{item.get('id')}",
                        "category": "inventory",
                        "action": item.get("movement_type") or "movement",
                        "actor": actor_name,
                        "description": description,
                        "date": item.get("created_at"),
                        "link": None,
                    }
                )

        if normalized_category in {"all", "api"}:
            conditions = ["1=1"]
            params = []

            # Hide legacy generic middleware rows and surface only curated business-action API logs.
            conditions.append("description NOT LIKE ?")
            params.append("% returned %")

            if actor:
                conditions.append("actor_name LIKE ?")
                params.append(f"%{actor}%")

            if search:
                like = f"%{search}%"
                conditions.append("(description LIKE ? OR path LIKE ? OR method LIKE ? OR actor_name LIKE ?)")
                params.extend([like, like, like, like])

            if start_date:
                conditions.append("created_at >= ?")
                params.append(start_date)

            if end_date:
                conditions.append("created_at <= ?")
                params.append(end_date)

            where_clause = " AND ".join(conditions)
            cursor = await db.execute(
                f"""SELECT id, actor_name, method, path, status_code, description, created_at
                    FROM api_activity_logs
                    WHERE {where_clause}
                    ORDER BY created_at DESC""",
                params,
            )
            rows = await cursor.fetchall()
            for row in rows:
                item = dict(row)
                path_value = str(item.get("path") or "")

                link = None
                if path_value.startswith("/activity/devices"):
                    link = "/devices"
                elif path_value.startswith("/activity/distributions"):
                    link = "/distributions"
                elif path_value.startswith("/activity/users"):
                    link = "/users"
                elif path_value.startswith("/activity/defects"):
                    link = "/defects"
                elif path_value.startswith("/activity/returns"):
                    link = "/returns"
                elif path_value.startswith("/activity/pending-dues"):
                    link = "/defects"
                elif path_value.startswith("/activity/reports"):
                    link = "/backup"
                elif path_value.startswith("/activity/external-inventory"):
                    link = "/external-inventory"

                activities.append(
                    {
                        "id": f"api-{item.get('id')}",
                        "category": "api",
                        "action": f"{item.get('method', 'API')} {item.get('path', '')}",
                        "actor": item.get("actor_name") or "Anonymous",
                        "description": item.get("description") or "API activity",
                        "date": item.get("created_at"),
                        "link": link,
                    }
                )

    activities.sort(key=lambda x: str(x.get("date") or ""), reverse=True)
    total = len(activities)
    start_idx = max(0, (page - 1) * page_size)
    end_idx = start_idx + page_size
    paged = activities[start_idx:end_idx]

    return {
        "data": paged,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": ((total + page_size - 1) // page_size) if page_size > 0 else 0,
        },
    }


async def track_client_activity(
    user: Dict[str, Any],
    action: str,
    description: str,
    context: Optional[str] = None,
) -> None:
    """Persist explicit client-side activity events (for UI-only actions)."""
    user_id = str(user.get("id") or user.get("_id") or user.get("user_id") or user.get("sub") or "")
    actor_name = str(user.get("name") or user.get("email") or "Unknown")
    actor_role = str(user.get("role") or "")
    normalized_action = str(action or "ui_action").strip() or "ui_action"
    final_description = str(description or "User action").strip() or "User action"
    path = f"/ui/{normalized_action}"
    if context:
        path = f"{path}/{str(context).strip()[:128]}"

    await log_api_activity(
        method="UI",
        path=path,
        status_code=200,
        actor_id=user_id,
        actor_name=actor_name,
        actor_role=actor_role,
        description=final_description,
    )


async def get_distribution_chart_data(start_date: Optional[str] = None,
                                      end_date: Optional[str] = None) -> list:
    """Get distribution data for charts"""
    data = []
    now = datetime.now().replace(tzinfo=None)

    async with get_db() as db:
        if start_date or end_date:
            cond, prm = _build_date_filter("status = 'delivered'", (), start_date, end_date)
            cursor = await db.execute(
                f"SELECT COUNT(*) FROM distributions WHERE {cond}", prm
            )
            total = (await cursor.fetchone())[0]
            data.append({
                "month": "Filtered",
                "distributions": total
            })
        else:
            for i in range(11, -1, -1):
                month_start = datetime(now.year, now.month, 1) - timedelta(days=i * 30)
                month_end = month_start + timedelta(days=30)

                cursor = await db.execute(
                    "SELECT COUNT(*) FROM distributions WHERE status = 'delivered' AND created_at >= ? AND created_at < ?",
                    (month_start.isoformat(), month_end.isoformat())
                )
                count = (await cursor.fetchone())[0]

                data.append({
                    "month": month_start.strftime("%b"),
                    "distributions": count
                })

    return data


async def get_defect_chart_data(start_date: Optional[str] = None,
                                end_date: Optional[str] = None) -> list:
    """Get defect data for charts"""
    data = []
    now = datetime.now().replace(tzinfo=None)

    async with get_db() as db:
        if start_date or end_date:
            cond, prm = _build_date_filter("1=1", (), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM defects WHERE {cond}", prm)
            reported = (await cursor.fetchone())[0]

            resolved_conds = ["status = 'resolved'"]
            resolved_params = []
            if start_date:
                resolved_conds.append("resolved_at >= ?")
                resolved_params.append(start_date)
            if end_date:
                resolved_conds.append("resolved_at <= ?")
                resolved_params.append(end_date)
            resolved_where = " AND ".join(resolved_conds)
            cursor = await db.execute(f"SELECT COUNT(*) FROM defects WHERE {resolved_where}", tuple(resolved_params))
            resolved = (await cursor.fetchone())[0]

            data.append({
                "month": "Filtered",
                "reported": reported,
                "resolved": resolved
            })
        else:
            for i in range(11, -1, -1):
                month_start = datetime(now.year, now.month, 1) - timedelta(days=i * 30)
                month_end = month_start + timedelta(days=30)

                cursor = await db.execute(
                    "SELECT COUNT(*) FROM defects WHERE created_at >= ? AND created_at < ?",
                    (month_start.isoformat(), month_end.isoformat())
                )
                reported = (await cursor.fetchone())[0]

                cursor = await db.execute(
                    "SELECT COUNT(*) FROM defects WHERE status = 'resolved' AND resolved_at >= ? AND resolved_at < ?",
                    (month_start.isoformat(), month_end.isoformat())
                )
                resolved = (await cursor.fetchone())[0]

                data.append({
                    "month": month_start.strftime("%b"),
                    "reported": reported,
                    "resolved": resolved
                })

    return data


async def get_scope_users(user: Dict[str, Any]) -> Dict[str, list]:
    """Return sub_distributors, clusters, and operators visible to the current user."""
    role = user.get("role")
    user_id = str(user.get("_id", user.get("id", "")))
    scope_root_id = _resolve_scope_root_for_sub_distribution_manager(user, user_id)

    async with get_db() as db:
        if role == "super_admin":
            cursor = await db.execute(
                "SELECT id, email, name, role, COALESCE(parent_id, '') AS parent_id FROM users WHERE role IN ('sub_distributor', 'cluster', 'operator') ORDER BY role, name"
            )
        elif role in ("manager", "md_director", "pdic_staff"):
            scope_ids = sorted({scope_root_id} | await _get_descendant_user_ids(db, scope_root_id))
            placeholders = ",".join(["?"] * len(scope_ids)) if scope_ids else "?"
            cursor = await db.execute(
                f"SELECT id, email, name, role, COALESCE(parent_id, '') AS parent_id FROM users WHERE role IN ('sub_distributor', 'cluster', 'operator') AND id IN ({placeholders}) ORDER BY role, name",
                tuple(scope_ids)
            )
        elif role == "sub_distributor":
            cursor = await db.execute(
                "SELECT id, email, name, role, COALESCE(parent_id, '') AS parent_id FROM users WHERE (role = 'cluster' AND parent_id = ?) OR (role = 'operator' AND parent_id IN (SELECT id FROM users WHERE role = 'cluster' AND parent_id = ?)) ORDER BY role, name",
                (int(user_id), int(user_id))
            )
        elif role == "cluster":
            cursor = await db.execute(
                "SELECT id, email, name, role, COALESCE(parent_id, '') AS parent_id FROM users WHERE role = 'operator' AND parent_id = ? ORDER BY name",
                (int(user_id),)
            )
        elif role == "sub_distribution_manager":
            scope_ids = sorted({scope_root_id} | await _get_descendant_user_ids(db, scope_root_id))
            placeholders = ",".join(["?"] * len(scope_ids)) if scope_ids else "?"
            cursor = await db.execute(
                f"SELECT id, email, name, role, COALESCE(parent_id, '') AS parent_id FROM users WHERE role IN ('sub_distributor', 'cluster', 'operator') AND id IN ({placeholders}) ORDER BY role, name",
                tuple(scope_ids)
            )
        else:
            return {"sub_distributors": [], "clusters": [], "operators": []}

        rows = await cursor.fetchall()
        users_list = [dict(r) for r in rows]
        for u in users_list:
            u["id"] = str(u["id"])
            u["parent_id"] = str(u["parent_id"]) if u["parent_id"] else ""

    role_map = {
        "sub_distributor": "sub_distributors",
        "cluster": "clusters",
        "operator": "operators",
    }
    result = {"sub_distributors": [], "clusters": [], "operators": []}
    for u in users_list:
        mapped_key = role_map.get(u["role"])
        if mapped_key:
            result[mapped_key].append(u)
    return result


async def get_user_kpi(current_user: Dict[str, Any], target_user_id: str,
                       start_date: Optional[str] = None,
                       end_date: Optional[str] = None) -> Dict[str, Any]:
    """Get KPI data for a specific user (devices, distributions, defects)."""
    empty = {"user": {}, "kpis": {}, "charts": {}}

    async with get_db() as db:
        cursor = await db.execute(
            "SELECT id, email, name, role FROM users WHERE id = ?",
            (int(target_user_id),)
        )
        target_user = await cursor.fetchone()
        if not target_user:
            return empty

        target_user = dict(target_user)
        target_user["id"] = str(target_user["id"])

        # Devices held by this user
        dc, dp = _build_date_filter("current_holder_id = ?", (target_user["id"],), start_date, end_date)
        cursor = await db.execute(f"SELECT COUNT(*) FROM devices WHERE {dc}", dp)
        devices_in_hand = (await cursor.fetchone())[0]

        # Devices in hierarchy (descendants)
        descendant_ids = sorted({target_user["id"]} | await _get_descendant_user_ids(db, target_user["id"]))
        placeholders = ",".join(["?"] * len(descendant_ids)) if descendant_ids else "?"
        hc, hp = _build_date_filter(f"current_holder_id IN ({placeholders})", tuple(descendant_ids), start_date, end_date)
        cursor = await db.execute(f"SELECT COUNT(*) FROM devices WHERE {hc}", hp)
        devices_in_hierarchy = (await cursor.fetchone())[0]

        # Active/inactive in hierarchy
        cursor = await db.execute(
            f"""SELECT status, COUNT(*) AS total FROM devices
                WHERE {hc}
                GROUP BY status""",
            hp
        )
        device_status_rows = await cursor.fetchall()
        device_status = {str(r[0]): int(r[1]) for r in device_status_rows}
        hierarchy_active = sum(
            v for k, v in device_status.items() if k in ACTIVE_DEVICE_STATUSES
        )
        hierarchy_inactive = max(0, devices_in_hierarchy - hierarchy_active)

        # Distributions where this user is the sender
        dc2, dp2 = _build_date_filter("from_user_id = ?", (target_user["id"],), start_date, end_date)
        cursor = await db.execute(f"SELECT COUNT(*) FROM distributions WHERE {dc2}", dp2)
        distributed_count = (await cursor.fetchone())[0]

        # Defects reported by this user
        dfc, dfp = _build_date_filter("CAST(reported_by AS CHAR) = ?", (target_user["id"],), start_date, end_date)
        cursor = await db.execute(f"SELECT COUNT(*) FROM defects WHERE {dfc}", dfp)
        total_defects = (await cursor.fetchone())[0]

        # Defect trend (12 months)
        now = datetime.now().replace(tzinfo=None)
        month_start = _month_start(now)
        defect_trend = []
        trend_range = range(11, -1, -1) if not (start_date or end_date) else range(0, 0)
        for i in trend_range:
            start = _shift_months(month_start, -i)
            end = _shift_months(start, 1)
            cursor = await db.execute(
                "SELECT COUNT(*) FROM defects WHERE CAST(reported_by AS CHAR) = ? AND created_at >= ? AND created_at < ?",
                (target_user["id"], start.isoformat(), end.isoformat())
            )
            reported = (await cursor.fetchone())[0]
            cursor = await db.execute(
                "SELECT COUNT(*) FROM defects WHERE CAST(reported_by AS CHAR) = ? AND status = 'resolved' AND resolved_at >= ? AND resolved_at < ?",
                (target_user["id"], start.isoformat(), end.isoformat())
            )
            resolved = (await cursor.fetchone())[0]
            defect_trend.append({
                "month": start.strftime("%b"),
                "reported": reported,
                "resolved": resolved,
            })

        # Distribution trend (12 months)
        distribution_trend = []
        for i in trend_range:
            start = _shift_months(month_start, -i)
            end = _shift_months(start, 1)
            cursor = await db.execute(
                "SELECT COUNT(*) FROM distributions WHERE from_user_id = ? AND created_at >= ? AND created_at < ?",
                (target_user["id"], start.isoformat(), end.isoformat())
            )
            total = (await cursor.fetchone())[0]
            cursor = await db.execute(
                "SELECT COUNT(*) FROM distributions WHERE from_user_id = ? AND status = 'delivered' AND created_at >= ? AND created_at < ?",
                (target_user["id"], start.isoformat(), end.isoformat())
            )
            delivered = (await cursor.fetchone())[0]
            distribution_trend.append({
                "month": start.strftime("%b"),
                "total": total,
                "delivered": delivered,
            })

    return {
        "user": {
            "id": target_user["id"],
            "name": target_user["name"],
            "role": target_user["role"],
            "email": target_user["email"],
        },
        "kpis": {
            "devices_in_hand": devices_in_hand,
            "devices_in_hierarchy": devices_in_hierarchy,
            "hierarchy_active_devices": hierarchy_active,
            "hierarchy_inactive_devices": hierarchy_inactive,
            "distributed_count": distributed_count,
            "total_defects": total_defects,
        },
        "charts": {
            "device_status": device_status,
            "defect_trend_12m": defect_trend,
            "distribution_trend_12m": distribution_trend,
        },
    }


async def get_system_alerts(user: Dict[str, Any]) -> list:
    """Get system alerts for dashboard"""
    role = user.get("role")
    alerts = []

    if role in ["super_admin", "md_director", "manager", "pdic_staff"]:
        async with get_db() as db:
            cursor = await db.execute(
                "SELECT COUNT(*) FROM defects WHERE severity = 'critical' AND status != 'resolved'"
            )
            critical_defects = (await cursor.fetchone())[0]
            if critical_defects > 0:
                alerts.append({
                    "type": "error",
                    "title": "Critical Defects",
                    "message": f"{critical_defects} critical defect(s) require attention",
                    "link": "/defects?severity=critical"
                })

            cursor = await db.execute("SELECT COUNT(*) FROM approvals WHERE status = 'pending'")
            pending_approvals = (await cursor.fetchone())[0]
            if pending_approvals > 0:
                alerts.append({
                    "type": "warning",
                    "title": "Pending",
                    "message": f"{pending_approvals} request(s) waiting for approval with Pending payments",
                    "link": "/payments"
                })

            cursor = await db.execute("SELECT COUNT(*) FROM devices WHERE status = 'available'")
            available_devices = (await cursor.fetchone())[0]
            if available_devices < 10:
                alerts.append({
                    "type": "warning",
                    "title": "Low Device Stock",
                    "message": f"Only {available_devices} devices available in stock",
                    "link": "/devices"
                })

    return alerts


async def get_advanced_dashboard_metrics(user: Dict[str, Any],
                                         start_date: Optional[str] = None,
                                         end_date: Optional[str] = None) -> Dict[str, Any]:
    """Get advanced analytics payload for management dashboards."""
    role = user.get("role")
    user_id = str(user.get("_id", user.get("id", "")))
    scope_root_id = _resolve_scope_root_for_sub_distribution_manager(user, user_id)

    if role not in ["super_admin", "md_director", "manager", "pdic_staff", "sub_distribution_manager", "sub_distributor", "cluster", "operator"]:
        return {"kpis": {}, "charts": {}, "alerts": [], "reliability": {"summary": {}, "trend": []}}

    # Role-scoped advanced payload for non-management dashboards.
    if role in ["sub_distribution_manager", "sub_distributor", "cluster", "operator"]:
        async with get_db() as db:
            if role == "sub_distribution_manager":
                scope_ids = sorted({scope_root_id} | await _get_descendant_user_ids(db, scope_root_id))
                placeholders = ",".join(["?"] * len(scope_ids)) if scope_ids else "?"
                cursor = await db.execute(
                    f"SELECT status, COUNT(*) AS total FROM devices WHERE current_holder_id IN ({placeholders}) GROUP BY status",
                    tuple(scope_ids)
                )
                rows = await cursor.fetchall()
                my_device_status = {str(row[0]): int(row[1]) for row in rows}
            else:
                my_device_status = await _get_device_status_counts_for_holder(db, user_id)
            my_device_active_split = _active_inactive_from_status_counts(my_device_status)

            charts = {
                "my_device_status": my_device_status,
                "my_device_active_split": my_device_active_split,
            }
            kpis = {
                "my_total_devices": int(sum(my_device_status.values())),
                "my_active_devices": int(my_device_active_split.get("active", 0)),
                "my_inactive_devices": int(my_device_active_split.get("inactive", 0)),
            }

            if role in ["sub_distribution_manager", "sub_distributor"]:
                if role == "sub_distribution_manager":
                    cursor = await db.execute(
                                                """
                                                SELECT id
                                                FROM users
                                                WHERE role = 'cluster'
                                                    AND (
                                                        parent_id = ?
                                                        OR parent_id IN (
                                                            SELECT id FROM users WHERE role = 'sub_distribution_manager' AND parent_id = ?
                                                        )
                                                    )
                                                """,
                        (int(scope_root_id), int(scope_root_id))
                    )
                    cluster_rows = await cursor.fetchall()
                    cluster_ids = [int(row[0]) for row in cluster_rows]
                    if cluster_ids:
                        placeholders = ",".join("?" * len(cluster_ids))
                        cursor = await db.execute(
                            f"""
                            SELECT
                                SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) AS active_total,
                                SUM(CASE WHEN status != 'active' THEN 1 ELSE 0 END) AS inactive_total
                            FROM users
                            WHERE role = 'cluster' AND id IN ({placeholders})
                            """,
                            tuple(cluster_ids)
                        )
                        cs_row = await cursor.fetchone()
                        cluster_status_split = {
                            "active": int((cs_row[0] if cs_row and cs_row[0] is not None else 0)),
                            "inactive": int((cs_row[1] if cs_row and cs_row[1] is not None else 0)),
                        }
                    else:
                        cluster_status_split = {"active": 0, "inactive": 0}
                else:
                    cluster_status_split = await _get_user_status_split_by_role(db, "cluster", user_id)
                    cursor = await db.execute(
                        "SELECT id FROM users WHERE role = 'cluster' AND parent_id = ?",
                        (int(user_id),)
                    )
                    cluster_rows = await cursor.fetchall()
                    cluster_ids = [int(row[0]) for row in cluster_rows]

                if cluster_ids:
                    placeholders = ",".join("?" * len(cluster_ids))
                    cursor = await db.execute(
                        f"""
                        SELECT
                            SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) AS active_total,
                            SUM(CASE WHEN status != 'active' THEN 1 ELSE 0 END) AS inactive_total
                        FROM users
                        WHERE role = 'operator' AND parent_id IN ({placeholders})
                        """,
                        tuple(cluster_ids)
                    )
                    op_row = await cursor.fetchone()
                    operator_status_split = {
                        "active": int((op_row[0] if op_row and op_row[0] is not None else 0)),
                        "inactive": int((op_row[1] if op_row and op_row[1] is not None else 0)),
                    }
                else:
                    operator_status_split = {"active": 0, "inactive": 0}

                charts["cluster_account_active_split"] = cluster_status_split
                charts["operator_account_active_split"] = operator_status_split
                kpis["my_total_clusters"] = int(cluster_status_split["active"] + cluster_status_split["inactive"])
                kpis["my_total_operators"] = int(operator_status_split["active"] + operator_status_split["inactive"])

            elif role == "cluster":
                operator_status_split = await _get_user_status_split_by_role(db, "operator", user_id)
                charts["operator_account_active_split"] = operator_status_split
                kpis["my_total_operators"] = int(operator_status_split["active"] + operator_status_split["inactive"])

            elif role == "operator":
                is_active = int(user.get("status", "active") == "active")
                charts["operator_account_active_split"] = {
                    "active": is_active,
                    "inactive": 0 if is_active else 1,
                }

        return {
            "kpis": kpis,
            "charts": charts,
            "alerts": [],
            "reliability": {"summary": {}, "trend": []},
        }

    now = datetime.now().replace(tzinfo=None)
    month_start = _month_start(now)
    year_start = datetime(now.year, 1, 1)

    effective_start = start_date or month_start.isoformat()
    effective_year_start = start_date or year_start.isoformat()

    device_stats = await device_service.get_device_stats(start_date, end_date)
    user_stats = await user_service.get_user_stats()
    defect_stats = await defect_service.get_defect_stats(start_date, end_date)
    return_stats = await return_service.get_return_stats(start_date, end_date)
    dist_stats = await distribution_service.get_distribution_stats(start_date, end_date)
    approval_stats = await approval_service.get_approval_stats()
    alerts = await get_system_alerts(user)

    async with get_db() as db:
        # Defect month/year totals (respect range if provided)
        dc_m, dp_m = _build_date_filter("1=1", (), effective_start, end_date)
        cursor = await db.execute(f"SELECT COUNT(*) FROM defects WHERE {dc_m}", dp_m)
        defects_this_month = (await cursor.fetchone())[0]

        dc_y, dp_y = _build_date_filter("1=1", (), effective_year_start, end_date)
        cursor = await db.execute(f"SELECT COUNT(*) FROM defects WHERE {dc_y}", dp_y)
        defects_this_year = (await cursor.fetchone())[0]

        # Replacement metrics (current state - no date filter)
        cursor = await db.execute(
            "SELECT COUNT(*) FROM defects WHERE replacement_device_id IS NOT NULL"
        )
        replacements_total = (await cursor.fetchone())[0]

        cursor = await db.execute(
            """SELECT COUNT(*) FROM defects
               WHERE replacement_device_id IS NOT NULL
               AND replacement_confirmed_at IS NOT NULL"""
        )
        replacements_confirmed = (await cursor.fetchone())[0]

        cursor = await db.execute(
            """SELECT COUNT(*) FROM defects
               WHERE replacement_device_id IS NOT NULL
               AND replacement_confirmed_at IS NULL"""
        )
        replacements_pending = (await cursor.fetchone())[0]

        # Role totals (current state - no date filter)
        cursor = await db.execute(
            "SELECT role, COUNT(*) AS total FROM users GROUP BY role"
        )
        role_rows = await cursor.fetchall()
        role_counts = {str(r[0]): int(r[1]) for r in role_rows}

        cursor = await db.execute(
            """
            SELECT
                role,
                SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) AS active_total,
                SUM(CASE WHEN status != 'active' THEN 1 ELSE 0 END) AS inactive_total
            FROM users
            WHERE role IN ('sub_distributor', 'cluster', 'operator')
            GROUP BY role
            """
        )
        role_status_rows = await cursor.fetchall()
        role_status_splits = {
            str(row[0]): {
                "active": int(row[1] or 0),
                "inactive": int(row[2] or 0),
            }
            for row in role_status_rows
        }

        # Device status distribution (current state - no date filter)
        cursor = await db.execute(
            "SELECT status, COUNT(*) AS total FROM devices GROUP BY status"
        )
        device_rows = await cursor.fetchall()
        device_status_counts = {str(r[0]): int(r[1]) for r in device_rows}

        cursor = await db.execute(
            """
            SELECT u.role, d.status, COUNT(*) AS total
            FROM devices d
            INNER JOIN users u
                ON d.current_holder_id REGEXP '^[0-9]+$'
               AND CAST(d.current_holder_id AS UNSIGNED) = u.id
            WHERE u.role IN ('sub_distributor', 'cluster', 'operator')
            GROUP BY u.role, d.status
            """
        )
        holder_rows = await cursor.fetchall()
        holder_role_status_counts = {
            "sub_distributor": {},
            "cluster": {},
            "operator": {},
        }
        for row in holder_rows:
            holder_role = str(row[0])
            status_name = str(row[1])
            total = int(row[2])
            holder_role_status_counts.setdefault(holder_role, {})[status_name] = total

        # Monthly defect trend (last 12 months, or filtered range)
        defect_trend = []
        distribution_trend = []
        trend_months = range(11, -1, -1) if not (start_date or end_date) else range(0, 0)
        for i in trend_months:
            start = _shift_months(month_start, -i)
            end = _shift_months(start, 1)

            cursor = await db.execute(
                "SELECT COUNT(*) FROM defects WHERE created_at >= ? AND created_at < ?",
                (start.isoformat(), end.isoformat())
            )
            reported = (await cursor.fetchone())[0]

            cursor = await db.execute(
                "SELECT COUNT(*) FROM defects WHERE status = 'resolved' AND resolved_at >= ? AND resolved_at < ?",
                (start.isoformat(), end.isoformat())
            )
            resolved = (await cursor.fetchone())[0]

            cursor = await db.execute(
                """SELECT COUNT(*) FROM defects
                   WHERE replacement_device_id IS NOT NULL
                   AND replacement_requested_at >= ? AND replacement_requested_at < ?""",
                (start.isoformat(), end.isoformat())
            )
            replaced = (await cursor.fetchone())[0]

            defect_trend.append({
                "month": start.strftime("%b"),
                "reported": reported,
                "resolved": resolved,
                "replaced": replaced
            })

            cursor = await db.execute(
                "SELECT COUNT(*) FROM distributions WHERE created_at >= ? AND created_at < ?",
                (start.isoformat(), end.isoformat())
            )
            total_dist = (await cursor.fetchone())[0]

            cursor = await db.execute(
                """SELECT COUNT(*) FROM distributions
                   WHERE status IN ('approved', 'delivered')
                   AND created_at >= ? AND created_at < ?""",
                (start.isoformat(), end.isoformat())
            )
            delivered = (await cursor.fetchone())[0]

            distribution_trend.append({
                "month": start.strftime("%b"),
                "total": total_dist,
                "delivered": delivered
            })

        # If date range provided, add a single aggregated entry
        if start_date or end_date:
            dc, dp = _build_date_filter("1=1", (), start_date, end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM defects WHERE {dc}", dp)
            reported_total = (await cursor.fetchone())[0]

            resolved_conds = ["status = 'resolved'"]
            resolved_params = []
            if start_date:
                resolved_conds.append("resolved_at >= ?")
                resolved_params.append(start_date)
            if end_date:
                resolved_conds.append("resolved_at <= ?")
                resolved_params.append(end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM defects WHERE {' AND '.join(resolved_conds)}", tuple(resolved_params))
            resolved_total = (await cursor.fetchone())[0]

            replaced_conds = ["replacement_device_id IS NOT NULL"]
            replaced_params = []
            if start_date:
                replaced_conds.append("replacement_requested_at >= ?")
                replaced_params.append(start_date)
            if end_date:
                replaced_conds.append("replacement_requested_at <= ?")
                replaced_params.append(end_date)
            cursor = await db.execute(f"SELECT COUNT(*) FROM defects WHERE {' AND '.join(replaced_conds)}", tuple(replaced_params))
            replaced_total = (await cursor.fetchone())[0]

            defect_trend.append({
                "month": "Filtered",
                "reported": reported_total,
                "resolved": resolved_total,
                "replaced": replaced_total,
            })

            cursor = await db.execute(f"SELECT COUNT(*) FROM distributions WHERE {dc}", dp)
            total_dist = (await cursor.fetchone())[0]

            dist_delivered_conds = ["status IN ('approved', 'delivered')"]
            if start_date:
                dist_delivered_conds.append("created_at >= ?")
            if end_date:
                dist_delivered_conds.append("created_at <= ?")
            dist_params = tuple(p for p in ([start_date] if start_date else []) + ([end_date] if end_date else []) if p)
            cursor = await db.execute(f"SELECT COUNT(*) FROM distributions WHERE {' AND '.join(dist_delivered_conds)}", dist_params)
            delivered_total = (await cursor.fetchone())[0]

            distribution_trend.append({
                "month": "Filtered",
                "total": total_dist,
                "delivered": delivered_total,
            })

        # Reliability analytics used by Defect Incidence cards
        reliability_start = start_date or (now - timedelta(days=60)).isoformat()
        if not start_date:
            reliability_start = (now - timedelta(days=60)).isoformat()

        rc, rp = _build_date_filter("1=1", (), reliability_start, end_date)
        cursor = await db.execute(f"SELECT COUNT(*) FROM defects WHERE {rc}", rp)
        defects_last_60_days = int((await cursor.fetchone())[0])

        # Build resolved_at-based date filter for repair rate queries
        r_conds, r_params = ["1=1"], []
        if start_date:
            r_conds.append("resolved_at >= ?")
            r_params.append(start_date)
        if end_date:
            r_conds.append("resolved_at <= ?")
            r_params.append(end_date)
        r_cond = " AND ".join(r_conds)
        r_params = tuple(r_params)

        cursor = await db.execute(
            f"""SELECT COUNT(*) FROM defects
               WHERE resolved_at IS NOT NULL
               AND TIMESTAMPDIFF(
                   DAY,
                   STR_TO_DATE(SUBSTRING(REPLACE(created_at, 'T', ' '), 1, 19), '%%Y-%%m-%%d %%H:%%i:%%s'),
                   STR_TO_DATE(SUBSTRING(REPLACE(resolved_at, 'T', ' '), 1, 19), '%%Y-%%m-%%d %%H:%%i:%%s')
               ) <= 60
               AND {r_cond}""", r_params
        )
        repaired_within_sla_devices = int((await cursor.fetchone())[0])

        cursor = await db.execute(
            f"SELECT COUNT(*) FROM defects WHERE resolved_at IS NOT NULL AND {r_cond}", r_params
        )
        total_resolved_defects = int((await cursor.fetchone())[0])

        total_devices_for_reliability = int(device_stats.get("total", 0))
        defect_incidence_percentage = (
            round((defects_last_60_days / total_devices_for_reliability) * 100, 2)
            if total_devices_for_reliability > 0 else 0.0
        )
        repaired_within_sla_percentage = (
            round((repaired_within_sla_devices / total_resolved_defects) * 100, 2)
            if total_resolved_defects > 0 else 0.0
        )

    active_devices = (
        device_stats.get("available", 0) +
        device_stats.get("distributed", 0) +
        device_stats.get("in_use", 0)
    )
    total_devices = int(device_stats.get("total", 0))
    inactive_devices = max(0, total_devices - active_devices)

    replacement_success_rate = (
        round((replacements_confirmed / replacements_total) * 100, 2)
        if replacements_total > 0 else 0
    )

    management_kpis = {
        "total_devices": total_devices,
        "active_devices": active_devices,
        "inactive_devices": inactive_devices,
        "total_users": int(user_stats.get("total", 0)),
        "total_staff": int(role_counts.get("pdic_staff", 0)),
        "total_operators": int(role_counts.get("operator", 0)),
        "total_sub_distributors": int(role_counts.get("sub_distributor", 0)),
        "total_clusters": int(role_counts.get("cluster", 0)),
        "defects_this_month": int(defects_this_month),
        "defects_this_year": int(defects_this_year),
        "replacements_total": int(replacements_total),
        "replacements_confirmed": int(replacements_confirmed),
        "replacements_pending": int(replacements_pending),
        "replacement_success_rate": replacement_success_rate,
        "pending_approvals": int(approval_stats.get("total_pending", 0)),
        "pending_receipts": int(dist_stats.get("pending_receipt", 0)),
    }

    charts = {
        "device_status": {
            "available": int(device_status_counts.get("available", 0)),
            "distributed": int(device_status_counts.get("distributed", 0)),
            "in_use": int(device_status_counts.get("in_use", 0)),
            "defective": int(device_status_counts.get("defective", 0)),
            "returned": int(device_status_counts.get("returned", 0)),
        },
        "device_active_split": {
            "active": active_devices,
            "inactive": inactive_devices,
        },
        "user_roles": {
            "pdic_staff": int(role_counts.get("pdic_staff", 0)),
            "sub_distributor": int(role_counts.get("sub_distributor", 0)),
            "cluster": int(role_counts.get("cluster", 0)),
            "operator": int(role_counts.get("operator", 0)),
            "manager": int(role_counts.get("manager", 0)),
            "super_admin": int(role_counts.get("super_admin", 0)),
        },
        "defect_severity": {
            "critical": int(defect_stats.get("by_severity", {}).get("critical", 0)),
            "high": int(defect_stats.get("by_severity", {}).get("high", 0)),
            "medium": int(defect_stats.get("by_severity", {}).get("medium", 0)),
            "low": int(defect_stats.get("by_severity", {}).get("low", 0)),
        },
        "defect_trend_12m": defect_trend,
        "distribution_trend_12m": distribution_trend,
        "replacement_pipeline": {
            "replaced": int(replacements_total),
            "confirmed": int(replacements_confirmed),
            "pending_confirmation": int(replacements_pending),
        },
        "returns_by_status": {
            "pending": int(return_stats.get("by_status", {}).get("pending", 0)),
            "approved": int(return_stats.get("by_status", {}).get("approved", 0)),
            "received": int(return_stats.get("by_status", {}).get("received", 0)),
            "rejected": int(return_stats.get("by_status", {}).get("rejected", 0)),
        },
        "sub_distributor_account_active_split": role_status_splits.get("sub_distributor", {"active": 0, "inactive": 0}),
        "cluster_account_active_split": role_status_splits.get("cluster", {"active": 0, "inactive": 0}),
        "operator_account_active_split": role_status_splits.get("operator", {"active": 0, "inactive": 0}),
        "sub_distributor_device_active_split": _active_inactive_from_status_counts(
            holder_role_status_counts.get("sub_distributor", {})
        ),
        "cluster_device_active_split": _active_inactive_from_status_counts(
            holder_role_status_counts.get("cluster", {})
        ),
        "operator_device_active_split": _active_inactive_from_status_counts(
            holder_role_status_counts.get("operator", {})
        ),
        "pending_action_queue": {
            "approvals": int(approval_stats.get("total_pending", 0)),
            "receipts": int(dist_stats.get("pending_receipt", 0)),
            "returns": int(return_stats.get("by_status", {}).get("pending", 0)),
        },
    }

    # Staff should not get governance-only user-role visibility for admin/manager counts.
    if role == "pdic_staff":
        charts["user_roles"].pop("super_admin", None)
        charts["user_roles"].pop("manager", None)

    return {
        "kpis": management_kpis,
        "charts": charts,
        "alerts": alerts,
        "reliability": {
            "summary": {
                "defect_incidence_percentage": defect_incidence_percentage,
                "repaired_within_sla_devices": repaired_within_sla_devices,
                "repaired_within_sla_percentage": repaired_within_sla_percentage,
                "defects_last_60_days": defects_last_60_days,
                "total_resolved_defects": total_resolved_defects,
            },
            "trend": defect_trend,
        },
    }


async def get_distribution_device_analytics(start_date: Optional[str] = None,
                                            end_date: Optional[str] = None) -> Dict[str, Any]:
    """Get distribution device analytics: sent, remaining, by type and manufacturer."""
    async with get_db() as db:
        date_cond, date_params = _build_date_filter("status IN ('distributed', 'in_use')", (), start_date, end_date)
        cursor = await db.execute(
            f"""SELECT
                   COALESCE(NULLIF(TRIM(device_type), ''), 'Unknown') AS device_type,
                   COUNT(*) AS total
               FROM devices
               WHERE {date_cond}
               GROUP BY COALESCE(NULLIF(TRIM(device_type), ''), 'Unknown')
               ORDER BY total DESC""",
            date_params
        )
        sent_by_type_rows = await cursor.fetchall()

        cursor = await db.execute(
            f"SELECT COUNT(*) FROM devices WHERE {date_cond}", date_params
        )
        total_sent = (await cursor.fetchone())[0]

        cursor = await db.execute(
            "SELECT COUNT(*) FROM devices WHERE status = 'available'"
        )
        remaining_available = (await cursor.fetchone())[0]

        cursor = await db.execute(
            """SELECT
                   COALESCE(NULLIF(TRIM(device_type), ''), 'Unknown') AS device_type,
                   COUNT(*) AS total
               FROM devices
               WHERE status = 'available'
               GROUP BY COALESCE(NULLIF(TRIM(device_type), ''), 'Unknown')
               ORDER BY total DESC"""
        )
        remaining_by_type_rows = await cursor.fetchall()

        cursor = await db.execute(
            f"""SELECT
                   COALESCE(NULLIF(TRIM(manufacturer), ''), 'Unknown') AS manufacturer,
                   COUNT(*) AS total
               FROM devices
               WHERE {date_cond}
               GROUP BY COALESCE(NULLIF(TRIM(manufacturer), ''), 'Unknown')
               ORDER BY total DESC""",
            date_params
        )
        by_manufacturer_rows = await cursor.fetchall()

        per_holder_cond = date_cond.replace("status IN ('distributed', 'in_use')", "d.status IN ('distributed', 'in_use')")
        cursor = await db.execute(
            f"""SELECT
                   CAST(d.current_holder_id AS TEXT) AS holder_id,
                   COALESCE(NULLIF(TRIM(d.current_holder_name), ''), 'Unknown') AS holder_name,
                   COALESCE(NULLIF(TRIM(d.device_type), ''), 'Unknown') AS device_type,
                   COUNT(*) AS total
               FROM devices d
               INNER JOIN users u ON CAST(d.current_holder_id AS UNSIGNED) = u.id AND u.role = 'sub_distributor'
               WHERE {per_holder_cond}
               AND d.current_holder_id IS NOT NULL
               GROUP BY
                   CAST(d.current_holder_id AS TEXT),
                   COALESCE(NULLIF(TRIM(d.current_holder_name), ''), 'Unknown'),
                   COALESCE(NULLIF(TRIM(d.device_type), ''), 'Unknown')
               ORDER BY holder_name, device_type""",
            date_params
        )
        per_holder_rows = await cursor.fetchall()

        cursor = await db.execute(
            """SELECT
                   CAST(d.current_holder_id AS TEXT) AS holder_id,
                   COALESCE(NULLIF(TRIM(d.current_holder_name), ''), 'Unknown') AS holder_name,
                   COUNT(*) AS total
               FROM devices d
               INNER JOIN users u ON CAST(d.current_holder_id AS UNSIGNED) = u.id AND u.role = 'sub_distributor'
               WHERE d.status = 'available'
               AND d.current_holder_id IS NOT NULL
               GROUP BY
                   CAST(d.current_holder_id AS TEXT),
                   COALESCE(NULLIF(TRIM(d.current_holder_name), ''), 'Unknown')
               ORDER BY total DESC"""
        )
        per_holder_available_rows = await cursor.fetchall()

    remaining_by_type: Dict[str, int] = {
        row["device_type"]: int(row["total"]) for row in remaining_by_type_rows
    }
    sent_by_type = [
        {
            "device_type": row["device_type"],
            "sent": int(row["total"]),
            "remaining": remaining_by_type.get(row["device_type"], 0),
        }
        for row in sent_by_type_rows
    ]

    by_manufacturer = [
        {"manufacturer": row["manufacturer"], "total": int(row["total"])}
        for row in by_manufacturer_rows
    ]

    holder_map: Dict[str, Dict[str, Any]] = {}
    for row in per_holder_rows:
        hid = row["holder_id"]
        if hid not in holder_map:
            holder_map[hid] = {
                "holder_id": hid,
                "holder_name": row["holder_name"],
                "total_sent": 0,
                "by_type": {},
            }
        count = int(row["total"])
        holder_map[hid]["total_sent"] += count
        holder_map[hid]["by_type"][row["device_type"]] = (
            holder_map[hid]["by_type"].get(row["device_type"], 0) + count
        )

    available_map: Dict[str, int] = {}
    for row in per_holder_available_rows:
        available_map[row["holder_id"]] = int(row["total"])

    per_holder = []
    for hid, payload in holder_map.items():
        per_holder.append({
            "holder_id": hid,
            "holder_name": payload["holder_name"],
            "total_sent": payload["total_sent"],
            "remaining_available": available_map.get(hid, 0),
            "type_breakdown": [
                {"device_type": dt, "count": c}
                for dt, c in sorted(payload["by_type"].items(), key=lambda x: x[1], reverse=True)
            ],
        })
    per_holder.sort(key=lambda x: x["total_sent"], reverse=True)

    return {
        "total_sent_to_distribution": total_sent,
        "remaining_available_devices": remaining_available,
        "sent_by_type": sent_by_type,
        "by_manufacturer": by_manufacturer,
        "per_holder_breakdown": per_holder,
    }


async def get_view_as_dashboard(
    target_user: Dict[str, Any],
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
) -> Dict[str, Any]:
    role = target_user.get("role")
    user_id = str(target_user.get("_id", target_user.get("id", "")))

    stats = await get_dashboard_stats(target_user, start_date, end_date)
    advanced = await get_advanced_dashboard_metrics(target_user, start_date, end_date)

    async with get_db() as db:
        scope_root_id = _resolve_scope_root_for_sub_distribution_manager(target_user, user_id)

        if role in ("sub_distributor",):
            cursor = await db.execute(
                "SELECT id, name, email, role, status, phone, location FROM users WHERE role = 'operator' AND parent_id IN (SELECT id FROM users WHERE role = 'cluster' AND parent_id IN (SELECT id FROM users WHERE role = 'sub_distribution_manager' AND parent_id = ?) UNION SELECT id FROM users WHERE role = 'sub_distribution_manager' AND parent_id = ?)",
                (int(user_id), int(user_id))
            )
            target_users = rows_to_list(await cursor.fetchall())
        elif role == "cluster":
            cursor = await db.execute(
                "SELECT id, name, email, role, status FROM users WHERE role = 'operator' AND parent_id = ?",
                (int(user_id),)
            )
            target_users = rows_to_list(await cursor.fetchall())
        else:
            target_users = []

        scope_ids = sorted({scope_root_id} | await _get_descendant_user_ids(db, scope_root_id)) if role in ("sub_distribution_manager", "sub_distributor") else [user_id]

        dev_conds = [f"current_holder_id IN ({','.join(['?'] * len(scope_ids))})"] if len(scope_ids) > 0 else ["1=0"]
        dev_params = list(scope_ids) if len(scope_ids) > 0 else []
        dc, dp = _build_date_filter(" AND ".join(dev_conds), tuple(dev_params), start_date, end_date)
        cursor = await db.execute(f"SELECT * FROM devices WHERE {dc}", dp)
        target_devices = rows_to_list(await cursor.fetchall())

        def_conds = [f"CAST(reported_by AS TEXT) IN ({','.join(['?'] * len(scope_ids))})"] if len(scope_ids) > 0 else ["1=0"]
        def_params = [str(s) for s in scope_ids] if len(scope_ids) > 0 else []
        dfc, dfp = _build_date_filter(" AND ".join(def_conds), tuple(def_params), start_date, end_date)
        cursor = await db.execute(f"SELECT * FROM defects WHERE {dfc}", dfp)
        target_defects = rows_to_list(await cursor.fetchall())

        ret_conds = [f"CAST(requested_by AS TEXT) IN ({','.join(['?'] * len(scope_ids))})"] if len(scope_ids) > 0 else ["1=0"]
        ret_params = [str(s) for s in scope_ids] if len(scope_ids) > 0 else []
        rc, rp = _build_date_filter(" AND ".join(ret_conds), tuple(ret_params), start_date, end_date)
        cursor = await db.execute(f"SELECT * FROM returns WHERE {rc}", rp)
        target_returns = rows_to_list(await cursor.fetchall())

        dist_conds = [f"(from_user_id IN ({','.join(['?'] * len(scope_ids))}) OR to_user_id IN ({','.join(['?'] * len(scope_ids))}))"] if len(scope_ids) > 0 else ["1=0"]
        dist_params = list(scope_ids) + list(scope_ids) if len(scope_ids) > 0 else []
        dic, dip = _build_date_filter(" AND ".join(dist_conds), tuple(dist_params), start_date, end_date)
        cursor = await db.execute(f"SELECT * FROM distributions WHERE {dic}", dip)
        target_distributions = rows_to_list(await cursor.fetchall())

    return {
        "user": {"id": target_user.get("id"), "name": target_user.get("name", ""), "role": target_user.get("role", "")},
        "stats": stats,
        "advanced": advanced,
        "devices": target_devices,
        "defects": target_defects,
        "returns": target_returns,
        "distributions": target_distributions,
        "users": target_users,
    }


async def generate_report(
    current_user: Dict[str, Any],
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
) -> Dict[str, Any]:
    """Generate a comprehensive system report as Excel with date-filtered data."""
    role = current_user.get("role")
    user_id = str(current_user.get("_id", current_user.get("id", "")))
    scope_root_id = _resolve_scope_root_for_sub_distribution_manager(current_user, user_id)

    async with get_db() as db:
        date_cond, date_params = _build_date_filter("1=1", (), start_date, end_date)

        scope_cond = "1=1"
        scope_params: List[Any] = []
        if role not in ["super_admin", "md_director", "manager", "pdic_staff"]:
            scoped_ids = sorted({scope_root_id} | await _get_descendant_user_ids(db, scope_root_id))
            placeholders = ",".join(["?"] * len(scoped_ids))
            scope_cond = f"current_holder_id IN ({placeholders})"
            scope_params = list(scoped_ids)

        device_stats = await device_service.get_device_stats(start_date, end_date)

        device_type_cond = f"({date_cond}) AND ({scope_cond})"
        device_type_params = list(date_params) + scope_params

        cursor = await db.execute(
            f"""SELECT
                   COALESCE(NULLIF(TRIM(device_type), ''), 'Unknown') AS device_type,
                   COUNT(*) AS total
               FROM devices
               WHERE {date_cond}
               GROUP BY COALESCE(NULLIF(TRIM(device_type), ''), 'Unknown')
               ORDER BY total DESC""",
            date_params
        )
        all_devices_by_type = {(r["device_type"]): int(r["total"]) for r in await cursor.fetchall()}

        cursor = await db.execute(
            f"""SELECT
                   COALESCE(NULLIF(TRIM(device_type), ''), 'Unknown') AS device_type,
                   COUNT(*) AS total
               FROM devices
               WHERE status IN ('distributed', 'in_use') AND {date_cond}
               GROUP BY COALESCE(NULLIF(TRIM(device_type), ''), 'Unknown')
               ORDER BY total DESC""",
            date_params
        )
        distributed_by_type = {(r["device_type"]): int(r["total"]) for r in await cursor.fetchall()}

        date_cond_d = date_cond.replace("created_at", "d.created_at")

        cursor = await db.execute(
            f"""SELECT
                   CAST(d.current_holder_id AS TEXT) AS holder_id,
                   COALESCE(NULLIF(TRIM(u.name), ''), 'Unknown') AS holder_name,
                   COUNT(*) AS total_sent
               FROM devices d
               LEFT JOIN users u ON CAST(d.current_holder_id AS UNSIGNED) = u.id
               WHERE d.status IN ('distributed', 'in_use') AND {date_cond_d}
               GROUP BY CAST(d.current_holder_id AS TEXT), COALESCE(NULLIF(TRIM(u.name), ''), 'Unknown')
               ORDER BY total_sent DESC""",
            date_params
        )
        subdistributor_rows = await cursor.fetchall()

        cursor = await db.execute(
            f"""SELECT *
               FROM devices
               WHERE {date_cond}
               ORDER BY id ASC""",
            date_params
        )
        all_device_rows = rows_to_list(await cursor.fetchall())

        cursor = await db.execute(
            f"""SELECT d.*,
                       COALESCE(NULLIF(TRIM(u.name), ''), 'Unknown') AS holder_name
               FROM devices d
               LEFT JOIN users u ON CAST(d.current_holder_id AS UNSIGNED) = u.id
               WHERE d.status IN ('distributed', 'in_use') AND {date_cond_d}
               ORDER BY d.id ASC""",
            date_params
        )
        distributed_device_rows = rows_to_list(await cursor.fetchall())

    total_distributed = sum(distributed_by_type.values())

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    date_label = "All Time"
    if start_date and end_date:
        date_label = f"{start_date[:10]} to {end_date[:10]}"
    elif start_date:
        date_label = f"From {start_date[:10]}"

    wb = Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.merge_cells("A1:D1")
    cell_title = ws1["A1"]
    cell_title.value = f"System Report - Generated {now_str}  |  Range: {date_label}"
    cell_title.font = Font(name="Calibri", size=14, bold=True)
    ws2 = wb.create_sheet("All Devices")
    ws3 = wb.create_sheet("Distributed Devices")
    ws4 = wb.create_sheet("Devices by Holder")
    ws5 = wb.create_sheet("Device Type Summary")

    for ws in [ws2, ws3, ws4, ws5]:
        ws["A1"] = f"System Report - Generated {now_str}  |  Range: {date_label}"

    summary_headers = ["Metric", "Value"]
    ws1.append(summary_headers)
    ws1.append(["Filter Period", date_label])
    ws1.append(["Total Devices (filtered)", device_stats.get("total", 0)])
    ws1.append(["Total Distributed/In Use", total_distributed])
    ws1.append(["Total Available (current)", device_stats.get("available", 0)])
    ws1.append(["Total Defective (filtered)", device_stats.get("defective", 0)])
    ws1.append([])
    ws1.append(["Distribution by Holder"])
    ws1.append(["Holder", "Total Devices Sent"])
    for row in subdistributor_rows:
        ws1.append([row["holder_name"], int(row["total_sent"])])

    # Sheet 2: All Devices
    all_headers = [
        "ID", "Device ID", "Serial Number", "MAC Address", "NUID",
        "Device Type", "Model", "Manufacturer", "Status",
        "Current Holder", "Current Location", "Created At"
    ]
    ws2.append(all_headers)
    for dev in all_device_rows:
        ws2.append([
            dev.get("id", ""),
            dev.get("device_id", ""),
            dev.get("serial_number", ""),
            dev.get("mac_address", ""),
            dev.get("nuid", ""),
            dev.get("device_type", ""),
            dev.get("model", ""),
            dev.get("manufacturer", ""),
            dev.get("status", ""),
            dev.get("current_holder_name", ""),
            dev.get("current_location", ""),
            dev.get("created_at", ""),
        ])

    # Sheet 3: Distributed Devices
    dist_headers = [
        "ID", "Device ID", "Serial Number", "MAC Address", "NUID",
        "Device Type", "Model", "Manufacturer",
        "Holder", "Current Location", "Status", "Created At"
    ]
    ws3.append(dist_headers)
    for dev in distributed_device_rows:
        ws3.append([
            dev.get("id", ""),
            dev.get("device_id", ""),
            dev.get("serial_number", ""),
            dev.get("mac_address", ""),
            dev.get("nuid", ""),
            dev.get("device_type", ""),
            dev.get("model", ""),
            dev.get("manufacturer", ""),
            dev.get("holder_name", ""),
            dev.get("current_location", ""),
            dev.get("status", ""),
            dev.get("created_at", ""),
        ])

    # Sheet 4: Devices by Holder
    sd_headers = ["Holder", "Total Devices Sent"]
    ws4.append(sd_headers)
    for row in subdistributor_rows:
        ws4.append([row["holder_name"], int(row["total_sent"])])

    # Sheet 5: Device Type Summary
    type_headers = ["Device Type", "Total (filtered)", "Distributed/In Use", "Remaining"]
    ws5.append(type_headers)
    all_types = sorted(set(list(all_devices_by_type.keys()) + list(distributed_by_type.keys())))
    for dt in all_types:
        total = all_devices_by_type.get(dt, 0)
        dist = distributed_by_type.get(dt, 0)
        remaining = total - dist
        ws5.append([dt, total, dist, remaining])

    payload = io.BytesIO()
    wb.save(payload)
    payload.seek(0)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    date_suffix = f"_{start_date[:10]}_{end_date[:10]}" if start_date and end_date else ""
    return {
        "content": payload.getvalue(),
        "filename": f"report{date_suffix}_{ts}.xlsx",
        "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

