"""Approval-request service for sub-distribution employees.

Employees (role='sub_distribution_employee') cannot directly create
distributions, defect reports, clusters, or operators. Instead they submit a
proposal stored in `approval_requests`. The branch sub-distributor and, when
one exists, the sub-distribution manager must both approve before the proposal
is applied. The stored payload is revalidated at the moment of the final
approval (the underlying service functions perform their own fresh checks) so a
stale proposal is never executed blindly.
"""
import json
import logging
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

from sqlalchemy import text

from app.database_sqlalchemy import async_session_factory
from app.models.distribution import DistributionCreate, DistributionStatus
from app.models.defect import DefectCreate, DefectStatus
from app.models.user import UserCreate, UserUpdate
from app.services import (
    user_service,
    distribution_service,
    defect_service,
    return_service,
    notification_service,
    bulk_upload_service,
)
from app.utils.helpers import get_pagination
from app.utils.roles import (
    SUB_DISTRIBUTOR,
    SUB_DISTRIBUTION_MANAGER,
    SUB_DISTRIBUTION_EMPLOYEE,
    SUPER_ADMIN,
    CLUSTER,
    OPERATOR,
    normalize_role,
)

logger = logging.getLogger(__name__)

VALID_REQUEST_TYPES = (
    "distribution",
    "defect",
    "cluster",
    "operator",
    "user_update",
    "user_delete",
    "user_reassign",
    "bulk_users",
    "bulk_distribution",
    "delivery_receipt",
    "return_status",
    "defect_status",
    "payment_confirmation",
)


def _now() -> datetime:
    return datetime.now().replace(tzinfo=None)


def _serialize_request(row: Dict[str, Any], *, include_approvals: bool = True) -> Dict[str, Any]:
    item = dict(row)
    item["_id"] = str(item.get("id") or "")
    for key in ("payload", "approvals", "execution_result"):
        raw = item.get(key)
        if isinstance(raw, str):
            try:
                item[key] = json.loads(raw)
            except (json.JSONDecodeError, TypeError):
                item[key] = None if raw is None or raw == "" else raw
    if item.get("required_roles") and isinstance(item["required_roles"], str):
        try:
            item["required_roles"] = json.loads(item["required_roles"])
        except (json.JSONDecodeError, TypeError):
            item["required_roles"] = []
    if not include_approvals:
        item.pop("payload", None)
    return item


async def _get_approver_users(session, sub_distribution_id: int) -> List[Dict[str, Any]]:
    """Return the eligible approver users for a branch:
    the sub-distributor plus any sub-distribution managers under it."""
    approvers: List[Dict[str, Any]] = []
    rows = (await session.execute(
        text("""
            SELECT id, name, role, parent_id FROM users
            WHERE (id = :sd_id AND role = :sd_role)
               OR (parent_id = :sd_id AND role = :sd_mgr_role)
            ORDER BY id
        """),
        {"sd_id": int(sub_distribution_id), "sd_role": SUB_DISTRIBUTOR, "sd_mgr_role": SUB_DISTRIBUTION_MANAGER},
    )).mappings().all()
    for row in rows:
        approvers.append({
            "id": int(row["id"]),
            "name": row["name"],
            "role": normalize_role(row["role"]),
            "parent_id": row["parent_id"],
        })
    return approvers


def _build_required_roles(approvers: List[Dict[str, Any]]) -> List[str]:
    roles = {a["role"] for a in approvers}
    required = []
    if SUB_DISTRIBUTOR in roles:
        required.append(SUB_DISTRIBUTOR)
    if SUB_DISTRIBUTION_MANAGER in roles:
        required.append(SUB_DISTRIBUTION_MANAGER)
    return required


async def _branch_contains_user(session, root_id: int, target_id: int) -> bool:
    row = (await session.execute(
        text("""
            WITH RECURSIVE descendants AS (
                SELECT id FROM users WHERE parent_id = :root
                UNION ALL
                SELECT u.id FROM users u
                INNER JOIN descendants d ON u.parent_id = d.id
            )
            SELECT 1 FROM descendants WHERE id = :target LIMIT 1
        """),
        {"root": int(root_id), "target": int(target_id)},
    )).mappings().first()
    return bool(row)


async def _branch_sub_distributor_actor(session, sub_distribution_id: int) -> Dict[str, Any]:
    """Return the branch sub-distributor as the acting identity for an approved request.

    Execution-time service calls run as the branch sub-distributor (rather than the
    employee) so role-gated services (bulk uploads, confirmations, etc.) behave as
    if the branch itself performed the action after its approval.
    """
    row = (await session.execute(
        text("SELECT * FROM users WHERE id = :id AND role = :role"),
        {"id": int(sub_distribution_id), "role": SUB_DISTRIBUTOR},
    )).mappings().first()
    if not row:
        raise ValueError("Sub distribution account not found")
    user = dict(row)
    return {
        "id": user["id"],
        "_id": user["id"],
        "name": user.get("name") or user.get("email") or "Sub Distributor",
        "role": normalize_role(user.get("role")),
        "parent_id": user.get("parent_id"),
        "email": user.get("email"),
    }


def _validate_bulk_rows(rows: Any) -> List[Dict[str, Any]]:
    if not isinstance(rows, list) or not rows:
        raise ValueError("rows must be a non-empty list")
    return rows


def _validate_defect_status_payload(payload: Dict[str, Any]) -> None:
    if not payload.get("defect_id"):
        raise ValueError("defect_id is required")
    if not payload.get("status"):
        raise ValueError("status is required")


def _validate_delivery_receipt_payload(payload: Dict[str, Any]) -> None:
    if not payload.get("distribution_id"):
        raise ValueError("distribution_id is required")


def _validate_return_status_payload(payload: Dict[str, Any]) -> None:
    if not payload.get("return_id"):
        raise ValueError("return_id is required")
    if not payload.get("status"):
        raise ValueError("status is required")


def _validate_user_edit_payload(payload: Dict[str, Any], request_type: str) -> None:
    if not payload.get("user_id"):
        raise ValueError("user_id is required")
    if request_type == "user_update" and not any(
        k in payload for k in ("name", "phone", "designation", "address", "pincode", "status", "network_name")
    ):
        raise ValueError("At least one field to update is required")
    if request_type == "user_reassign" and not payload.get("new_parent_id"):
        raise ValueError("new_parent_id is required")


def _validate_distribution_payload(payload: Dict[str, Any]) -> None:
    to_user_id = payload.get("to_user_id")
    device_ids = payload.get("device_ids")
    if not to_user_id:
        raise ValueError("to_user_id is required")
    if not isinstance(device_ids, list) or not device_ids:
        raise ValueError("device_ids must be a non-empty list")


def _validate_defect_payload(payload: Dict[str, Any]) -> None:
    if not payload.get("device_id"):
        raise ValueError("device_id is required")
    if not payload.get("defect_type"):
        raise ValueError("defect_type is required")
    if not payload.get("severity"):
        raise ValueError("severity is required")
    if not payload.get("description") or len(str(payload["description"]).strip()) < 10:
        raise ValueError("description is required (minimum 10 characters)")


def _validate_user_payload(payload: Dict[str, Any], request_type: str) -> None:
    email = str(payload.get("email") or "").strip()
    name = str(payload.get("name") or "").strip()
    password = str(payload.get("password") or "")
    phone = str(payload.get("phone") or "").strip()
    if not email or "@" not in email:
        raise ValueError("A valid email is required")
    if not name:
        raise ValueError("name is required")
    if len(password) < 8:
        raise ValueError("password must be at least 8 characters")
    if len(phone) < 10:
        raise ValueError("phone must be at least 10 characters")
    if request_type == "cluster":
        if "sub_distribution_id" not in payload or not str(payload.get("sub_distribution_id") or "").strip():
            raise ValueError("sub_distribution_id is required for cluster creation")


async def _validate_payload_and_scope(session, requester: Dict[str, Any], request_type: str, payload: Dict[str, Any]) -> None:
    """Validate the proposal at submission time (hard scope constraints)."""
    if request_type == "distribution":
        _validate_distribution_payload(payload)
        to_user_id = int(payload["to_user_id"])
        recipient = (await session.execute(
            text("SELECT id, role, status FROM users WHERE id = :uid"), {"uid": to_user_id}
        )).mappings().first()
        if not recipient:
            raise ValueError("Selected recipient does not exist")
        if normalize_role(recipient["role"]) not in {CLUSTER, OPERATOR}:
            raise ValueError("Recipient must be a cluster or operator")
        if not await _branch_contains_user(session, int(requester["parent_id"]), to_user_id):
            raise ValueError("Selected recipient is outside your sub distribution")
    elif request_type == "defect":
        _validate_defect_payload(payload)
        device = (await session.execute(
            text("SELECT id FROM devices WHERE id = :did"), {"did": int(payload["device_id"])}
        )).mappings().first()
        if not device:
            raise ValueError("Selected device does not exist")
    elif request_type in {"cluster", "operator"}:
        _validate_user_payload(payload, request_type)
        if request_type == "cluster":
            sd_id = int(payload["sub_distribution_id"])
            if sd_id != int(requester["parent_id"]):
                raise ValueError("Cluster must be created under your own sub distribution")
            sd = (await session.execute(
                text("SELECT id FROM users WHERE id = :uid AND role = :role"),
                {"uid": sd_id, "role": SUB_DISTRIBUTOR},
            )).mappings().first()
            if not sd:
                raise ValueError("Invalid sub distribution selected")
        elif request_type == "operator":
            parent_id = payload.get("parent_id")
            if parent_id:
                parent = (await session.execute(
                    text("SELECT id, role, status FROM users WHERE id = :uid"), {"uid": int(parent_id)}
                )).mappings().first()
                if not parent:
                    raise ValueError("Selected operator parent does not exist")
                if normalize_role(parent["role"]) not in {CLUSTER, SUB_DISTRIBUTOR}:
                    raise ValueError("Operator parent must be a cluster or sub distributor")
                if not await _branch_contains_user(session, int(requester["parent_id"]), int(parent_id)):
                    raise ValueError("Selected parent is outside your sub distribution")
    elif request_type in {"user_update", "user_delete", "user_reassign"}:
        _validate_user_edit_payload(payload, request_type)
        target = (await session.execute(
            text("SELECT id, role, status FROM users WHERE id = :uid"),
            {"uid": int(payload["user_id"])},
        )).mappings().first()
        if not target:
            raise ValueError("Target user does not exist")
        if not await _branch_contains_user(session, int(requester["parent_id"]), int(payload["user_id"])):
            raise ValueError("Target user is outside your sub distribution")
        if request_type == "user_reassign":
            new_parent = (await session.execute(
                text("SELECT id, role, status FROM users WHERE id = :uid"),
                {"uid": int(payload["new_parent_id"])},
            )).mappings().first()
            if not new_parent:
                raise ValueError("New parent does not exist")
            if not await _branch_contains_user(session, int(requester["parent_id"]), int(payload["new_parent_id"])):
                raise ValueError("New parent is outside your sub distribution")
            if normalize_role(new_parent["role"]) not in {CLUSTER, SUB_DISTRIBUTOR}:
                raise ValueError("New parent must be a cluster or sub distributor")
    elif request_type == "bulk_users":
        _validate_bulk_rows(payload.get("rows"))
        target_role = payload.get("role")
        if target_role != OPERATOR:
            raise ValueError("Employees can only bulk-upload operators")
        parent_id = payload.get("parent_id")
        if not parent_id:
            raise ValueError("A parent is required when bulk-uploading operators")
        if not await _branch_contains_user(session, int(requester["parent_id"]), int(parent_id)):
            raise ValueError("Parent is outside your sub distribution")
    elif request_type == "bulk_distribution":
        _validate_bulk_rows(payload.get("rows"))
        to_user_id = payload.get("to_user_id")
        if not to_user_id:
            raise ValueError("to_user_id is required")
        recipient = (await session.execute(
            text("SELECT id, role, status FROM users WHERE id = :uid"), {"uid": int(to_user_id)}
        )).mappings().first()
        if not recipient:
            raise ValueError("Selected recipient does not exist")
        if normalize_role(recipient["role"]) not in {CLUSTER, OPERATOR}:
            raise ValueError("Recipient must be a cluster or operator")
        if not await _branch_contains_user(session, int(requester["parent_id"]), int(to_user_id)):
            raise ValueError("Recipient is outside your sub distribution")
    elif request_type == "delivery_receipt":
        _validate_delivery_receipt_payload(payload)
        dist = (await session.execute(
            text("SELECT id, to_user_id, status FROM distributions WHERE distribution_id = :id"),
            {"id": str(payload["distribution_id"])},
        )).mappings().first()
        if not dist:
            raise ValueError("Distribution not found")
        if not await _branch_contains_user(session, int(requester["parent_id"]), int(dist["to_user_id"])):
            raise ValueError("Distribution recipient is outside your sub distribution")
    elif request_type == "return_status":
        _validate_return_status_payload(payload)
        ret = (await session.execute(
            text("SELECT return_id, requested_by FROM returns WHERE return_id = :id"),
            {"id": str(payload["return_id"])},
        )).mappings().first()
        if not ret:
            raise ValueError("Return request not found")
        if int(ret["requested_by"]) and not await _branch_contains_user(session, int(requester["parent_id"]), int(ret["requested_by"])):
            raise ValueError("Return request is outside your sub distribution")
    elif request_type == "defect_status":
        _validate_defect_status_payload(payload)
        defect = (await session.execute(
            text("SELECT id, reported_by FROM defects WHERE id = :id"),
            {"id": int(payload["defect_id"])},
        )).mappings().first()
        if not defect:
            raise ValueError("Defect report not found")
        if int(defect["reported_by"]) and not await _branch_contains_user(session, int(requester["parent_id"]), int(defect["reported_by"])):
            raise ValueError("Defect report is outside your sub distribution")
    elif request_type == "payment_confirmation":
        if not payload.get("defect_id"):
            raise ValueError("defect_id is required")
        defect = (await session.execute(
            text("SELECT id, reported_by FROM defects WHERE id = :id"),
            {"id": int(payload["defect_id"])},
        )).mappings().first()
        if not defect:
            raise ValueError("Defect report not found")
        if int(defect["reported_by"]) and not await _branch_contains_user(session, int(requester["parent_id"]), int(defect["reported_by"])):
            raise ValueError("Defect report is outside your sub distribution")


async def _find_duplicate_pending(session, requester_id: int, request_type: str, payload: Dict[str, Any]) -> Optional[int]:
    """Return the id of an existing pending request that duplicates this one."""
    pending = (await session.execute(
        text("""
            SELECT id, payload FROM approval_requests
            WHERE requested_by = :rb AND request_type = :rt AND status = 'pending'
            ORDER BY created_at DESC
        """),
        {"rb": int(requester_id), "rt": request_type},
    )).mappings().all()

    for row in pending:
        try:
            existing_payload = json.loads(row["payload"]) if row["payload"] else {}
        except (json.JSONDecodeError, TypeError):
            continue

        if request_type == "distribution":
            if (str(existing_payload.get("to_user_id")) == str(payload.get("to_user_id"))
                    and set(str(d) for d in (existing_payload.get("device_ids") or []))
                    == set(str(d) for d in (payload.get("device_ids") or []))):
                return row["id"]
        elif request_type == "defect":
            if str(existing_payload.get("device_id")) == str(payload.get("device_id")):
                return row["id"]
        elif request_type in {"cluster", "operator"}:
            if str(existing_payload.get("email") or "").strip().lower() == str(payload.get("email") or "").strip().lower():
                return row["id"]
        elif request_type in {"user_update", "user_delete"}:
            if str(existing_payload.get("user_id")) == str(payload.get("user_id")):
                return row["id"]
        elif request_type == "user_reassign":
            if (str(existing_payload.get("user_id")) == str(payload.get("user_id"))
                    and str(existing_payload.get("new_parent_id")) == str(payload.get("new_parent_id"))):
                return row["id"]
        elif request_type == "bulk_users":
            if (str(existing_payload.get("parent_id")) == str(payload.get("parent_id"))
                    and str(existing_payload.get("role")) == str(payload.get("role"))):
                return row["id"]
        elif request_type == "bulk_distribution":
            if str(existing_payload.get("to_user_id")) == str(payload.get("to_user_id")):
                return row["id"]
        elif request_type == "delivery_receipt":
            if str(existing_payload.get("distribution_id")) == str(payload.get("distribution_id")):
                return row["id"]
        elif request_type == "return_status":
            if (str(existing_payload.get("return_id")) == str(payload.get("return_id"))
                    and str(existing_payload.get("status")) == str(payload.get("status"))):
                return row["id"]
        elif request_type in {"defect_status", "payment_confirmation"}:
            if (str(existing_payload.get("defect_id")) == str(payload.get("defect_id"))
                    and str(existing_payload.get("status")) == str(payload.get("status"))):
                return row["id"]
    return None


async def submit_request(
    requester: Dict[str, Any],
    request_type: str,
    payload: Dict[str, Any],
    summary: Optional[str] = None,
) -> Dict[str, Any]:
    role = normalize_role(requester.get("role"))
    if role != SUB_DISTRIBUTION_EMPLOYEE:
        raise PermissionError("Only sub distribution employees can submit approval requests")

    parent_id = requester.get("parent_id")
    if not parent_id:
        raise ValueError("Employee account is not assigned to a sub distribution")

    parent = await user_service.get_user_by_id(str(parent_id))
    if not parent or normalize_role(parent.get("role")) != SUB_DISTRIBUTOR:
        raise ValueError("Employee is not assigned to a valid sub distribution")

    now = _now()
    request_id = f"APR-{uuid.uuid4().hex[:8].upper()}"
    approvers: List[Dict[str, Any]] = []
    required_roles: List[str] = []

    async with async_session_factory() as session:
        await _validate_payload_and_scope(session, requester, request_type, payload)

        duplicate_id = await _find_duplicate_pending(session, int(requester["id"]), request_type, payload)
        if duplicate_id:
            raise ValueError("You already have a pending request for this item")

        approvers = await _get_approver_users(session, int(parent_id))
        required_roles = _build_required_roles(approvers)
        if not required_roles:
            raise ValueError("No approver configured for this sub distribution")

        await session.execute(
            text("""
                INSERT INTO approval_requests
                    (request_id, request_type, requested_by, requested_by_name,
                     sub_distribution_id, summary, payload, status, required_roles,
                     approvals, rejection_reason, execution_result, executed_at,
                     created_at, updated_at)
                VALUES
                    (:request_id, :request_type, :requested_by, :requested_by_name,
                     :sub_distribution_id, :summary, :payload, 'pending', :required_roles,
                     '[]', NULL, NULL, NULL, :now, :now)
            """),
            {
                "request_id": request_id,
                "request_type": request_type,
                "requested_by": int(requester["id"]),
                "requested_by_name": requester.get("name") or requester.get("email") or "Employee",
                "sub_distribution_id": int(parent_id),
                "summary": (summary or "")[:1000],
                "payload": json.dumps(payload),
                "required_roles": json.dumps(required_roles),
                "now": now,
            },
        )
        await session.commit()

        approver_payloads = [
            {
                "user_id": a["id"],
                "title": "New Employee Approval Request",
                "message": (
                    f"{requester.get('name') or requester.get('email')} submitted a "
                    f"{request_type} request ({request_id}) for your approval."
                ),
                "notification_type": "warning",
                "category": "approval",
                "link": "/sub-distribution-approvals",
                "metadata": {
                    "action": "employee_approval_request",
                    "request_id": request_id,
                    "request_type": request_type,
                    "requested_by": str(requester.get("id") or ""),
                },
            }
            for a in approvers
        ]
        if approver_payloads:
            await notification_service.bulk_create_notifications(approver_payloads)

    return {
        "success": True,
        "message": "Approval request submitted successfully",
        "data": {
            "request_id": request_id,
            "request_type": request_type,
            "status": "pending",
            "required_roles": required_roles,
        },
    }


async def _build_request_rows(session, where: str, params: Dict[str, Any], page: int, page_size: int) -> Dict[str, Any]:
    total = (await session.execute(
        text(f"SELECT COUNT(*) FROM approval_requests WHERE {where}"), params
    )).scalar()

    offset = (page - 1) * page_size
    rows = (await session.execute(
        text(f"SELECT * FROM approval_requests WHERE {where} ORDER BY created_at DESC LIMIT :limit OFFSET :offset"),
        {**params, "limit": page_size, "offset": offset},
    )).mappings().all()

    items = [_serialize_request(dict(r)) for r in rows]
    return {"data": items, "pagination": get_pagination(page, page_size, total)}


async def stage_bulk_payload(
    requester: Dict[str, Any],
    kind: str,
    contents: bytes,
    filename: str,
    *,
    role: Optional[str] = None,
    parent_id: Optional[str] = None,
    to_user_id: Optional[str] = None,
    notes: Optional[str] = None,
    date_of_distribution: Optional[str] = None,
) -> Dict[str, Any]:
    """Parse an uploaded bulk file into the JSON payload stored on an approval request.

    Keeps bulk files out of the (64KB TEXT) payload column by persisting the
    normalized parsed rows instead of the raw file bytes.
    """
    if normalize_role(requester.get("role")) != SUB_DISTRIBUTION_EMPLOYEE:
        raise PermissionError("Only sub distribution employees can stage bulk uploads")

    filename = filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in {"csv", "xlsx", "xls"}:
        raise ValueError("Unsupported file format. Use .csv, .xlsx, or .xls")

    if kind == "users":
        bulk_upload_service.check_bulk_upload_file(contents, f".{ext}")
        rows = bulk_upload_service.parse_file(contents, ext)
        bulk_upload_service.check_bulk_upload_row_count(rows)
        if not rows:
            raise ValueError("File is empty or has no data rows")
        if not parent_id:
            raise ValueError("A parent is required when bulk-uploading operators")
        return {
            "kind": "users",
            "rows": rows,
            "role": normalize_role(role) if role else None,
            "parent_id": str(parent_id),
        }

    if kind == "distribution":
        parsed = _parse_distribution_identifier_file(contents, filename)
        rows, date_values = parsed
        if not rows:
            raise ValueError("No identifier rows found in file")
        normalized_date = None
        if date_values:
            if len(date_values) > 1:
                raise ValueError("Multiple date_of_distribution values found in file; provide a single date")
            normalized_date = date_values.pop()
        if date_of_distribution and normalized_date and date_of_distribution.strip() != normalized_date:
            raise ValueError("date_of_distribution in form and file do not match")
        return {
            "kind": "distribution",
            "rows": rows,
            "to_user_id": str(to_user_id) if to_user_id else None,
            "notes": notes,
            "date_of_distribution": normalized_date or (date_of_distribution.strip() if date_of_distribution else None),
        }

    raise ValueError(f"Unsupported bulk kind: {kind}")


def _parse_distribution_identifier_file(contents: bytes, filename: str) -> tuple:
    """Parse a bulk-distribution file into identifier rows + set of dates."""
    import csv
    import io

    from app.services.bulk_upload_service import check_bulk_upload_file, check_bulk_upload_row_count, MAX_BULK_ROWS

    filename_lower = filename.lower()
    if not filename_lower.endswith((".xlsx", ".xls", ".csv")):
        raise ValueError("Only Excel (.xlsx, .xls) or CSV (.csv) files are supported")

    check_bulk_upload_file(contents, filename_lower)

    if filename_lower.endswith(".csv"):
        decoded = contents.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(decoded))
        all_rows = list(reader)
        if not all_rows:
            raise ValueError("CSV file is empty")
        headers = [str(h).strip().lower() for h in all_rows[0]]
        data_rows = all_rows[1:]
        check_bulk_upload_row_count(data_rows)
        row_iter = enumerate(data_rows, start=2)
    else:
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1), None)
        if not header_row:
            raise ValueError("Excel file is empty")
        headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in header_row]
        row_count = 0

        def _excel_rows():
            nonlocal row_count
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_count += 1
                if row_count > MAX_BULK_ROWS:
                    raise ValueError(f"Too many rows. Maximum is {MAX_BULK_ROWS}")
                yield row_count + 1, row

        row_iter = _excel_rows()

    if "mac_address" not in headers and "serial_number" not in headers and "nuid" not in headers:
        raise ValueError("Missing required columns: add at least one of mac_address, serial_number, or nuid")

    identifier_rows = []
    date_values = set()
    for row_idx, row in row_iter:
        row_data = {
            headers[i]: (str(row[i]).strip() if i < len(row) and row[i] is not None else "")
            for i in range(len(headers))
        }
        mac_address = row_data.get("mac_address", "")
        serial_number = row_data.get("serial_number", "")
        nuid = row_data.get("nuid", "")
        row_date = row_data.get("date_of_distribution", "")
        if row_date:
            date_values.add(row_date)
        if not mac_address and not serial_number and not nuid:
            if not any(v for v in row_data.values()):
                continue
        identifier_rows.append({
            "row": row_idx,
            "mac_address": mac_address,
            "serial_number": serial_number,
            "nuid": nuid,
        })

    return identifier_rows, date_values


async def get_my_requests(
    requester: Dict[str, Any],
    page: int = 1,
    page_size: int = 20,
    status: Optional[str] = None,
) -> Dict[str, Any]:
    if normalize_role(requester.get("role")) != SUB_DISTRIBUTION_EMPLOYEE:
        raise PermissionError("Only employees can list their own approval requests")

    conditions = ["requested_by = :rb"]
    params: Dict[str, Any] = {"rb": int(requester["id"])}
    if status:
        conditions.append("status = :status")
        params["status"] = status

    where = " AND ".join(conditions)
    async with async_session_factory() as session:
        return await _build_request_rows(session, where, params, page, page_size)


async def _is_eligible_approver(session, approver: Dict[str, Any], request_row: Dict[str, Any]) -> bool:
    role = normalize_role(approver.get("role"))
    sub_distribution_id = int(request_row["sub_distribution_id"])

    if role == SUPER_ADMIN:
        return True

    if role == SUB_DISTRIBUTOR:
        return str(approver.get("id")) == str(sub_distribution_id)

    if role == SUB_DISTRIBUTION_MANAGER:
        return str(approver.get("parent_id")) == str(sub_distribution_id)

    return False


async def get_requests_for_approver(
    approver: Dict[str, Any],
    page: int = 1,
    page_size: int = 20,
    status: Optional[str] = None,
    request_type: Optional[str] = None,
) -> Dict[str, Any]:
    role = normalize_role(approver.get("role"))
    if role not in {SUPER_ADMIN, SUB_DISTRIBUTOR, SUB_DISTRIBUTION_MANAGER}:
        raise PermissionError("You are not an approver for employee requests")

    conditions = []
    params: Dict[str, Any] = {}

    if role == SUB_DISTRIBUTOR:
        conditions.append("sub_distribution_id = :sd_id")
        params["sd_id"] = int(approver["id"])
    elif role == SUB_DISTRIBUTION_MANAGER:
        if not approver.get("parent_id"):
            return {"data": [], "pagination": get_pagination(page, page_size, 0)}
        conditions.append("sub_distribution_id = :sd_id")
        params["sd_id"] = int(approver["parent_id"])
    if status:
        conditions.append("status = :status")
        params["status"] = status
    if request_type:
        conditions.append("request_type = :request_type")
        params["request_type"] = request_type

    where = " AND ".join(conditions) if conditions else "1=1"
    async with async_session_factory() as session:
        rows_result = await _build_request_rows(session, where, params, page, page_size)
        if role != SUPER_ADMIN:
            eligible = [
                row for row in rows_result["data"]
                if await _is_eligible_approver(session, approver, row)
            ]
            rows_result["data"] = eligible
        return rows_result


async def get_request_detail(request_id: str, viewer: Dict[str, Any]) -> Dict[str, Any]:
    async with async_session_factory() as session:
        row = (await session.execute(
            text("SELECT * FROM approval_requests WHERE request_id = :rid"), {"rid": request_id}
        )).mappings().first()
        if not row:
            raise LookupError("Approval request not found")

        item = _serialize_request(dict(row))

        viewer_id = str(viewer.get("id"))
        viewer_role = normalize_role(viewer.get("role"))
        is_requester = str(row["requested_by"]) == viewer_id
        is_approver = await _is_eligible_approver(session, viewer, row)

        if not is_requester and not is_approver:
            raise PermissionError("You are not authorized to view this request")

        if is_approver and not is_requester:
            approvers = await _get_approver_users(session, int(row["sub_distribution_id"]))
            item["approver_can_approve"] = True
            item["approvers"] = approvers
        return item


async def cancel_request(request_id: str, requester: Dict[str, Any]) -> Dict[str, Any]:
    if normalize_role(requester.get("role")) != SUB_DISTRIBUTION_EMPLOYEE:
        raise PermissionError("Only the requesting employee can cancel an approval request")

    now = _now()
    async with async_session_factory() as session:
        result = await session.execute(
            text("""
                UPDATE approval_requests
                SET status = 'cancelled', updated_at = :now
                WHERE request_id = :rid AND requested_by = :rb AND status = 'pending'
            """),
            {"rid": request_id, "rb": int(requester["id"]), "now": now},
        )
        updated = result.rowcount
        if updated == 0:
            existing = (await session.execute(
                text("SELECT status FROM approval_requests WHERE request_id = :rid"), {"rid": request_id}
            )).mappings().first()
            if not existing:
                raise LookupError("Approval request not found")
            raise ValueError("Only pending requests can be cancelled")
        await session.commit()

    return {"success": True, "message": "Approval request cancelled"}


async def _execute_request(request_row: Dict[str, Any], requester: Dict[str, Any]) -> Dict[str, Any]:
    """Apply the approved proposal with fresh validation at approval time.

    Every branch re-runs the underlying service's own validation so stale or
    invalid payloads are surfaced here instead of being applied blindly.
    """
    request_type = request_row["request_type"]
    try:
        payload = json.loads(request_row["payload"]) if request_row["payload"] else {}
    except (json.JSONDecodeError, TypeError) as e:
        raise ValueError("Stored payload is invalid; ask the employee to resubmit") from e

    actor = {
        "id": requester["id"],
        "_id": requester["id"],
        "name": requester.get("name") or requester.get("email") or "Employee",
        "role": SUB_DISTRIBUTION_EMPLOYEE,
    }

    if request_type == "distribution":
        dist = await distribution_service.create_distribution(
            dist_data=DistributionCreate(**payload),
            from_user=actor,
        )
        return {"request_type": request_type, "result_id": dist.get("distribution_id") or dist.get("id")}
    if request_type == "defect":
        defect = await defect_service.create_defect(
            defect_data=DefectCreate(**payload),
            reporter=actor,
        )
        return {"request_type": request_type, "result_id": defect.get("report_id") or defect.get("id")}
    if request_type in {"cluster", "operator"}:
        create_payload = dict(payload)
        create_payload["role"] = request_type
        if request_type == "cluster":
            create_payload["parent_id"] = str(request_row["sub_distribution_id"])
            create_payload.pop("sub_distribution_id", None)
        if request_type == "operator" and not create_payload.get("parent_id"):
            create_payload["parent_id"] = str(request_row["sub_distribution_id"])
        user = await user_service.create_user(
            user_data=UserCreate(**create_payload),
            creator_id=int(requester["id"]),
        )
        return {"request_type": request_type, "result_id": str(user.get("id"))}

    # Branch actions run as the branch sub-distributor so role-gated services
    # behave as if the branch performed the approved action itself.
    async with async_session_factory() as session:
        branch_actor = await _branch_sub_distributor_actor(session, int(request_row["sub_distribution_id"]))

    if request_type == "user_update":
        allowed_fields = ("name", "phone", "designation", "address", "pincode", "status", "network_name")
        update_payload = {k: v for k, v in payload.items() if k in allowed_fields}
        updated = await user_service.update_user(str(payload["user_id"]), UserUpdate(**update_payload))
        if not updated:
            raise ValueError("Target user no longer exists")
        return {"request_type": request_type, "result_id": str(payload["user_id"])}
    if request_type == "user_delete":
        deleted = await user_service.delete_user(str(payload["user_id"]))
        if not deleted:
            raise ValueError("Target user no longer exists")
        return {"request_type": request_type, "result_id": str(payload["user_id"])}
    if request_type == "user_reassign":
        target_user = await user_service.get_user_by_id(str(payload["user_id"]))
        new_parent = await user_service.get_user_by_id(str(payload["new_parent_id"]))
        if not target_user or not new_parent:
            raise ValueError("Target user or new parent no longer exists")
        result = await user_service.reassign_user(
            user_id=str(payload["user_id"]),
            target_user=target_user,
            new_parent_id=str(payload["new_parent_id"]),
            new_parent=new_parent,
            performed_by=branch_actor,
        )
        return {"request_type": request_type, "result_id": str(payload["user_id"])}
    if request_type == "bulk_users":
        rows = _validate_bulk_rows(payload.get("rows"))
        bulk_result = await bulk_upload_service.process_bulk_user_upload(
            rows=rows,
            current_user=branch_actor,
            target_role=payload.get("role"),
            parent_id=int(payload["parent_id"]),
        )
        created = int((bulk_result or {}).get("created", 0) or 0)
        return {"request_type": request_type, "result_id": f"{created} users created", "created_count": created}
    if request_type == "bulk_distribution":
        rows = _validate_bulk_rows(payload.get("rows"))
        distribution_date = None
        if payload.get("date_of_distribution"):
            try:
                from datetime import date as _date
                distribution_date = _date.fromisoformat(str(payload["date_of_distribution"]).strip())
            except ValueError:
                raise ValueError("date_of_distribution must be in YYYY-MM-DD format")
        dist = await distribution_service.create_distribution_from_identifiers(
            to_user_id=str(payload["to_user_id"]),
            identifier_rows=rows,
            from_user=branch_actor,
            notes=payload.get("notes"),
            date_of_distribution=distribution_date,
        )
        return {
            "request_type": request_type,
            "result_id": dist.get("distribution_id") or dist.get("id"),
            "errors": (dist or {}).get("errors", []),
        }
    if request_type == "delivery_receipt":
        result = await distribution_service.confirm_receipt(
            distribution_id=str(payload["distribution_id"]),
            received=bool(payload.get("received")),
            user=branch_actor,
            notes=payload.get("notes"),
        )
        return {"request_type": request_type, "result_id": str(payload["distribution_id"])}
    if request_type == "return_status":
        result = await return_service.update_return_status(
            return_id=str(payload["return_id"]),
            status=str(payload["status"]),
            user=branch_actor,
            notes=payload.get("notes"),
            return_amount=payload.get("return_amount"),
            payment_bill_url=payload.get("payment_bill_url"),
        )
        if not result:
            raise ValueError("Return request no longer exists")
        return {"request_type": request_type, "result_id": str(payload["return_id"])}
    if request_type == "defect_status":
        result = await defect_service.update_defect_status(
            defect_id=str(payload["defect_id"]),
            status=str(payload["status"]),
            user=branch_actor,
            notes=payload.get("notes"),
            return_amount=payload.get("return_amount"),
        )
        if not result:
            raise ValueError("Defect report no longer exists")
        return {"request_type": request_type, "result_id": str(payload["defect_id"])}
    if request_type == "payment_confirmation":
        result = await defect_service.confirm_defect_payment(
            defect_id=str(payload["defect_id"]),
            confirmer=branch_actor,
            notes=payload.get("notes"),
        )
        if not result:
            raise ValueError("Defect report no longer exists")
        return {"request_type": request_type, "result_id": str(payload["defect_id"])}

    raise ValueError(f"Unsupported request type: {request_type}")


async def decide_request(
    approver: Dict[str, Any],
    request_id: str,
    action: str,
    review_note: Optional[str] = None,
) -> Dict[str, Any]:
    action = (action or "").strip().lower()
    if action not in {"approve", "reject"}:
        raise ValueError("action must be 'approve' or 'reject'")

    role = normalize_role(approver.get("role"))
    if role not in {SUPER_ADMIN, SUB_DISTRIBUTOR, SUB_DISTRIBUTION_MANAGER}:
        raise PermissionError("You are not an approver for employee requests")

    now = _now()
    approver_entry = {
        "role": role,
        "user_id": str(approver.get("id") or ""),
        "user_name": approver.get("name") or approver.get("email") or "Approver",
        "decision": action,
        "note": (review_note or "")[:1000],
        "decided_at": now.isoformat(),
    }

    async with async_session_factory() as session:
        row = (await session.execute(
            text("SELECT * FROM approval_requests WHERE request_id = :rid"), {"rid": request_id}
        )).mappings().first()
        if not row:
            raise LookupError("Approval request not found")
        request_row = dict(row)

        if not await _is_eligible_approver(session, approver, request_row):
            raise PermissionError("You are not an approver for this request")

        if request_row["status"] != "pending":
            raise ValueError("Request has already been reviewed")

        if action == "reject":
            await session.execute(
                text("""
                    UPDATE approval_requests
                    SET status = 'rejected', rejection_reason = :reason,
                        approvals = :approvals, updated_at = :now
                    WHERE id = :id AND status = 'pending'
                """),
                {
                    "id": request_row["id"],
                    "reason": (review_note or "")[:1000],
                    "approvals": json.dumps([approver_entry]),
                    "now": now,
                },
            )
            await session.commit()
            await _notify_requester(
                request_row,
                title="Approval Request Rejected",
                message=(
                    f"Your {request_row['request_type']} request ({request_id}) was rejected"
                    + (f" by {approver.get('name')}." if approver.get("name") else ".")
                    + (f" Reason: {review_note}" if review_note else "")
                ),
                notification_type="warning",
            )
            return {"success": True, "message": "Request rejected"}

        # approve
        try:
            existing_approvals = json.loads(request_row["approvals"]) if request_row["approvals"] else []
        except (json.JSONDecodeError, TypeError):
            existing_approvals = []

        already_decided = any(
            a.get("role") == role and str(a.get("user_id")) == str(approver.get("id"))
            for a in existing_approvals
            if isinstance(a, dict)
        )
        if already_decided:
            raise ValueError("You have already reviewed this request")

        new_approvals = existing_approvals + [approver_entry]
        approved_roles = {str(a.get("role")) for a in new_approvals if a.get("decision") == "approve"}

        try:
            required_roles = json.loads(request_row["required_roles"]) if request_row["required_roles"] else []
        except (json.JSONDecodeError, TypeError):
            required_roles = []

        all_approved = set(required_roles) <= approved_roles

        if not all_approved:
            await session.execute(
                text("UPDATE approval_requests SET approvals = :approvals, updated_at = :now WHERE id = :id AND status = 'pending'"),
                {"approvals": json.dumps(new_approvals), "now": now, "id": request_row["id"]},
            )
            await session.commit()
            requester = {
                "id": request_row["requested_by"],
                "name": request_row["requested_by_name"],
                "role": SUB_DISTRIBUTION_EMPLOYEE,
            }
            result = await _notify_requester(
                request_row,
                title="Approval Request Partially Approved",
                message=f"One approver approved your {request_row['request_type']} request ({request_id}). Waiting for the remaining approver.",
                notification_type="info",
                silent=True,
            )
            return {"success": True, "message": "Approval recorded; waiting for the remaining approver", "data": result}

        # Final approval -> execute with revalidation.
        requester = {
            "id": request_row["requested_by"],
            "name": request_row["requested_by_name"],
            "role": SUB_DISTRIBUTION_EMPLOYEE,
        }
        execution: Optional[Dict[str, Any]] = None
        execution_error: Optional[str] = None
        try:
            execution = await _execute_request(request_row, requester)
        except ValueError as e:
            execution_error = str(e)
        except Exception as e:  # pragma: no cover - defensive
            logger.exception("Approval request execution failed: %s", request_id)
            execution_error = "An internal error occurred while applying the request"

        if execution_error:
            await session.execute(
                text("""
                    UPDATE approval_requests
                    SET status = 'rejected', rejection_reason = :reason,
                        approvals = :approvals, updated_at = :now
                    WHERE id = :id AND status = 'pending'
                """),
                {
                    "id": request_row["id"],
                    "reason": f"Could not be applied: {execution_error}"[:1000],
                    "approvals": json.dumps(new_approvals),
                    "now": now,
                },
            )
            await session.commit()
            await _notify_requester(
                request_row,
                title="Approval Request Could Not Be Applied",
                message=(
                    f"Your {request_row['request_type']} request ({request_id}) was approved but "
                    f"could not be applied: {execution_error}"
                ),
                notification_type="warning",
            )
            return {"success": False, "message": "Approved but could not be applied", "data": {"error": execution_error}}

        await session.execute(
            text("""
                UPDATE approval_requests
                SET status = 'approved', approvals = :approvals,
                    execution_result = :execution_result, executed_at = :now, updated_at = :now
                WHERE id = :id AND status = 'pending'
            """),
            {
                "id": request_row["id"],
                "approvals": json.dumps(new_approvals),
                "execution_result": json.dumps(execution),
                "now": now,
            },
        )
        await session.commit()

    await _notify_requester(
        request_row,
        title="Approval Request Applied",
        message=f"Your {request_row['request_type']} request ({request_id}) was approved and applied.",
        notification_type="success",
    )
    return {"success": True, "message": "Request approved and applied", "data": execution}


async def _notify_requester(
    request_row: Dict[str, Any],
    *,
    title: str,
    message: str,
    notification_type: str,
    silent: bool = False,
):
    if silent:
        return None
    try:
        await notification_service.create_notification(
            user_id=int(request_row["requested_by"]),
            title=title,
            message=message,
            notification_type=notification_type,
            category="approval",
            link="/approval-requests",
            metadata={
                "action": "employee_approval_reviewed",
                "request_id": request_row.get("request_id"),
                "request_type": request_row.get("request_type"),
                "status": request_row.get("status"),
            },
        )
    except Exception:
        logger.exception("Failed to notify requester for request %s", request_row.get("request_id"))
    return None
