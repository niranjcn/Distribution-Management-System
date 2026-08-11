import logging
import json
import uuid
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, status, Depends, Query, UploadFile, File
from typing import Optional, Dict, Any, List
from app.models.device import DeviceCreate, DeviceUpdate, DeviceType, DeviceEditRequest
from app.services import device_service, notification_service, defect_service
from app.middleware.auth_middleware import get_current_user, require_admin_or_manager,require_management
from app.core.activity_logger import build_field_change_summary, log_business_activity
from app.core.cache_version import bump_cache_version
from app.utils.roles import normalize_role
from app.database_sqlalchemy import async_session_factory
from sqlalchemy import text
from app.services.bulk_upload_service import (
    BULK_UPLOAD_CHUNKED_COMMIT_THRESHOLD,
    build_bulk_result,
    chunked_executemany,
    chunks,
)

router = APIRouter()

logger = logging.getLogger(__name__)
MANAGEMENT_ROLES = {"super_admin", "md_director", "manager", "pdic_staff"}


def _extract_duplicate_message(error: Exception) -> str:
    lowered = str(error).lower()
    if "duplicate" not in lowered and "unique" not in lowered:
        return ""
    if "serial_number" in lowered:
        return "Serial number already exists"
    if "mac_address" in lowered:
        return "MAC address already exists"
    if "nuid" in lowered:
        return "NUID already exists"
    if "device_id" in lowered:
        return "Generated device ID collision. Retry upload."
    return "Duplicate value already exists"


def _build_bulk_device_id(device_type: str) -> str:
    # Avoid random 4-digit collisions during large imports.
    prefix_map = {
        "ONU": "ONU",
        "ONT": "ONT",
        "Router": "RTR",
        "Switch": "SWT",
        "Modem": "MDM",
        "Access Point": "AP",
        "Set-top box": "SB",
        "Other": "DEV",
    }
    prefix = prefix_map.get(device_type, "DEV")
    return f"{prefix}-{datetime.now().year}-{uuid.uuid4().hex[:10].upper()}"


async def _fetch_existing_values(session, column: str, values: List[str]) -> set:
    if not values:
        return set()

    existing = set()
    for batch in chunks(values, 1000):
        placeholders = ",".join([f":val{i}" for i in range(len(batch))])
        params = {f"val{i}": v for i, v in enumerate(batch)}
        result = await session.execute(
            text(f"SELECT {column} FROM devices WHERE {column} IN ({placeholders})"),
            params,
        )
        rows = result.mappings().all()
        for row in rows:
            value = row.get(column)
            if value:
                # Normalize to match the case-insensitive utf8mb4_general_ci
                # collation used by the unique indexes.
                existing.add(str(value).strip().lower())
    return existing


def _is_likely_text(content: bytes) -> bool:
    if not content:
        return True
    return b"\x00" not in content


def _validate_upload_signature(filename_lower: str, content: bytes) -> None:
    if filename_lower.endswith(".xlsx"):
        if not content.startswith(b"PK\x03\x04"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid XLSX file content"
            )
        return

    if filename_lower.endswith(".xls"):
        if not content.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid XLS file content"
            )
        return

    if filename_lower.endswith(".csv"):
        if not _is_likely_text(content[:2048]):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid CSV file content"
            )
        return


def _parse_csv_ids(raw_ids: Optional[str]) -> List[str]:
    if not raw_ids:
        return []
    return [item.strip() for item in str(raw_ids).split(",") if item and item.strip()]


async def _get_hierarchy_holder_ids(root_id: str) -> List[str]:
    """Return ids of the root user plus holder-tier users (sub_distributor,
    cluster, operator) under it, recursively."""
    normalized_id = str(root_id or "").strip()
    if not normalized_id or not normalized_id.isdigit():
        return []

    async with async_session_factory() as session:
        result = await session.execute(
            text("""
                WITH RECURSIVE descendants AS (
                    SELECT id, role FROM users WHERE id = :root
                    UNION ALL
                    SELECT u.id, u.role FROM users u
                    INNER JOIN descendants d ON u.parent_id = d.id
                )
                SELECT id FROM descendants
                WHERE role IN ('sub_distributor', 'cluster', 'operator')
            """),
            {"root": int(normalized_id)},
        )
        rows = result.mappings().all()
        return [str(row.get("id")) for row in rows if row.get("id") is not None]


async def _resolve_management_holder_scope(
    sub_distributor_id: Optional[str],
    cluster_id: Optional[str],
    operator_id: Optional[str] = None,
) -> Optional[List[str]]:
    sub_scope = None
    cluster_scope = None
    operator_scope = None

    normalized_sub_id = str(sub_distributor_id or "").strip()
    if normalized_sub_id:
        sub_scope = set(await _get_hierarchy_holder_ids(normalized_sub_id))

    normalized_cluster_id = str(cluster_id or "").strip()
    if normalized_cluster_id:
        cluster_scope = set(await _get_hierarchy_holder_ids(normalized_cluster_id))

    normalized_operator_id = str(operator_id or "").strip()
    if normalized_operator_id and normalized_operator_id.isdigit():
        operator_scope = {normalized_operator_id}

    scopes = [s for s in (sub_scope, cluster_scope, operator_scope) if s is not None]
    if not scopes:
        return None

    resolved = set(scopes[0])
    for scope in scopes[1:]:
        resolved = resolved.intersection(scope)
    return sorted(resolved)


@router.get("", summary="Get all devices with pagination and filters")
async def get_devices(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1),
    req_status: Optional[str] = Query(None, alias="status"),
    device_type: Optional[str] = None,
    manufacturer: Optional[str] = None,
    holder_id: Optional[str] = None,
    holder_ids: Optional[str] = None,
    search_by: Optional[str] = None,
    search: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all devices with pagination and filters"""
    try:
        parsed_holder_ids = _parse_csv_ids(holder_ids)

        # Filter by holder for non-admin/manager/staff users
        if current_user["role"] not in MANAGEMENT_ROLES:
            holder_id = current_user["id"]
            parsed_holder_ids = []

        result = await device_service.get_devices(
            page=page,
            page_size=page_size,
            status=req_status,
            device_type=device_type,
            manufacturer=manufacturer,
            holder_id=holder_id,
            holder_ids=parsed_holder_ids,
            search_by=search_by,
            search=search,
            start_date=start_date,
            end_date=end_date
        )

        return {
            "success": True,
            "message": "Devices retrieved successfully",
            "data": result["data"],
            "pagination": result["pagination"]
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.get("/for-replacement", summary="Get devices available as replacements with pagination and search.")
async def get_devices_for_replacement(
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1),
    exclude_device_id: Optional[str] = None,
    device_type: Optional[str] = None,
    search: Optional[str] = None,
    search_by: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get devices available as replacements (status=available or returned) with pagination and search.
    Management only - returns full stock regardless of holder. Used in the Replace Device modal."""
    if current_user["role"] not in ["super_admin", "manager", "pdic_staff"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only management can access replacement device pool"
        )
    try:
        result = await device_service.get_devices_for_replacement(
            exclude_device_id=exclude_device_id,
            page=page,
            page_size=page_size,
            device_type=device_type,
            search=search,
            search_by=search_by,
        )
        return {
            "success": True,
            "message": "Replacement-eligible devices retrieved successfully",
            "data": result["data"],
            "pagination": result["pagination"]
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.get("/available", summary="Get devices available to distribute for the current user.")
async def get_available_devices(
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=1000),
    search: Optional[str] = None,
    search_by: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get devices available to distribute for the current user.
    - admin/manager/staff: PDIC stock (status='available')
    - sub_distributor/cluster/operator: all devices they currently hold
    Supports pagination and server-side search across all matching devices."""
    try:
        role = current_user["role"]
        if role in ["super_admin", "md_director", "manager", "pdic_staff"]:
            result = await device_service.get_available_devices(
                holder_id=None,
                page=page,
                page_size=page_size,
                search=search,
                search_by=search_by,
            )
        elif role == "sub_distribution_employee":
            # Sub-distribution employees act on behalf of their assigned branch:
            # they can distribute the devices held by their parent sub-distributor
            # (and confirm them) rather than only devices personally assigned to them.
            holder_id = current_user.get("parent_id") or current_user["id"]
            result = await device_service.get_held_devices(
                holder_id=holder_id,
                page=page,
                page_size=page_size,
                search=search,
                search_by=search_by,
            )
        else:
            # Sub-level roles can redistribute any device they hold
            result = await device_service.get_held_devices(
                holder_id=current_user["id"],
                page=page,
                page_size=page_size,
                search=search,
                search_by=search_by,
            )

        return {
            "success": True,
            "message": "Available devices retrieved successfully",
            "data": result["data"],
            "pagination": result["pagination"],
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.get("/my-overview", summary="Get comprehensive device overview: devices in hand + under hierarchy + distribution stats.")
async def get_my_device_overview(
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1),
    show_all: bool = Query(False),
    paginate: bool = Query(False),
    limit: Optional[int] = Query(None, ge=1, le=5000),
    scope: str = Query("all"),
    req_status: Optional[str] = Query(None, alias="status"),
    device_type: Optional[str] = None,
    manufacturer: Optional[str] = None,
    sub_distributor_id: Optional[str] = None,
    cluster_id: Optional[str] = None,
    operator_id: Optional[str] = None,
    search_by: Optional[str] = None,
    search: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get comprehensive device overview: devices in hand + under hierarchy + distribution stats.
    - admin/manager/staff: all system devices with full stats
    - sub_distributor: held + cluster + operator devices in their chain
    - cluster: held + operator devices under them
    - operator: only their held devices"""
    try:
        role = current_user["role"]
        if role in MANAGEMENT_ROLES:
            effective_page_size = 1_000_000 if show_all else min(page_size, 1000)
            holder_scope = await _resolve_management_holder_scope(sub_distributor_id, cluster_id, operator_id)
            result = await device_service.get_devices(
                page=page,
                page_size=effective_page_size,
                status=req_status,
                device_type=device_type,
                manufacturer=manufacturer,
                holder_ids=holder_scope,
                search_by=search_by,
                search=search,
                start_date=start_date,
                end_date=end_date,
            )
            all_devices = result["data"]
            stats = await device_service.get_device_stats()
            insights = await device_service.get_management_insights()
            pagination = result.get("pagination", {})
            total_count = int(pagination.get("total", stats.get("total", 0) or 0))
            has_next = (not show_all) and ((page * effective_page_size) < total_count)
            return {
                "success": True,
                "data": {
                    "held_by_me": all_devices,
                    "under_subordinates": [],
                    "all_under_me": all_devices,
                    "meta": {
                        "page": page,
                        "page_size": effective_page_size,
                        "show_all": show_all,
                        "loaded_count": len(all_devices),
                        "total_count": total_count,
                        "has_next": has_next,
                    },
                    "insights": insights,
                    "stats": {
                        "in_my_hand": stats.get("total", 0),
                        "under_subordinates": 0,
                        "total_in_chain": stats.get("total", 0),
                        "total_devices_received": 0,
                        "total_devices_sent": 0,
                        "total_distributions_received": 0,
                        "total_distributions_sent": 0,
                        **stats
                    }
                }
            }
        else:
            scope_normalized = str(scope or "all").strip().lower()

            # Paginated requests slice the scoped chain in SQL (LIMIT/OFFSET in
            # MySQL) so only one page of rows is loaded per request. This avoids
            # loading the entire chain into memory for every page / "load more".
            if paginate:
                holder_scope = await _resolve_management_holder_scope(
                    sub_distributor_id, cluster_id, operator_id
                )
                result = await device_service.get_user_device_page(
                    user_id=current_user["id"],
                    user_role=role,
                    page=page,
                    page_size=page_size,
                    show_all=show_all,
                    scope=scope_normalized,
                    status=req_status,
                    device_type=device_type,
                    manufacturer=manufacturer,
                    search_by=search_by,
                    search=search,
                    start_date=start_date,
                    end_date=end_date,
                    holder_ids=holder_scope,
                )
                page_devices = result["page_devices"]
                total_count = int(result["total_count"])
                effective_page_size = result["page_size_used"]
                holder_id_str = str(current_user["id"])
                held = [
                    d for d in page_devices
                    if str(d.get("current_holder_id") or "") == holder_id_str
                ]
                subordinate = [
                    d for d in page_devices
                    if str(d.get("current_holder_id") or "") != holder_id_str
                ]
                has_next = (not show_all) and ((page * effective_page_size) < total_count)

                return {
                    "success": True,
                    "data": {
                        "held_by_me": held,
                        "under_subordinates": subordinate,
                        "all_under_me": page_devices,
                        "meta": {
                            "page": page,
                            "page_size": effective_page_size,
                            "show_all": show_all,
                            "loaded_count": len(page_devices),
                            "total_count": total_count,
                            "has_next": has_next,
                        },
                        "stats": result["stats"],
                        "insights": result.get("insights", {"by_type": [], "by_vendor": []}),
                    },
                }

            # Non-paginated overview (overview cards / dashboards): fetch a small
            # window plus cheap aggregate stats.
            effective_limit = limit if limit is not None else (
                10 if show_all else min(page_size, 1000)
            )
            overview = await device_service.get_user_device_overview(
                user_id=current_user["id"],
                user_role=role,
                limit=effective_limit
            )
            if scope_normalized == "mine":
                chain_devices = overview.get("held_by_me") or []
            elif scope_normalized == "hierarchy":
                chain_devices = overview.get("under_subordinates") or []
            else:
                chain_devices = overview.get("all_under_me") or []

            if start_date:
                chain_devices = [
                    d for d in chain_devices
                    if str(d.get("created_at") or "") >= start_date
                ]
            if end_date:
                chain_devices = [
                    d for d in chain_devices
                    if str(d.get("created_at") or "") <= end_date
                ]

            stats = overview.get("stats") or {}
            total_count = int(stats.get("total_in_chain", len(chain_devices)) or len(chain_devices))
            overview["meta"] = {
                "page": 1,
                "page_size": len(chain_devices),
                "show_all": show_all,
                "loaded_count": len(chain_devices),
                "total_count": total_count,
                "has_next": False,
            }
            return {"success": True, "data": overview}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.get("/management-holder-insights", summary="Get management holder insights")
async def get_management_holder_insights(
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in MANAGEMENT_ROLES:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only management can access holder insights"
        )
    try:
        insights = await device_service.get_management_holder_insights()
        return {
            "success": True,
            "data": insights,
        }
    except HTTPException:
        raise
    except Exception:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.post("/{device_id}/repair-holder", summary="Admin/Manager: repair a device's current_holder by replaying the most recent")
async def repair_device_holder(
    device_id: str,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Admin/Manager: repair a device's current_holder by replaying the most recent
    distributed history entry. Use when a double-approval has overwritten the holder."""
    try:
        device = await device_service.repair_device_holder_from_history(device_id)
        if not device:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found or no distribution history available"
            )
        return {
            "success": True,
            "message": "Device holder repaired successfully",
            "data": device
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )



@router.post("/{device_id}/request-edit", summary="PDIC staff: request an edit to a device.")
async def request_device_edit(
    device_id: str,
    payload: DeviceEditRequest,
    current_user: dict = Depends(get_current_user)
):
    """PDIC staff: request an edit to a device.
    The device is not modified until a manager or super admin approves it from Change Requests."""
    if current_user["role"] != "pdic_staff":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only PDIC staff can submit device edit approval requests"
        )

    try:
        if not str(device_id).isdigit():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid device id")

        device = await device_service.get_device_by_id(device_id)
        if not device:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Device not found")

        proposed_changes = payload.model_dump(exclude_none=True)
        if not proposed_changes:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No valid device edit fields provided")

        now = datetime.now().replace(tzinfo=None)
        request_id = f"CR-{uuid.uuid4().hex[:8].upper()}"

        async with async_session_factory() as session:
            await session.execute(
                text("""INSERT INTO change_requests
                   (request_id, requested_by, requested_by_name, requested_by_role,
                    request_type, device_id, reason, status, created_at, updated_at)
                   VALUES (:request_id, :requested_by, :requested_by_name, :requested_by_role, 'device_edit_change', :device_id, :reason, 'pending', :created_at, :updated_at)"""),
                {
                    "request_id": request_id,
                    "requested_by": int(current_user["id"]),
                    "requested_by_name": current_user.get("name") or current_user.get("email") or "PDIC Staff",
                    "requested_by_role": current_user.get("role"),
                    "device_id": str(device_id),
                    "reason": json.dumps({"changes": proposed_changes}),
                    "created_at": now,
                    "updated_at": now,
                }
            )

            result = await session.execute(text("SELECT id FROM users WHERE role IN ('super_admin', 'manager') AND status = 'active'"))
            reviewer_rows = result.mappings().all()
            reviewer_ids = [int(row["id"]) for row in reviewer_rows]
            await bump_cache_version(session)
            await session.commit()

        proposer_name = current_user.get("name") or current_user.get("email", "pdic_staff")
        changes_summary = ", ".join(
            f"{k}: '{v}'" for k, v in proposed_changes.items() if v
        )
        message = (
            f"Staff Edit Request from {proposer_name}:\n"
            f"Device: {device.get('device_id')} (Serial: {device.get('serial_number')})\n"
            f"Proposed Changes: {changes_summary or 'No changes specified'}"
        )

        notified_count = len(reviewer_ids)
        notification_failures = 0
        try:
            notifications = [
                {
                    "user_id": reviewer_id,
                    "title": "Device Edit Approval Request",
                    "message": message,
                    "notification_type": "device_edit_request",
                    "category": "approval",
                    "link": "/devices/edit-requests",
                    "metadata": {
                        "action": "device_edit_change",
                        "request_id": request_id,
                        "device_id": str(device_id),
                        "requested_by": str(current_user.get("id") or ""),
                    },
                }
                for reviewer_id in reviewer_ids
            ]
            await notification_service.bulk_create_notifications(notifications)
        except Exception:
            notification_failures = notified_count
            notified_count = 0
            logger.exception(
                "Failed sending device edit notifications | request_id=%s",
                request_id,
            )

        message_text = f"Edit request submitted. Sent to {notified_count} reviewer(s) for approval."
        if notification_failures:
            message_text += f" {notification_failures} notification(s) failed to deliver."

        return {
            "success": True,
            "message": message_text,
            "data": {
                "request_id": request_id,
                "notified_reviewers": notified_count,
                "notification_failures": notification_failures,
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.get("/track/{serial_number}", summary="Track device by serial number, NUID, or MAC with full history")
async def track_device_by_serial(
    serial_number: str,
    current_user: dict = Depends(get_current_user)
):
    """Track device by serial number, NUID, or MAC with full history"""
    try:
        device = await device_service.track_device_by_serial(
            serial_number,
            user_id=str(current_user.get("id", current_user.get("_id", ""))),
            user_role=str(current_user.get("role", "")).lower(),
        )

        if not device:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found"
            )

        return {
            "success": True,
            "message": "Device tracked successfully",
            "data": device
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.get("/{device_id}", summary="Get device by ID")
async def get_device(
    device_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get device by ID"""
    try:
        device = await device_service.get_device_by_id(device_id)

        if not device:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found"
            )

        if not await device_service.user_can_view_device(current_user, device):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Insufficient permissions"
            )

        return {
            "success": True,
            "message": "Device retrieved successfully",
            "data": device
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.get("/{device_id}/history", summary="Get device history")
async def get_device_history(
    device_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get device history"""
    try:
        device = await device_service.get_device_by_id(device_id)

        if not device:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found"
            )

        if not await device_service.user_can_view_device(current_user, device):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Insufficient permissions"
            )

        history = await device_service.get_device_history(device_id)

        return {
            "success": True,
            "message": "Device history retrieved successfully",
            "data": history
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.post("/bulk-upload", status_code=status.HTTP_201_CREATED, summary="Bulk upload devices from an Excel file.")
async def bulk_upload_devices(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_management)
):
    """Bulk upload devices from an Excel file.

    Supported schemas (case-insensitive headers):
    - SB sheet: vendor, device_type, model, nuid, box_type
    - Regular sheet: vendor, device_type, model, serial_number, mac_address, band_type

    Alias support: manufacturer -> vendor, SB/set-top box/set top box/stb -> Set-top box
    """
    filename_lower = file.filename.lower()
    if not filename_lower.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel (.xlsx, .xls) or CSV (.csv) files are supported"
        )

    actor_name = current_user.get("name") or current_user.get("email") or "User"

    async def _log_bulk_upload_summary(created_count: int, skipped_count: int, error_count: int) -> None:
        await log_business_activity(
            user=current_user,
            path="/activity/devices/bulk-upload",
            description=(
                f"{actor_name} used bulk upload for devices: "
                f"{created_count} created, {skipped_count} skipped, {error_count} errors"
            ),
        )

    try:
        import io

        contents = await file.read()

        from app.services.bulk_upload_service import (
            check_bulk_upload_file,
            check_bulk_upload_file_row_count,
            check_bulk_upload_row_count,
            MAX_BULK_ROWS,
        )
        check_bulk_upload_file(contents, filename_lower)

        _validate_upload_signature(filename_lower, contents)
        check_bulk_upload_file_row_count(contents, filename_lower)

        if filename_lower.endswith('.csv'):
            import csv
            decoded = contents.decode('utf-8-sig')  # strip BOM if present
            reader = csv.reader(io.StringIO(decoded))
            all_rows = list(reader)
            if not all_rows:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV file is empty")
            headers = [h.strip().lower() for h in all_rows[0]]
            data_rows = all_rows[1:]
            check_bulk_upload_row_count(data_rows)

            def iter_data_rows():
                for row in data_rows:
                    # Pad short rows to header length
                    padded = row + [''] * (len(headers) - len(row))
                    yield tuple(padded[:len(headers)])
        else:
            if filename_lower.endswith('.xls'):
                import xlrd
                wb = xlrd.open_workbook(file_contents=contents)
                ws = wb.sheet_by_index(0)
                check_bulk_upload_row_count(range(ws.nrows - 1))
                headers = [str(ws.cell_value(0, col)).strip().lower() for col in range(ws.ncols)]

                def iter_data_rows():
                    for row_idx in range(1, ws.nrows):
                        yield tuple(ws.cell_value(row_idx, col) for col in range(ws.ncols))
            else:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
                ws = wb.active
                headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

                row_count = 0

                def iter_data_rows():
                    nonlocal row_count
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_count += 1
                        if row_count > MAX_BULK_ROWS:
                            raise HTTPException(
                                status_code=status.HTTP_400_BAD_REQUEST,
                                detail=f"Too many rows. Maximum is {MAX_BULK_ROWS}",
                            )
                        yield row

        normalized_headers = ["vendor" if h == "manufacturer" else h for h in headers]
        header_set = set(normalized_headers)

        sb_required = {"vendor", "device_type", "model", "nuid", "box_type"}
        regular_required = {"vendor", "device_type", "model", "serial_number"}

        has_sb_schema = sb_required.issubset(header_set)
        has_regular_schema = regular_required.issubset(header_set)
        if not has_sb_schema and not has_regular_schema:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "Missing required columns. Expected either SB schema "
                    "(vendor, device_type, model, nuid, box_type) or regular schema "
                    "(vendor, device_type, model, serial_number). mac_address is optional for non-SB devices."
                )
            )

        valid_types = {t.value.lower(): t.value for t in DeviceType}
        valid_types.update({
            "sb": DeviceType.SETUP_BOX.value,
            "set top box": DeviceType.SETUP_BOX.value,
            "stb": DeviceType.SETUP_BOX.value,
        })
        valid_bands = {
            "single_band": "single_band",
            "single band": "single_band",
            "single": "single_band",
            "dual_band": "dual_band",
            "dual band": "dual_band",
            "dual": "dual_band",
        }
        created, skipped, errors = [], [], []
        prepared_rows = []
        seen_serials = set()
        seen_macs = set()
        seen_nuids = set()

        for row_idx, row in enumerate(iter_data_rows(), start=2):
            if row is None:
                row = ()

            row_data = {
                normalized_headers[i]: (str(row[i]).strip() if i < len(row) and row[i] is not None else "")
                for i in range(len(normalized_headers))
            }

            # Skip completely empty rows
            if not any(row_data.values()):
                continue

            # Normalise device_type
            raw_type = row_data.get("device_type", "").lower()
            device_type_val = valid_types.get(raw_type)
            if not device_type_val:
                errors.append({"row": row_idx, "error": f"Invalid device_type '{row_data.get('device_type')}'"})
                continue

            is_sb_row = device_type_val == DeviceType.SETUP_BOX.value
            vendor = row_data.get("vendor", "")
            model = row_data.get("model", "")
            nuid = row_data.get("nuid", "")
            serial = row_data.get("serial_number", "")
            mac = row_data.get("mac_address", "")

            if not vendor:
                errors.append({"row": row_idx, "serial": serial or nuid or "", "error": "Missing vendor"})
                continue
            if not model:
                errors.append({"row": row_idx, "serial": serial or nuid or "", "error": "Missing model"})
                continue

            if is_sb_row:
                if not nuid:
                    errors.append({"row": row_idx, "error": "Missing nuid for SB row"})
                    continue
                box_type = str(row_data.get("box_type", "")).strip().upper()
                if not box_type:
                    errors.append({"row": row_idx, "error": "Missing box_type for SB row"})
                    continue
                if box_type not in {"HD", "OTT"}:
                    errors.append({"row": row_idx, "error": "Invalid box_type. Use HD or OTT"})
                    continue
                band_type_val = None
                normalized_nuid = str(nuid).strip().lower()
                if normalized_nuid in seen_nuids:
                    skipped.append({"row": row_idx, "serial": normalized_nuid, "reason": "Duplicate nuid in file"})
                    continue
                seen_nuids.add(normalized_nuid)
            else:
                if not serial:
                    errors.append({"row": row_idx, "error": "Missing serial_number"})
                    continue
                raw_band = row_data.get("band_type", "").lower()
                band_type_val = valid_bands.get(raw_band) if raw_band else None
                if raw_band and not band_type_val:
                    errors.append({"row": row_idx, "serial": serial, "error": f"Invalid band_type '{row_data.get('band_type')}'"})
                    continue

                normalized_serial = str(serial).strip().lower()
                if normalized_serial in seen_serials:
                    skipped.append({"row": row_idx, "serial": normalized_serial, "reason": "Duplicate serial_number in file"})
                    continue
                seen_serials.add(normalized_serial)

                normalized_mac = str(mac).strip().lower()
                if normalized_mac:
                    if normalized_mac in seen_macs:
                        skipped.append({"row": row_idx, "serial": normalized_serial, "reason": "Duplicate mac_address in file"})
                        continue
                    seen_macs.add(normalized_mac)

            prepared_rows.append({
                "row": row_idx,
                "device_type": device_type_val,
                "model": model,
                "manufacturer": vendor,
                "serial_number": None if is_sb_row else str(serial).strip(),
                "mac_address": None if is_sb_row else (str(mac).strip() or None),
                "band_type": None if is_sb_row else (band_type_val or "single_band"),
                "nuid": str(nuid).strip() or None,
                "box_type": box_type if is_sb_row else None,
                "metadata": None,
                "is_sb": is_sb_row,
            })

        if not prepared_rows:
            await _log_bulk_upload_summary(0, len(skipped), len(errors))
            return {
                "success": True,
                "message": f"Bulk upload complete: 0 created, {len(skipped)} skipped, {len(errors)} errors",
                "data": build_bulk_result(created, skipped, errors),
            }

        async with async_session_factory() as session:
            serials = [item["serial_number"] for item in prepared_rows if item["serial_number"]]
            macs = [item["mac_address"] for item in prepared_rows if item["mac_address"]]
            nuids = [item["nuid"] for item in prepared_rows if item["nuid"]]

            existing_serials = await _fetch_existing_values(session, "serial_number", serials)
            existing_macs = await _fetch_existing_values(session, "mac_address", macs)
            existing_nuids = await _fetch_existing_values(session, "nuid", nuids)

            insertable_rows = []
            for item in prepared_rows:
                row_idx = item["row"]
                serial = item["serial_number"]
                mac = item["mac_address"]
                nuid = item["nuid"]

                if nuid and nuid.lower() in existing_nuids:
                    skipped.append({"row": row_idx, "serial": nuid, "reason": "NUID already exists"})
                    continue
                if serial and serial.lower() in existing_serials:
                    skipped.append({"row": row_idx, "serial": serial, "reason": "Serial number already exists"})
                    continue
                if mac and mac.lower() in existing_macs:
                    skipped.append({"row": row_idx, "serial": serial or nuid or "", "reason": "MAC address already exists"})
                    continue

                insertable_rows.append(item)

            if not insertable_rows:
                await _log_bulk_upload_summary(0, len(skipped), len(errors))
                return {
                    "success": True,
                    "message": f"Bulk upload complete: 0 created, {len(skipped)} skipped, {len(errors)} errors",
                    "data": build_bulk_result(created, skipped, errors),
                }

            now = datetime.now().replace(tzinfo=None)
            created_by_name = current_user.get("name") or current_user.get("email") or "PDIC Staff"
            insert_sql = """INSERT INTO devices (
                device_id, device_type, model, serial_number, mac_address,
                manufacturer, band_type, nuid, box_type, status, current_location,
                current_holder_id, current_holder_name, current_holder_type,
                registered_by_name, purchase_date, warranty_expiry, metadata,
                created_at, updated_at
            ) VALUES (:device_id, :device_type, :model, :serial_number, :mac_address, :manufacturer, :band_type, :nuid, :box_type, :status, :current_location, :current_holder_id, :current_holder_name, :current_holder_type, :registered_by_name, :purchase_date, :warranty_expiry, :metadata, :created_at, :updated_at)"""

            partial = False
            use_chunked_commits = len(insertable_rows) > BULK_UPLOAD_CHUNKED_COMMIT_THRESHOLD

            payload_rows = []
            for item in insertable_rows:
                generated_id = _build_bulk_device_id(item["device_type"])
                payload_rows.append({
                    "row": item["row"],
                    "_generated_id": generated_id,
                    "device_id": generated_id,
                    "device_type": item["device_type"],
                    "model": item["model"],
                    "serial_number": item["serial_number"],
                    "mac_address": item["mac_address"],
                    "manufacturer": item["manufacturer"],
                    "band_type": item["band_type"],
                    "nuid": item["nuid"],
                    "box_type": item["box_type"],
                    "status": "available",
                    "current_location": "PDIC",
                    "current_holder_id": None,
                    "current_holder_name": "PDIC (Distribution)",
                    "current_holder_type": "noc",
                    "registered_by_name": created_by_name,
                    "purchase_date": None,
                    "warranty_expiry": None,
                    "metadata": json.dumps(item["metadata"]) if item["metadata"] else None,
                    "created_at": now,
                    "updated_at": now,
                })

            # Tracks how many rows were recorded for the current batch so an
            # unexpected error can drop them from `created` (they are rolled
            # back with the transaction).
            batch_created_start = [len(created)]

            async def _device_batch_success(session, batch):
                # Insert history for every device that was actually created.
                # Joining back to `devices` avoids a separate
                # device_id -> numeric id lookup round trip per batch.
                batch_device_ids = [item["_generated_id"] for item in batch]
                history_device_ids = ",".join([f":dev{i}" for i in range(len(batch_device_ids))])
                history_params = {f"dev{i}": bid for i, bid in enumerate(batch_device_ids)}
                history_params["performed_by"] = int(current_user["id"])
                history_params["performed_by_name"] = created_by_name
                history_params["timestamp"] = now
                await session.execute(
                    text(f"""INSERT INTO device_history (
                        device_id, action, from_user_id, from_user_name, to_user_id, to_user_name,
                        status_before, status_after, location, notes, performed_by, performed_by_name, timestamp
                    )
                    SELECT d.id, 'bulk_registered', NULL, NULL, NULL, NULL, NULL, 'available', 'PDIC',
                           'Device registered in system', :performed_by, :performed_by_name, :timestamp
                    FROM devices d
                    WHERE d.device_id IN ({history_device_ids})"""),
                    history_params,
                )
                created.extend(batch_device_ids)

            async def _device_row_duplicate(session, item, err):
                skipped.append({
                    "row": item["row"],
                    "serial": item.get("serial_number") or item.get("nuid") or "",
                    "reason": _extract_duplicate_message(err),
                })

            async def _device_row_error(session, item, err):
                errors.append({
                    "row": item["row"],
                    "serial": item.get("serial_number") or item.get("nuid") or "",
                    "error": str(err),
                })
                # Rows inserted earlier in this batch are rolled back with the
                # transaction, so drop them from the created list to keep the
                # reported counts honest.
                del created[batch_created_start[0]:]

            async def _device_batch_start():
                batch_created_start[0] = len(created)

            async def _device_batch_complete(ok, can_commit):
                if use_chunked_commits and can_commit:
                    await bump_cache_version(session)
                    await session.commit()

            should_commit = await chunked_executemany(
                session,
                insert_sql,
                payload_rows,
                on_batch_success=_device_batch_success,
                on_row_duplicate=_device_row_duplicate,
                on_row_error=_device_row_error,
                on_batch_start=_device_batch_start,
                on_batch_complete=_device_batch_complete,
                abort_on_error=True,
            )

            if use_chunked_commits:
                # Prior batches are already committed. Roll back any in-flight
                # changes from the batch that hit an unexpected error and report
                # a partial result so the user can fix the remaining rows.
                if not should_commit:
                    await session.rollback()
                    partial = True
            elif should_commit:
                await bump_cache_version(session)
                await session.commit()
            else:
                await session.rollback()
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Bulk upload was rolled back due to an unexpected insert error. Please retry."
                )

        await _log_bulk_upload_summary(len(created), len(skipped), len(errors))

        if partial:
            message = (
                f"Bulk upload partially completed: {len(created)} created, "
                f"{len(skipped)} skipped, {len(errors)} errors (stopped at an unexpected insert error). "
                "Already-imported rows will be skipped on retry."
            )
        else:
            message = f"Bulk upload complete: {len(created)} created, {len(skipped)} skipped, {len(errors)} errors"

        data = build_bulk_result(created, skipped, errors)
        data["partial"] = partial
        return {
            "success": True,
            "message": message,
            "data": data,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.post("", status_code=status.HTTP_201_CREATED, summary="Register a new device")
async def create_device(
    device_data: DeviceCreate,
    current_user: dict = Depends(get_current_user)
):
    """Register a new device"""
    try:
        role = normalize_role(current_user.get("role"))
        if role not in {"super_admin", "manager", "pdic_staff"}:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only management users can register devices"
            )

        device = await device_service.create_device(
            device_data=device_data,
            created_by=int(current_user["id"]),
            created_by_name=(current_user.get("name") or current_user.get("email") or "PDIC Staff")
        )

        actor_name = current_user.get("name") or current_user.get("email") or "User"
        await log_business_activity(
            user=current_user,
            path="/activity/devices/create",
            description=(
                f"{actor_name} added device {device.get('device_id') or device.get('id')} "
                f"({device.get('device_type') or 'Unknown type'})"
            ),
        )

        return {
            "success": True,
            "message": "Device registered successfully",
            "data": device
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.put("/{device_id}", summary="Update device")
async def update_device(
    device_id: str,
    device_data: DeviceUpdate,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Update device"""
    try:
        before = await device_service.get_device_by_id(device_id)
        device = await device_service.update_device(device_id, device_data)

        if not device:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found"
            )

        actor_name = current_user.get("name") or current_user.get("email") or "User"
        edited_fields = list(device_data.model_dump(exclude_unset=True).keys())
        change_summary = build_field_change_summary(
            before=before or {},
            after=device or {},
            fields=edited_fields,
            exclude_fields={"updated_at"},
        )
        await log_business_activity(
            user=current_user,
            path="/activity/devices/update",
            description=(
                f"{actor_name} updated device "
                f"{device.get('device_id') or before.get('device_id') if before else device_id}; "
                f"changes: {change_summary}"
            ),
        )

        return {
            "success": True,
            "message": "Device updated successfully",
            "data": device
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.delete("/{device_id}", summary="Delete device")
async def delete_device(
    device_id: str,
    current_user: dict = Depends(require_admin_or_manager)
):
    """Delete device"""
    try:
        device = await device_service.get_device_by_id(device_id)
        success = await device_service.delete_device(device_id)

        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found"
            )

        actor_name = current_user.get("name") or current_user.get("email") or "User"
        device_ref = (device or {}).get("device_id") or device_id
        await log_business_activity(
            user=current_user,
            path="/activity/devices/delete",
            description=f"{actor_name} deleted device {device_ref}",
        )

        return {
            "success": True,
            "message": "Device deleted successfully"
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.post("/bulk-delete", summary="Bulk delete devices (super admin / manager only)")
async def bulk_delete_devices(
    payload: Dict[str, Any],
    current_user: dict = Depends(require_admin_or_manager),
):
    """Delete multiple devices at once. Staff must use the approval flow instead."""
    device_ids = payload.get("device_ids") or []
    if not device_ids or not isinstance(device_ids, list):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="device_ids is required")

    try:
        result = await device_service.bulk_delete_devices([str(i) for i in device_ids])

        deleted_count = len(result["deleted"])
        if deleted_count > 0:
            actor_name = current_user.get("name") or current_user.get("email") or "User"
            await log_business_activity(
                user=current_user,
                path="/activity/devices/bulk-delete",
                description=f"{actor_name} bulk deleted {deleted_count} device(s)",
            )

        return {
            "success": True,
            "message": f"{deleted_count} device(s) deleted successfully",
            "data": result,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.patch("/{device_id}/status", summary="Update device status")
async def update_device_status(
    device_id: str,
    status_update: dict,
    current_user: dict = Depends(get_current_user)
):
    """Update device status"""
    status_value = status_update.get("status")
    notes = status_update.get("notes")
    
    if not status_value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Status is required"
        )

    try:
        if current_user.get("role") == "md_director":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="MD/Director has read-only access to devices"
            )
        if current_user.get("role") == "sub_distribution_manager":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Sub Distribution MD/Manager has read-only access to devices"
            )

        before = await device_service.get_device_by_id(device_id)

        device = await device_service.update_device_status(
            device_id=device_id,
            status=status_value,
            performed_by=int(current_user["id"]),
            performed_by_name=current_user["name"],
            notes=notes
        )

        if str(status_value).lower() == "defective":
            await defect_service.create_or_get_active_defect_for_device(
                device_id=device_id,
                reporter=current_user,
                notes=notes
            )

        if not device:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Device not found"
            )

        actor_name = current_user.get("name") or current_user.get("email") or "User"
        change_summary = build_field_change_summary(
            before=before or {},
            after=device or {},
            fields=["status"],
            exclude_fields={"updated_at"},
        )
        await log_business_activity(
            user=current_user,
            path="/activity/devices/status-update",
            description=(
                f"{actor_name} updated device status for "
                f"{device.get('device_id') or (before or {}).get('device_id') or device_id}; "
                f"changes: {change_summary}"
            ),
        )

        return {
            "success": True,
            "message": "Device status updated successfully",
            "data": device
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )




