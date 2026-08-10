import logging
import csv
import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from sqlalchemy import text

from app.core.activity_logger import build_field_change_summary, log_business_activity

from app.middleware.auth_middleware import require_any_role, require_management
from app.models.inventory import (
    ExternalBulkDistributionCreate,
    ExternalDistributionCreate,
    InventoryItemCreate,
    InventoryItemUpdate,
)
from app.services import inventory_service
from app.services.bulk_upload_service import (
    BULK_UPLOAD_CHUNKED_COMMIT_THRESHOLD,
    build_bulk_result,
    chunked_executemany,
    chunks,
)

router = APIRouter()
logger = logging.getLogger(__name__)


def _is_management_user(current_user: dict) -> bool:
    role = str(current_user.get("role") or "").strip().lower()
    return role in {"super_admin", "manager", "pdic_staff"}


def _resolve_actor_id(current_user: dict) -> int:
    actor_id = (
        current_user.get("id")
        or current_user.get("_id")
        or current_user.get("user_id")
        or current_user.get("sub")
        or 0
    )
    if not actor_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authenticated user id is missing",
        )
    return int(actor_id)


def _read_bulk_rows(filename_lower: str, contents: bytes) -> tuple:
    """Normalize an Excel (.xls/.xlsx) workbook into ``(headers, data_rows)``.

    Row values are returned as lists so callers can pad short rows to the header
    count in the same way CSV rows are handled. ``headers`` are stripped and
    lowercased. The caller is responsible for size/row-count limits.
    """
    if filename_lower.endswith(".xls"):
        import xlrd

        wb = xlrd.open_workbook(file_contents=contents)
        ws = wb.sheet_by_index(0)
        if ws.nrows == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file is empty",
            )
        headers = [str(ws.cell_value(0, col)).strip().lower() for col in range(ws.ncols)]
        rows = [
            [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
            for row_idx in range(1, ws.nrows)
        ]
        return headers, rows

    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    worksheet = workbook.active
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file is empty",
        )
    headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]
    rows = [list(row) for row in worksheet.iter_rows(min_row=2, values_only=True)]
    workbook.close()
    return headers, rows


@router.get("/items", summary="Get external inventory items")
async def get_external_inventory_items(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1),
    search: Optional[str] = None,
    search_by: Optional[str] = Query("all"),
    device_type: Optional[str] = Query(None, alias="type"),
    status_filter: Optional[str] = Query(None, alias="status"),
    identifier_type: Optional[str] = Query(None),
    warranty: Optional[str] = Query(None),
    current_user: dict = Depends(require_any_role),
):
    try:
        result = await inventory_service.get_items(
            page=page,
            page_size=page_size,
            search=search,
            search_by=search_by,
            device_type=device_type,
            status_filter=status_filter,
            identifier_type=identifier_type,
            warranty=warranty,
            management=_is_management_user(current_user),
        )
        return {
            "success": True,
            "message": "External inventory items retrieved successfully",
            "data": result["data"],
            "pagination": result["pagination"],
        }
    except HTTPException:
        raise
    except Exception:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later.",
        )


@router.post("/items", status_code=status.HTTP_201_CREATED, summary="Create external inventory item")
async def create_external_inventory_item(
    item_data: InventoryItemCreate,
    current_user: dict = Depends(require_management),
):
    try:
        item = await inventory_service.create_item(item_data=item_data, user=current_user)
        actor_name = current_user.get("name") or current_user.get("email") or "User"
        await log_business_activity(
            user=current_user,
            path="/activity/external-inventory/item-create",
            description=f"{actor_name} created external inventory item {item.get('name') or 'unknown'}",
        )
        return {
            "success": True,
            "message": "External inventory item created successfully",
            "data": item,
        }
    except HTTPException:
        raise
    except Exception:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later.",
        )


@router.post("/items/bulk-upload", status_code=status.HTTP_201_CREATED, summary="Bulk upload external inventory items from CSV")
async def bulk_upload_external_inventory_items(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_management),
):
    """Bulk upload external inventory items from CSV or Excel.

    Required columns: name
    Optional columns: identifier_type, identifier, device_type, price, quantity,
    supplier_name, location, warranty_start_date (YYYY-MM-DD), warranty_duration (months),
    notes
    """
    filename_lower = (file.filename or "").lower()
    if not filename_lower.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel (.xlsx, .xls) or CSV (.csv) files are supported",
        )

    try:
        contents = await file.read()

        from app.services.bulk_upload_service import (
            check_bulk_upload_file,
            check_bulk_upload_row_count,
            validate_upload_signature,
        )
        # Enforces the 10 MB size cap, the xlsx decompressed-size (zip bomb) guard,
        # and rejects mis-typed files via magic-byte signature checks.
        check_bulk_upload_file(contents, filename_lower)
        validate_upload_signature(filename_lower, contents)

        if filename_lower.endswith(".csv"):
            decoded = contents.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(decoded))
            all_rows = list(reader)
            if not all_rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="CSV file is empty",
                )
            headers = [h.strip().lower() for h in all_rows[0]]
            data_rows = all_rows[1:]
        else:
            headers, data_rows = _read_bulk_rows(filename_lower, contents)
        check_bulk_upload_row_count(data_rows)

        required = {"name"}
        missing = required - set(headers)
        if missing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing required columns: {', '.join(sorted(missing))}",
            )

        created = []
        skipped = []
        errors = []
        prepared_rows = []
        seen_identifiers = set()

        for row_idx, row in enumerate(data_rows, start=2):
            padded = row + [""] * (len(headers) - len(row))
            row_data = {
                headers[i]: (str(padded[i]).strip() if padded[i] is not None else "")
                for i in range(len(headers))
            }

            if not any(row_data.values()):
                continue

            name = row_data.get("name", "").strip()
            if not name:
                errors.append({"row": row_idx, "name": "", "error": "Missing name"})
                continue

            identifier_type_value = row_data.get("identifier_type", "").strip()
            identifier_value = row_data.get("identifier", "").strip()
            if identifier_value and not identifier_type_value:
                errors.append({"row": row_idx, "name": name, "error": "identifier_type is required when identifier is provided"})
                continue

            if identifier_type_value and identifier_value:
                # Normalize to lowercase so the file-level dedupe matches the
                # case-insensitive unique index (uq_external_inventory_items_identifier)
                # instead of letting case-colliding rows slip through to the
                # per-row binary-split fallback on the INSERT.
                identifier_pair = (identifier_type_value.lower(), identifier_value.lower())
                if identifier_pair in seen_identifiers:
                    errors.append({
                        "row": row_idx,
                        "name": name,
                        "error": "Duplicate identifier_type and identifier in file",
                    })
                    continue
                seen_identifiers.add(identifier_pair)

            quantity_raw = row_data.get("quantity", "").strip()
            try:
                quantity_value = int(quantity_raw) if quantity_raw else 1
            except ValueError:
                errors.append({"row": row_idx, "name": name, "error": "Invalid quantity value"})
                continue
            if quantity_value < 1:
                errors.append({"row": row_idx, "name": name, "error": "Quantity must be at least 1"})
                continue

            price_raw = row_data.get("price", "").strip()
            if price_raw:
                try:
                    price_value = float(price_raw)
                except ValueError:
                    errors.append({"row": row_idx, "name": name, "error": "Invalid price value"})
                    continue
            else:
                price_value = 0.0

            warranty_start_date_raw = row_data.get("warranty_start_date", "").strip()
            if warranty_start_date_raw:
                try:
                    warranty_start_date_value = datetime.strptime(warranty_start_date_raw, "%Y-%m-%d").date()
                except ValueError:
                    errors.append({
                        "row": row_idx,
                        "name": name,
                        "error": "Invalid warranty_start_date value (expected YYYY-MM-DD)",
                    })
                    continue
            else:
                warranty_start_date_value = None

            warranty_duration_raw = row_data.get("warranty_duration", "").strip()
            if warranty_duration_raw:
                try:
                    warranty_duration_value = int(warranty_duration_raw)
                except ValueError:
                    errors.append({"row": row_idx, "name": name, "error": "Invalid warranty_duration value"})
                    continue
                if warranty_duration_value < 0:
                    errors.append({"row": row_idx, "name": name, "error": "Warranty duration must be at least 0"})
                    continue
            else:
                warranty_duration_value = None

            prepared_rows.append({
                "row": row_idx,
                "name": name,
                "identifier_type": identifier_type_value or None,
                "identifier": identifier_value or None,
                "device_type": row_data.get("device_type") or None,
                "price": price_value,
                "quantity": quantity_value,
                "supplier_name": row_data.get("supplier_name") or None,
                "location": row_data.get("location") or None,
                "warranty_start_date": warranty_start_date_value,
                "warranty_duration": warranty_duration_value,
                "notes": row_data.get("notes") or None,
            })

        if prepared_rows:
            actor_id = _resolve_actor_id(current_user)
            now = datetime.now().replace(tzinfo=None)

            from app.database_sqlalchemy import async_session_factory
            from app.core.cache_version import bump_cache_version

            insert_sql = """INSERT INTO external_inventory_items (
                name, identifier_type, identifier, device_type, price, quantity,
                supplier_name, location, status, notes, warranty_start_date, warranty_duration,
                created_by, created_at, updated_at
            ) VALUES (:name, :identifier_type, :identifier, :device_type, :price, :quantity,
                :supplier_name, :location, 'active', :notes, :warranty_start_date, :warranty_duration,
                :created_by, :created_at, :updated_at)"""

            async with async_session_factory() as session:
                existing_pairs = await _fetch_existing_identifier_pairs(
                    session,
                    [
                        (item["identifier_type"].lower(), item["identifier"].lower())
                        for item in prepared_rows
                        if item["identifier_type"] and item["identifier"]
                    ],
                )

                insertable_rows = []
                for item in prepared_rows:
                    if item["identifier_type"] and item["identifier"]:
                        pair = (item["identifier_type"].lower(), item["identifier"].lower())
                        if pair in existing_pairs:
                            errors.append({
                                "row": item["row"],
                                "name": item["name"],
                                "error": "Identifier type and identifier already exist",
                            })
                            continue
                    insertable_rows.append(item)

                payload_rows = []
                for item in insertable_rows:
                    payload_rows.append({
                        "row": item["row"],
                        "name": item["name"],
                        "identifier_type": item["identifier_type"],
                        "identifier": item["identifier"],
                        "device_type": item["device_type"],
                        "price": item["price"],
                        "quantity": item["quantity"],
                        "supplier_name": item["supplier_name"],
                        "location": item["location"],
                        "warranty_start_date": item["warranty_start_date"],
                        "warranty_duration": item["warranty_duration"],
                        "notes": item["notes"],
                        "created_by": actor_id,
                        "created_at": now,
                        "updated_at": now,
                    })

                # Large uploads commit per batch (mirroring the device bulk
                # upload) so one huge transaction does not accumulate a giant
                # undo log / binlog entry; small uploads keep a single atomic
                # transaction.
                use_chunked_commits = len(payload_rows) > BULK_UPLOAD_CHUNKED_COMMIT_THRESHOLD

                async def _item_batch_success(session, batch):
                    created.extend(item["name"] for item in batch)

                async def _item_row_error(session, item, err):
                    errors.append({
                        "row": item["row"],
                        "name": item["name"],
                        "error": str(err),
                    })

                async def _item_batch_complete(ok, can_commit):
                    if use_chunked_commits and can_commit:
                        await bump_cache_version(session)
                        await session.commit()

                should_commit = await chunked_executemany(
                    session,
                    insert_sql,
                    payload_rows,
                    chunk_size=2000,
                    on_batch_success=_item_batch_success,
                    on_row_error=_item_row_error,
                    on_batch_complete=_item_batch_complete,
                    abort_on_error=False,
                )

                if use_chunked_commits:
                    # Prior batches are already committed; a hard error here
                    # only rolls back the in-flight batch.
                    if not should_commit:
                        await session.rollback()
                elif should_commit:
                    await bump_cache_version(session)
                    await session.commit()
                else:
                    await session.rollback()
                    raise HTTPException(
                        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        detail="Bulk upload was rolled back due to an unexpected insert error. Please retry."
                    )

        actor_name = current_user.get("name") or current_user.get("email") or "User"
        await log_business_activity(
            user=current_user,
            path="/activity/external-inventory/item-bulk-import",
            description=(
                f"{actor_name} imported external inventory items: "
                f"{len(created)} created, {len(skipped)} skipped, {len(errors)} errors"
            ),
        )

        return {
            "success": True,
            "message": f"Import complete: {len(created)} created, {len(skipped)} skipped, {len(errors)} errors",
            "data": build_bulk_result(created, skipped, errors),
        }
    except HTTPException:
        raise
    except Exception:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later.",
        )


@router.put("/items/{item_id}", summary="Update external inventory item")
async def update_external_inventory_item(
    item_id: int,
    item_data: InventoryItemUpdate,
    current_user: dict = Depends(require_management),
):
    try:
        before = await inventory_service.get_item_by_id(item_id)
        updated = await inventory_service.update_item(
            item_id=item_id,
            item_data=item_data,
            user=current_user,
        )
        if not updated:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="External inventory item not found",
            )

        actor_name = current_user.get("name") or current_user.get("email") or "User"
        edited_fields = list(item_data.model_dump(exclude_unset=True).keys())
        change_summary = build_field_change_summary(
            before=before or {},
            after=updated or {},
            fields=edited_fields,
            exclude_fields={"updated_at"},
        )
        await log_business_activity(
            user=current_user,
            path="/activity/external-inventory/item-update",
            description=(
                f"{actor_name} updated external inventory item "
                f"{(before or {}).get('name') or item_id}; changes: {change_summary}"
            ),
        )

        return {
            "success": True,
            "message": "External inventory item updated successfully",
            "data": updated,
        }
    except HTTPException:
        raise
    except Exception:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later.",
        )


@router.delete("/items/{item_id}", summary="Delete external inventory item")
async def delete_external_inventory_item(
    item_id: int,
    current_user: dict = Depends(require_management),
):
    try:
        deleted = await inventory_service.delete_item(item_id=item_id)
        if not deleted:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="External inventory item not found",
            )

        actor_name = current_user.get("name") or current_user.get("email") or "User"
        await log_business_activity(
            user=current_user,
            path="/activity/external-inventory/item-delete",
            description=f"{actor_name} deleted external inventory item {deleted.get('name') or item_id}",
        )

        return {
            "success": True,
            "message": "External inventory item deleted successfully",
        }
    except HTTPException:
        raise
    except Exception:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later.",
        )


@router.post("/distributions", status_code=status.HTTP_201_CREATED, summary="Distribute external inventory item")
async def distribute_external_inventory_item(
    payload: ExternalDistributionCreate,
    current_user: dict = Depends(require_management),
):
    """Distribute an external inventory item to a recipient.

    The distribution completes immediately (no recipient confirmation step),
    reduces the item's available quantity, records a history entry, and
    notifies the recipient.
    """
    try:
        result = await inventory_service.distribute_item(payload=payload, user=current_user)
        actor_name = current_user.get("name") or current_user.get("email") or "User"
        await log_business_activity(
            user=current_user,
            path="/activity/external-inventory/distribute",
            description=(
                f"{actor_name} distributed {result.get('quantity')} x "
                f"{result.get('item_name')} to {result.get('recipient_name')}"
            ),
        )
        return {
            "success": True,
            "message": "External inventory item distributed successfully",
            "data": result,
        }
    except HTTPException:
        raise
    except Exception:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later.",
        )


@router.post("/distributions/bulk", status_code=status.HTTP_201_CREATED, summary="Bulk distribute external inventory items")
async def bulk_distribute_external_inventory_items(
    payload: ExternalBulkDistributionCreate,
    current_user: dict = Depends(require_management),
):
    try:
        result = await inventory_service.bulk_distribute(payload=payload, user=current_user)
        actor_name = current_user.get("name") or current_user.get("email") or "User"
        await log_business_activity(
            user=current_user,
            path="/activity/external-inventory/distribute-bulk",
            description=(
                f"{actor_name} bulk distributed external inventory items: "
                f"{result.get('created_count')} created, {result.get('error_count')} errors"
            ),
        )
        return {
            "success": True,
            "message": (
                f"Bulk distribution complete: {result.get('created_count')} distributed, "
                f"{result.get('error_count')} errors"
            ),
            "data": result,
        }
    except HTTPException:
        raise
    except Exception:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later.",
        )


@router.post("/distributions/bulk-upload", status_code=status.HTTP_201_CREATED, summary="Bulk distribute external inventory items from CSV/Excel")
async def bulk_distribute_external_inventory_from_file(
    file: UploadFile = File(...),
    to_user_id: str = Form(...),
    notes: Optional[str] = Form(None),
    current_user: dict = Depends(require_management),
):
    """Bulk distribute external inventory items to a single recipient from an
    uploaded CSV/Excel file.

    Required columns: ``identifier_type``, ``identifier``
    Optional columns: ``quantity``
    """
    filename_lower = (file.filename or "").lower()
    if not filename_lower.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel (.xlsx, .xls) or CSV (.csv) files are supported",
        )

    try:
        contents = await file.read()

        from app.services.bulk_upload_service import (
            check_bulk_upload_file,
            check_bulk_upload_row_count,
            validate_upload_signature,
        )

        # Enforces the 10 MB size cap, the xlsx decompressed-size (zip bomb) guard,
        # and rejects mis-typed files via magic-byte signature checks.
        check_bulk_upload_file(contents, filename_lower)
        validate_upload_signature(filename_lower, contents)

        if filename_lower.endswith(".csv"):
            decoded = contents.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(decoded))
            all_rows = list(reader)
            if not all_rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="CSV file is empty",
                )
            headers = [h.strip().lower() for h in all_rows[0]]
            data_rows = all_rows[1:]
        else:
            headers, data_rows = _read_bulk_rows(filename_lower, contents)
        check_bulk_upload_row_count(data_rows)

        required = {"identifier_type", "identifier"}
        missing = required - set(headers)
        if missing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing required columns: {', '.join(sorted(missing))}",
            )

        identifier_rows = []
        for row_idx, row in enumerate(data_rows, start=2):
            padded = row + [""] * (len(headers) - len(row))
            row_data = {
                headers[i]: (str(padded[i]).strip() if padded[i] is not None else "")
                for i in range(len(headers))
            }

            if not any(row_data.values()):
                continue

            identifier_rows.append({
                "row": row_idx,
                "identifier_type": row_data.get("identifier_type", "").strip(),
                "identifier": row_data.get("identifier", "").strip(),
                "quantity": row_data.get("quantity", "").strip(),
                "notes": notes,
            })

        result = await inventory_service.bulk_distribute_from_file(
            identifier_rows=identifier_rows,
            to_user_id=to_user_id,
            user=current_user,
        )
        actor_name = current_user.get("name") or current_user.get("email") or "User"
        await log_business_activity(
            user=current_user,
            path="/activity/external-inventory/distribute-bulk-upload",
            description=(
                f"{actor_name} bulk distributed external inventory items from file: "
                f"{result.get('created_count')} created, {result.get('error_count')} errors"
            ),
        )

        return {
            "success": True,
            "message": (
                f"Bulk distribution complete: {result.get('created_count')} distributed, "
                f"{result.get('error_count')} errors"
            ),
            "data": result,
        }
    except HTTPException:
        raise
    except Exception:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later.",
        )


async def _fetch_existing_identifier_pairs(session, pairs: list[tuple[str, str]]) -> set:
    """Return the identifier pairs already present in the catalog.

    Runs one row-value ``IN`` query per batch (``(a, b) IN ((...), ...)``) so
    MySQL can use the composite unique index
    ``uq_external_inventory_items_identifier`` directly, instead of a slow
    multi-hundred-way ``OR`` that degrades as the table grows. Both sides are
    normalized to lowercase so case collisions with the case-insensitive
    collation are caught here rather than triggering the per-row binary-split
    fallback on the INSERT.
    """
    if not pairs:
        return set()

    existing = set()
    for batch in chunks(pairs, 1000):
        row_constructors = ", ".join(f"(:t_{i}, :i_{i})" for i in range(len(batch)))
        params = {}
        for i, (identifier_type, identifier) in enumerate(batch):
            params[f"t_{i}"] = identifier_type
            params[f"i_{i}"] = identifier
        result = await session.execute(
            text(
                "SELECT identifier_type, identifier FROM external_inventory_items "
                f"WHERE (identifier_type, identifier) IN ({row_constructors})"
            ),
            params,
        )
        for row in result.mappings().all():
            existing.add(
                (
                    str(row.get("identifier_type") or "").strip().lower(),
                    str(row.get("identifier") or "").strip().lower(),
                )
            )
    return existing


@router.get("/distributions", summary="Get external inventory distribution history")
async def get_external_inventory_distributions(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1),
    search: Optional[str] = None,
    search_by: Optional[str] = Query("all"),
    item_id: Optional[int] = None,
    identifier_type: Optional[str] = Query(None),
    device_type: Optional[str] = Query(None, alias="type"),
    warranty: Optional[str] = Query(None),
    current_user: dict = Depends(require_management),
):
    """Management-only reporting page listing every completed distribution."""
    try:
        result = await inventory_service.get_distribution_history(
            page=page,
            page_size=page_size,
            search=search,
            search_by=search_by,
            item_id=item_id,
            identifier_type=identifier_type,
            device_type=device_type,
            warranty=warranty,
        )
        return {
            "success": True,
            "message": "External inventory distributions retrieved successfully",
            "data": result["data"],
            "pagination": result["pagination"],
        }
    except HTTPException:
        raise
    except Exception:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later.",
        )

