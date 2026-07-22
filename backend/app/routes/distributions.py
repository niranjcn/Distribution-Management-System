import logging
import csv
import io

from fastapi import APIRouter, HTTPException, status, Depends, Query, UploadFile, File, Form
from fastapi.responses import FileResponse, Response
from typing import Optional
from pydantic import BaseModel
from datetime import date
from app.models.distribution import DistributionCreate, DistributionStatusUpdate
from app.services import distribution_service
from app.middleware.auth_middleware import get_current_user, require_admin_or_manager, require_management
from app.core.activity_logger import build_field_change_summary, log_business_activity

router = APIRouter()

logger = logging.getLogger(__name__)


def _ensure_not_md_director(current_user: dict) -> None:
    if current_user.get("role") == "md_director":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="MD/Director has read-only access to distributions"
        )


def _ensure_distribution_create_access(current_user: dict) -> None:
    if current_user.get("role") == "sub_distribution_manager":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Sub Distribution MD/Manager cannot create or bulk upload distributions"
        )


def _ensure_sub_distribution_manager_read_only(current_user: dict) -> None:
    if current_user.get("role") == "sub_distribution_manager":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Sub Distribution MD/Manager has read-only access to distributions",
        )


def _is_likely_text(content: bytes) -> bool:
    if not content:
        return True
    return b"\x00" not in content


def _validate_upload_signature(filename_lower: str, content: bytes) -> None:
    if filename_lower.endswith(".xlsx"):
        if not content.startswith(b"PK\x03\x04"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid XLSX file content"
            )
        return

    if filename_lower.endswith(".xls"):
        if not content.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid XLS file content"
            )
        return

    if filename_lower.endswith(".csv"):
        if not _is_likely_text(content[:2048]):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid CSV file content"
            )
        return


class ReceiptConfirmation(BaseModel):
    received: bool
    notes: Optional[str] = None


class ReturnConfirmation(BaseModel):
    notes: Optional[str] = None


@router.post("/bulk-upload")
async def bulk_upload_distribution(
    file: UploadFile = File(...),
    to_user_id: str = Form(...),
    notes: Optional[str] = Form(None),
    date_of_distribution: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Create a distribution from uploaded CSV/Excel rows using mac_address, serial_number, and/or nuid."""
    filename_lower = (file.filename or "").lower()
    _ensure_not_md_director(current_user)
    _ensure_distribution_create_access(current_user)

    if not filename_lower.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel (.xlsx, .xls) or CSV (.csv) files are supported"
        )

    try:
        contents = await file.read()

        _validate_upload_signature(filename_lower, contents)

        if filename_lower.endswith(".csv"):
            decoded = contents.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(decoded))
            all_rows = list(reader)
            if not all_rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="CSV file is empty"
                )
            headers = [str(h).strip().lower() for h in all_rows[0]]
            data_rows = all_rows[1:]

            def iter_data_rows():
                for row in data_rows:
                    padded = row + [""] * (len(headers) - len(row))
                    yield tuple(padded[:len(headers)])

        else:
            import openpyxl

            workbook = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            worksheet = workbook.active
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1), None)
            if not header_row:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Excel file is empty"
                )

            headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in header_row]

            def iter_data_rows():
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    yield row

        if "mac_address" not in headers and "serial_number" not in headers and "nuid" not in headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Missing required columns: add at least one of mac_address, serial_number, or nuid"
            )

        identifier_rows = []
        date_values = set()
        for row_idx, row in enumerate(iter_data_rows(), start=2):
            row_data = {
                headers[i]: (str(row[i]).strip() if i < len(row) and row[i] is not None else "")
                for i in range(len(headers))
            }

            mac_address = row_data.get("mac_address", "")
            serial_number = row_data.get("serial_number", "")
            nuid = row_data.get("nuid", "")
            row_date = row_data.get("date_of_distribution", "")

            if row_date:
                date_values.add(row_date)

            if not mac_address and not serial_number and not nuid:
                # Skip fully empty lines, otherwise keep for validation.
                if not any(v for v in row_data.values()):
                    continue

            identifier_rows.append({
                "row": row_idx,
                "mac_address": mac_address,
                "serial_number": serial_number,
                "nuid": nuid,
            })

        if not identifier_rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No identifier rows found in file"
            )

        def _parse_distribution_date(value: str) -> date:
            try:
                return date.fromisoformat(value.strip())
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="date_of_distribution must be in YYYY-MM-DD format",
                )

        normalized_form_date = _parse_distribution_date(date_of_distribution) if date_of_distribution else None
        normalized_file_date = None

        if date_values:
            normalized_dates = {_parse_distribution_date(value).isoformat() for value in date_values if value}
            if len(normalized_dates) > 1:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Multiple date_of_distribution values found in file; provide a single date",
                )
            normalized_file_date = _parse_distribution_date(next(iter(normalized_dates)))

        if normalized_form_date and normalized_file_date and normalized_form_date != normalized_file_date:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="date_of_distribution in form and file do not match",
            )

        final_distribution_date = normalized_file_date or normalized_form_date

        result = await distribution_service.create_distribution_from_identifiers(
            to_user_id=to_user_id,
            identifier_rows=identifier_rows,
            from_user=current_user,
            notes=notes,
            date_of_distribution=final_distribution_date,
        )

        if result.get("created") and result.get("distribution"):
            distribution = result["distribution"]
            actor_name = current_user.get("name") or current_user.get("email") or "User"
            await log_business_activity(
                user=current_user,
                path="/activity/distributions/bulk-create",
                description=(
                    f"{actor_name} created distribution {distribution.get('distribution_id')} "
                    f"to {distribution.get('to_user_name') or distribution.get('to_user_id')} "
                    f"via bulk upload ({result.get('created_count', 0)} device(s))"
                ),
            )

        if result["created"]:
            if result.get("error_count"):
                message = (
                    f"Distribution created with {result['created_count']} device(s); "
                    f"{result['error_count']} row error(s)."
                )
            else:
                message = (
                    f"Distribution created successfully with {result['created_count']} device(s)"
                )
        else:
            message = (
                f"Upload validation failed: {result['error_count']} row error(s). "
                "Fix errors and upload again."
            )

        return {
            "success": True,
            "message": message,
            "data": result,
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.post("/sync-devices")
async def sync_distribution_devices(
    current_user: dict = Depends(require_admin_or_manager)
):
    """Sync device holders for all approved distributions (admin fix endpoint)"""
    try:
        result = await distribution_service.sync_approved_distributions(user=current_user)
        return {
            "success": True,
            "message": f"Synced {result['devices_synced']} device(s) from {result['total_distributions']} approved distribution(s)",
            "data": result
        }
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later.")


@router.get("")
async def get_distributions(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1),
    status: Optional[str] = None,
    to_user_id: Optional[str] = None,
    search: Optional[str] = None,
    search_by: Optional[str] = Query("all"),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all distributions with pagination and filters"""
    try:
        result = await distribution_service.get_distributions(
            page=page,
            page_size=page_size,
            status=status,
            to_user_id=to_user_id,
            current_user=current_user,
            search=search,
            search_by=search_by,
            start_date=start_date,
            end_date=end_date,
        )

        return {
            "success": True,
            "message": "Distributions retrieved successfully",
            "data": result["data"],
            "pagination": result["pagination"]
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.get("/pending")
async def get_pending_distributions(
    current_user: dict = Depends(require_management)
):
    """Get pending distributions for approval"""
    try:
        distributions = await distribution_service.get_pending_distributions()

        return {
            "success": True,
            "message": "Pending distributions retrieved successfully",
            "data": distributions
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.get("/{distribution_id}/manifest")
async def download_distribution_manifest(
    distribution_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Download generated Excel manifest for a distribution."""
    try:
        manifest = await distribution_service.get_distribution_manifest_file(distribution_id, current_user)
        if not manifest:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Distribution manifest not found"
            )

        return FileResponse(
            path=manifest["path"],
            filename=manifest["filename"],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.get("/{distribution_id}/export-mac-nuid")
async def download_distribution_mac_nuid(
    distribution_id: str,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    current_user: dict = Depends(get_current_user)
):
    """Download distribution devices as MAC/NUID export in CSV or XLSX format."""
    try:
        export_data = await distribution_service.get_distribution_mac_nuid_export(
            distribution_id=distribution_id,
            user=current_user,
            file_format=format,
        )

        return Response(
            content=export_data["content"],
            media_type=export_data["media_type"],
            headers={
                "Content-Disposition": f"attachment; filename={export_data['filename']}"
            },
        )
    except ValueError as e:
        message = str(e)
        if "not found" in message.lower():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=message
            )
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=message
        )
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.get("/{distribution_id}")
async def get_distribution(
    distribution_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get distribution by ID"""
    try:
        distribution = await distribution_service.get_distribution_by_id(distribution_id)

        if not distribution:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Distribution not found"
            )

        return {
            "success": True,
            "message": "Distribution retrieved successfully",
            "data": distribution
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.post("", status_code=status.HTTP_201_CREATED)
async def create_distribution(
    dist_data: DistributionCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new distribution request.
    - admin/manager/staff: can distribute PDIC devices to any sub-level user
    - sub_distributor: can distribute their held devices to clusters or operators under them
    - cluster: can distribute their held devices to operators under them
    - operator: can distribute their held devices to operators in the same cluster
    """
    
    _ensure_not_md_director(current_user)
    _ensure_distribution_create_access(current_user)

    try:
        distribution = await distribution_service.create_distribution(
            dist_data=dist_data,
            from_user=current_user
        )

        if not distribution:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Distribution created but could not be loaded. Please try again."
            )

        actor_name = current_user.get("name") or current_user.get("email") or "User"
        await log_business_activity(
            user=current_user,
            path="/activity/distributions/create",
            description=(
                f"{actor_name} created distribution {distribution.get('distribution_id')} "
                f"to {distribution.get('to_user_name') or distribution.get('to_user_id')} "
                f"({distribution.get('device_count', 0)} device(s))"
            ),
        )

        return {
            "success": True,
            "message": "Distribution created successfully",
            "data": distribution
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.post("/{distribution_id}/receipt")
async def confirm_distribution_receipt(
    distribution_id: str,
    body: ReceiptConfirmation,
    current_user: dict = Depends(get_current_user)
):
    """Recipient confirms or disputes receipt of a distribution.
    - received=true  -> Distribution becomes APPROVED; receiver can now redistribute devices
    - received=false -> Distribution becomes DISPUTED; admin/manager + sender are notified
    """
    _ensure_not_md_director(current_user)
    _ensure_sub_distribution_manager_read_only(current_user)

    try:
        distribution = await distribution_service.confirm_receipt(
            distribution_id=distribution_id,
            received=body.received,
            user=current_user,
            notes=body.notes
        )
        action = "confirmed" if body.received else "disputed"
        return {
            "success": True,
            "message": f"Receipt {action} successfully",
            "data": distribution
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.post("/{distribution_id}/confirm-return")
async def confirm_disputed_distribution_return(
    distribution_id: str,
    body: ReturnConfirmation,
    current_user: dict = Depends(require_management)
):
    """PDIC confirms disputed devices are physically back with sender and unlocks redistribution."""
    _ensure_not_md_director(current_user)

    try:
        distribution = await distribution_service.confirm_disputed_return(
            distribution_id=distribution_id,
            user=current_user,
            notes=body.notes,
        )
        return {
            "success": True,
            "message": "Disputed return confirmed successfully",
            "data": distribution,
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.patch("/{distribution_id}/status")
async def update_distribution_status(
    distribution_id: str,
    status_update: DistributionStatusUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update distribution status"""
    _ensure_not_md_director(current_user)
    _ensure_sub_distribution_manager_read_only(current_user)

    try:
        before = await distribution_service.get_distribution_by_id(distribution_id)
        distribution = await distribution_service.update_distribution_status(
            distribution_id=distribution_id,
            status=status_update.status.value,
            user=current_user,
            notes=status_update.notes
        )

        if not distribution:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Distribution not found"
            )

        actor_name = current_user.get("name") or current_user.get("email") or "User"
        change_summary = build_field_change_summary(
            before=before or {},
            after=distribution or {},
            fields=["status"],
            exclude_fields={"updated_at"},
        )
        await log_business_activity(
            user=current_user,
            path="/activity/distributions/status-update",
            description=(
                f"{actor_name} updated distribution status for "
                f"{distribution.get('distribution_id') or (before or {}).get('distribution_id') or distribution_id}; "
                f"changes: {change_summary}"
            ),
        )

        return {
            "success": True,
            "message": "Distribution status updated successfully",
            "data": distribution
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except PermissionError as e:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )


@router.delete("/{distribution_id}")
async def cancel_distribution(
    distribution_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Cancel a distribution (only by creator)"""
    _ensure_not_md_director(current_user)
    _ensure_sub_distribution_manager_read_only(current_user)

    try:
        success = await distribution_service.cancel_distribution(
            distribution_id=distribution_id,
            user=current_user
        )

        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Distribution not found"
            )

        return {
            "success": True,
            "message": "Distribution cancelled successfully"
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.exception("Unhandled route exception")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An internal error occurred. Please try again later."
        )





